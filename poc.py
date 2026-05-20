"""
╔══════════════════════════════════════════════════════════════════════════════════╗
║         ENTERPRISE VEHICLE COUNTER — SINGLE-FILE EDITION                        ║
║         Colombia: ANI (CAT I–VII) · INVIAS (CAT I–V) · COL (C1–C5)             ║
║         YOLOv8 · ByteTrack · Kalman Filter · White-Line Lane Detection          ║
║         Speed Estimation · Density Analytics · HUD Engine                       ║
║         Reports: CSV · JSON · Excel (5 sheets) · PDF · HTML                     ║
╚══════════════════════════════════════════════════════════════════════════════════╝

INSTALL:
    pip install ultralytics opencv-python numpy openpyxl reportlab

RUN (basic):
    python vehicle_counter_colombia_final.py --video footage.mp4

RUN (full options):
    python vehicle_counter_colombia_final.py \\
        --video footage.mp4 \\
        --output annotated.mp4 \\
        --lanes 3 \\
        --direction NS \\
        --classification ANI \\
        --report my_report \\
        --camera-id CAM1 \\
        --model yolov8s.pt \\
        --site "Peaje Niquía" \\
        --no-preview

CLASSIFICATION OPTIONS:
    --classification ANI      →  CAT I to CAT VII  (default)
    --classification INVIAS   →  CAT I to CAT V
    --classification COL      →  C1 to C5

DIRECTION OPTIONS:
    --direction NS   →  North→South (toward Medellín)
    --direction SN   →  South→North (toward Copacabana)

BATCH PROCESSING:
    python vehicle_counter_colombia_final.py --batch videos.txt
    # Each line: video_path,lanes,direction,camera_id
"""

from __future__ import annotations

# ── stdlib ─────────────────────────────────────────────────────────────────────
import argparse
import csv
import json
import math
import os
import statistics
import sys
import time
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ── third-party ────────────────────────────────────────────────────────────────
import cv2
import numpy as np

try:
    from ultralytics import YOLO
except ImportError:
    raise SystemExit("ERROR: Run  pip install ultralytics opencv-python numpy")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAVE_EXCEL = True
except ImportError:
    HAVE_EXCEL = False

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, PageBreak, KeepTogether,
    )
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    HAVE_PDF = True
except ImportError:
    HAVE_PDF = False


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — CLASSIFICATION SYSTEMS
# ══════════════════════════════════════════════════════════════════════════════

YOLO_VEHICLE_CLASSES: Dict[int, str] = {
    2: "car",
    3: "motorcycle",
    5: "bus",
    7: "truck",
}

# ── ANI (Agencia Nacional de Infraestructura) CAT I–VII ──────────────────────
ANI_CATEGORIES: Dict[str, dict] = {
    "CAT I": {
        "label": "CAT I", "name": "Motocicletas y bicicletas",
        "desc":  "Vehículos de dos ruedas motorizados o no motorizados",
        "yolo":  ["motorcycle"], "axles": 2,
        "color": (180, 120, 255), "hex": "#B478FF", "icon": "🏍️",
    },
    "CAT II": {
        "label": "CAT II", "name": "Automóviles, camperos, camionetas y microbuses",
        "desc":  "Vehículos livianos de pasajeros hasta 8 asientos",
        "yolo":  ["car"], "axles": 2,
        "color": (0, 220, 80), "hex": "#00DC50", "icon": "🚗",
    },
    "CAT III": {
        "label": "CAT III", "name": "Buses",
        "desc":  "Vehículos de transporte masivo de pasajeros",
        "yolo":  ["bus"], "axles": 2,
        "color": (0, 160, 255), "hex": "#00A0FF", "icon": "🚌",
    },
    "CAT IV": {
        "label": "CAT IV", "name": "Camiones de 2 ejes",
        "desc":  "Vehículos de carga con dos ejes simples",
        "yolo":  ["truck"], "axles": 2,
        "color": (255, 200, 0), "hex": "#FFC800", "icon": "🚚",
    },
    "CAT V": {
        "label": "CAT V", "name": "Camiones de 3 ejes",
        "desc":  "Vehículos de carga con tres ejes",
        "yolo":  ["truck"], "axles": 3,
        "color": (255, 120, 0), "hex": "#FF7800", "icon": "🚛",
    },
    "CAT VI": {
        "label": "CAT VI", "name": "Camiones de 4 ejes",
        "desc":  "Vehículos de carga con cuatro ejes",
        "yolo":  ["truck"], "axles": 4,
        "color": (220, 50, 50), "hex": "#DC3232", "icon": "🚛",
    },
    "CAT VII": {
        "label": "CAT VII", "name": "Camiones de 5 o más ejes",
        "desc":  "Tracto-camiones y vehículos articulados pesados",
        "yolo":  ["truck"], "axles": 5,
        "color": (140, 0, 0), "hex": "#8C0000", "icon": "🚛",
    },
}

# ── INVIAS (Instituto Nacional de Vías) CAT I–V ───────────────────────────────
INVIAS_CATEGORIES: Dict[str, dict] = {
    "CAT I":  {
        "label": "CAT I",  "name": "Motos",
        "desc": "Motocicletas y ciclomotores",
        "yolo": ["motorcycle"], "color": (180, 120, 255), "hex": "#B478FF", "icon": "🏍️",
    },
    "CAT II": {
        "label": "CAT II", "name": "Automóviles y camperos",
        "desc": "Vehículos livianos de hasta 8 pasajeros",
        "yolo": ["car"], "color": (0, 220, 80), "hex": "#00DC50", "icon": "🚗",
    },
    "CAT III":{
        "label": "CAT III","name": "Buses y busetas",
        "desc": "Vehículos de transporte público colectivo",
        "yolo": ["bus"], "color": (0, 160, 255), "hex": "#00A0FF", "icon": "🚌",
    },
    "CAT IV": {
        "label": "CAT IV", "name": "Camiones livianos (2 ejes)",
        "desc": "Camiones de dos ejes con peso bruto ≤ 10.5 t",
        "yolo": ["truck"], "color": (255, 200, 0), "hex": "#FFC800", "icon": "🚚",
    },
    "CAT V":  {
        "label": "CAT V",  "name": "Camiones pesados (3+ ejes)",
        "desc": "Vehículos de carga con tres o más ejes",
        "yolo": ["truck"], "color": (220, 50, 50), "hex": "#DC3232", "icon": "🚛",
    },
}

# ── Colombian Standard C1–C5 ──────────────────────────────────────────────────
COL_CATEGORIES: Dict[str, dict] = {
    "C1": {
        "label": "C1",  "name": "Automóviles, microbuses y motos",
        "desc": "Vehículos livianos de pasajeros y motos",
        "yolo": ["car", "motorcycle"], "color": (0, 220, 80), "hex": "#00DC50", "icon": "🚗",
    },
    "C2P":{
        "label": "C2P", "name": "Camión 2 ejes pequeño",
        "desc": "Camión ligero de dos ejes",
        "yolo": ["truck"], "color": (255, 200, 0), "hex": "#FFC800", "icon": "🚚",
    },
    "C2G":{
        "label": "C2G", "name": "Camión 2 ejes grande / bus",
        "desc": "Camión pesado de dos ejes o bus",
        "yolo": ["bus", "truck"], "color": (0, 160, 255), "hex": "#00A0FF", "icon": "🚌",
    },
    "C3": {
        "label": "C3",  "name": "Camión 3 ejes",
        "desc": "Vehículo articulado de tres ejes",
        "yolo": ["truck"], "color": (255, 120, 0), "hex": "#FF7800", "icon": "🚛",
    },
    "C4": {
        "label": "C4",  "name": "Camión 4 ejes",
        "desc": "Vehículo articulado de cuatro ejes",
        "yolo": ["truck"], "color": (220, 50, 50), "hex": "#DC3232", "icon": "🚛",
    },
    "C5": {
        "label": "C5",  "name": "Camión 5+ ejes",
        "desc": "Tracto-camión con más de cuatro ejes",
        "yolo": ["truck"], "color": (140, 0, 0), "hex": "#8C0000", "icon": "🚛",
    },
}

CLASSIFICATION_SYSTEMS: Dict[str, Dict[str, dict]] = {
    "ANI":    ANI_CATEGORIES,
    "INVIAS": INVIAS_CATEGORIES,
    "COL":    COL_CATEGORIES,
}


def classify_vehicle(yolo_type: str, box_w: int, box_h: int,
                     system: str = "ANI", conf: float = 1.0) -> str:
    """
    Map a YOLO detection to a classification category key.
    Uses bounding-box aspect ratio + area heuristics to distinguish truck axle counts.
    """
    area   = box_w * box_h
    aspect = box_w / max(box_h, 1)

    if system == "ANI":
        if yolo_type == "motorcycle": return "CAT I"
        if yolo_type == "car":        return "CAT II"
        if yolo_type == "bus":        return "CAT III"
        if yolo_type == "truck":
            if aspect > 3.5 or area > 90_000: return "CAT VII"
            if aspect > 2.8 or area > 60_000: return "CAT VI"
            if aspect > 2.2 or area > 38_000: return "CAT V"
            return "CAT IV"

    elif system == "INVIAS":
        if yolo_type == "motorcycle": return "CAT I"
        if yolo_type == "car":        return "CAT II"
        if yolo_type == "bus":        return "CAT III"
        if yolo_type == "truck":
            if aspect > 2.5 or area > 45_000: return "CAT V"
            return "CAT IV"

    elif system == "COL":
        if yolo_type in ("car", "motorcycle"): return "C1"
        if yolo_type == "bus":                 return "C2G"
        if yolo_type == "truck":
            if aspect > 3.5 or area > 90_000: return "C5"
            if aspect > 2.8 or area > 60_000: return "C4"
            if aspect > 2.2 or area > 38_000: return "C3"
            if aspect > 1.5:                   return "C2G"
            return "C2P"

    cats = CLASSIFICATION_SYSTEMS.get(system, {})
    for key, cat in cats.items():
        if yolo_type in cat.get("yolo", []):
            return key
    return list(cats.keys())[0] if cats else "CAT II"


def cat_name(cat: str, system: str) -> str:
    return CLASSIFICATION_SYSTEMS.get(system, {}).get(cat, {}).get("name", cat)


def cat_color_bgr(cat: str, system: str) -> Tuple[int, int, int]:
    color = CLASSIFICATION_SYSTEMS.get(system, {}).get(cat, {}).get("color", (200, 200, 200))
    return (int(color[2]), int(color[1]), int(color[0]))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — GEOMETRY HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def line_intersection(p1, p2, p3, p4) -> Optional[Tuple[float, float]]:
    x1,y1=p1; x2,y2=p2; x3,y3=p3; x4,y4=p4
    denom=(x1-x2)*(y3-y4)-(y1-y2)*(x3-x4)
    if abs(denom)<1e-9: return None
    t=((x1-x3)*(y3-y4)-(y1-y3)*(x3-x4))/denom
    return x1+t*(x2-x1), y1+t*(y2-y1)


def poly_contains_point(poly, px: int, py: int) -> bool:
    n=len(poly); inside=False; j=n-1
    for i in range(n):
        xi,yi=poly[i]; xj,yj=poly[j]
        if ((yi>py)!=(yj>py)) and (px<(xj-xi)*(py-yi)/(yj-yi+1e-9)+xi):
            inside=not inside
        j=i
    return inside


def euclidean(p1, p2) -> float:
    return float(np.sqrt((p1[0]-p2[0])**2+(p1[1]-p2[1])**2))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — SPEED ESTIMATOR
# ══════════════════════════════════════════════════════════════════════════════

class SpeedEstimator:
    """
    Estimates vehicle speed from Kalman velocity (pixels/frame).

    Parameters
    ----------
    fps          : video frames per second
    px_per_meter : how many pixels = 1 metre (calibrate per camera)
    smoothing    : rolling window size for stable readout
    """

    def __init__(self, fps: float, px_per_meter: float, smoothing: int = 8):
        self.fps          = max(fps, 1.0)
        self.px_per_meter = max(px_per_meter, 0.1)
        self.smoothing    = smoothing
        self._raw: Dict[int, deque] = {}

    def update(self, track_id: int, vx_px: float, vy_px: float):
        if track_id not in self._raw:
            self._raw[track_id] = deque(maxlen=self.smoothing)
        self._raw[track_id].append(float(np.sqrt(vx_px**2 + vy_px**2)))

    def get_kmh(self, track_id: int) -> float:
        if track_id not in self._raw or not self._raw[track_id]:
            return 0.0
        avg_px  = float(np.mean(self._raw[track_id]))
        m_frame = avg_px / self.px_per_meter
        return round(m_frame * self.fps * 3.6, 1)

    def remove(self, track_id: int):
        self._raw.pop(track_id, None)

    def reset(self):
        self._raw.clear()


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — DENSITY ANALYZER
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class LaneDensityState:
    lane_n:        int
    flow_window:   deque = field(default_factory=lambda: deque(maxlen=3600))
    active_ids:    set   = field(default_factory=set)
    total_counted: int   = 0
    peak_flow:     float = 0.0
    last_count_ts: Optional[float] = None
    headway_buf:   deque = field(default_factory=lambda: deque(maxlen=50))

    def record_crossing(self, ts: float):
        if self.last_count_ts is not None:
            hw = ts - self.last_count_ts
            if 0.3 < hw < 120:
                self.headway_buf.append(hw)
        self.last_count_ts = ts
        self.flow_window.append(ts)
        self.total_counted += 1

    @property
    def avg_headway(self) -> float:
        return float(np.mean(self.headway_buf)) if self.headway_buf else 0.0

    def flow_rate_vph(self, window_s: float = 60.0) -> float:
        now = time.time()
        recent = sum(1 for t in self.flow_window if t >= now - window_s)
        return recent * (3600 / window_s)

    def update_peak(self, window_s: float = 60.0):
        r = self.flow_rate_vph(window_s)
        if r > self.peak_flow:
            self.peak_flow = r


class DensityAnalyzer:
    """Aggregates lane-level density, flow and congestion metrics."""

    def __init__(self, n_lanes: int, fps: float = 25.0):
        self.n_lanes = n_lanes
        self.fps     = fps
        self.lanes: Dict[int, LaneDensityState] = {
            i: LaneDensityState(lane_n=i) for i in range(1, n_lanes + 1)
        }
        self.session_start = time.time()

    def record_count(self, lane_n: int):
        if lane_n in self.lanes:
            self.lanes[lane_n].record_crossing(time.time())
            self.lanes[lane_n].update_peak()

    def update_active(self, lane_n: int, track_ids: set):
        if lane_n in self.lanes:
            self.lanes[lane_n].active_ids = track_ids

    def occupancy(self, lane_n: int) -> float:
        if lane_n not in self.lanes: return 0.0
        return min(len(self.lanes[lane_n].active_ids) / 10.0, 1.0)

    def congestion_level(self, lane_n: int) -> str:
        occ = self.occupancy(lane_n)
        if occ < 0.25: return "LIBRE"
        if occ < 0.55: return "MODERADO"
        if occ < 0.80: return "CONGESTIONADO"
        return "SATURADO"

    def congestion_color(self, lane_n: int) -> tuple:
        return {
            "LIBRE":         (0, 200, 80),
            "MODERADO":      (0, 200, 200),
            "CONGESTIONADO": (0, 120, 255),
            "SATURADO":      (0, 0, 220),
        }.get(self.congestion_level(lane_n), (200, 200, 200))

    def get_stats(self, lane_n: int) -> dict:
        if lane_n not in self.lanes: return {}
        s = self.lanes[lane_n]
        return {
            "lane":          lane_n,
            "flow_vph_1m":   round(s.flow_rate_vph(60),  1),
            "flow_vph_5m":   round(s.flow_rate_vph(300), 1),
            "peak_vph":      round(s.peak_flow,  1),
            "occupancy":     round(self.occupancy(lane_n), 3),
            "congestion":    self.congestion_level(lane_n),
            "avg_headway_s": round(s.avg_headway, 2),
            "total":         s.total_counted,
        }

    def get_all_stats(self) -> Dict[int, dict]:
        return {ln: self.get_stats(ln) for ln in self.lanes}

    def session_duration_s(self) -> float:
        return time.time() - self.session_start


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 5 — KALMAN BOX TRACKER
# ══════════════════════════════════════════════════════════════════════════════

class KalmanBoxTracker:
    """Kalman filter for smooth 2D bounding-box tracking (x, y, s, r, dx, dy, ds)."""

    count = 0

    def __init__(self, bbox: np.ndarray):
        KalmanBoxTracker.count += 1
        self.id  = KalmanBoxTracker.count
        self.kf  = cv2.KalmanFilter(7, 4)
        self.kf.measurementMatrix   = np.eye(4, 7, dtype=np.float32)
        self.kf.transitionMatrix    = np.eye(7, dtype=np.float32)
        for i in range(4):
            self.kf.transitionMatrix[i, i + 3] = 1.0
        self.kf.processNoiseCov     = np.eye(7, dtype=np.float32) * 1e-2
        self.kf.measurementNoiseCov = np.eye(4, dtype=np.float32) * 1e-1
        self.kf.errorCovPost        = np.eye(7, dtype=np.float32)
        self.kf.statePost           = self._bbox_to_state(bbox)
        self.time_since_update = 0
        self.hits       = 1
        self.hit_streak = 1
        self.history: List[np.ndarray] = []

    @staticmethod
    def _bbox_to_state(bbox: np.ndarray) -> np.ndarray:
        """Convert bbox to full 7x1 state vector for initialisation."""
        w=bbox[2]-bbox[0]; h=bbox[3]-bbox[1]
        x=bbox[0]+w/2;     y=bbox[1]+h/2
        s=w*h;             r=w/float(max(h,1))
        return np.array([[x],[y],[s],[r],[0.],[0.],[0.]], dtype=np.float32)

    @staticmethod
    def _bbox_to_z(bbox: np.ndarray) -> np.ndarray:
        """Convert bbox to 4x1 measurement vector for kf.correct()."""
        w=bbox[2]-bbox[0]; h=bbox[3]-bbox[1]
        x=bbox[0]+w/2;     y=bbox[1]+h/2
        s=w*h;             r=w/float(max(h,1))
        return np.array([[x],[y],[s],[r]], dtype=np.float32)

    @staticmethod
    def _z_to_bbox(state: np.ndarray) -> np.ndarray:
        w=math.sqrt(abs(state[2,0]*state[3,0])); h=abs(state[2,0])/max(w,1)
        return np.array([state[0,0]-w/2, state[1,0]-h/2,
                         state[0,0]+w/2, state[1,0]+h/2], dtype=np.float32)

    def predict(self) -> np.ndarray:
        if self.time_since_update > 0:
            self.hit_streak = 0
        self.time_since_update += 1
        self.kf.predict()
        return self._z_to_bbox(self.kf.statePost)

    def update(self, bbox: np.ndarray):
        self.time_since_update = 0
        self.hits      += 1
        self.hit_streak += 1
        self.kf.correct(self._bbox_to_z(bbox))   # 4x1 measurement ✓
        self.history.append(self.get_state())

    def get_state(self) -> np.ndarray:
        return self._z_to_bbox(self.kf.statePost)

    @property
    def velocity(self) -> Tuple[float, float]:
        st = self.kf.statePost
        return float(st[4,0]), float(st[5,0])


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 6 — BYTE TRACKER
# ══════════════════════════════════════════════════════════════════════════════

def _iou_matrix(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    area_a=(a[:,2]-a[:,0])*(a[:,3]-a[:,1])
    area_b=(b[:,2]-b[:,0])*(b[:,3]-b[:,1])
    ix1=np.maximum(a[:,None,0],b[None,:,0]); iy1=np.maximum(a[:,None,1],b[None,:,1])
    ix2=np.minimum(a[:,None,2],b[None,:,2]); iy2=np.minimum(a[:,None,3],b[None,:,3])
    inter=np.maximum(0,ix2-ix1)*np.maximum(0,iy2-iy1)
    union=area_a[:,None]+area_b[None,:]-inter
    return np.where(union>0, inter/union, 0.0)


def _hungarian(cost: np.ndarray, threshold: float
               ) -> Tuple[List[Tuple[int,int]], List[int], List[int]]:
    matched, unmatched_a, assigned_b = [], [], set()
    if cost.size == 0:
        return matched, list(range(cost.shape[0])), list(range(cost.shape[1]))
    for i in np.argsort(cost.min(axis=1)):
        for j in np.argsort(cost[i]):
            if j not in assigned_b and cost[i,j] < threshold:
                matched.append((int(i), int(j))); assigned_b.add(j); break
        else:
            unmatched_a.append(int(i))
    unmatched_b=[j for j in range(cost.shape[1]) if j not in assigned_b]
    return matched, unmatched_a, unmatched_b


@dataclass
class Track:
    """Full enterprise track with Kalman filter, trail, speed and direction voting."""
    id:           int
    kalman:       KalmanBoxTracker
    cls_id:       int
    cat_key:      str
    lane:         int
    conf:         float
    state:        int   = 1   # 1=TENTATIVE 2=CONFIRMED 3=DELETED
    counted:      bool  = False
    counted_line: int   = -1
    age:          int   = 1
    misses:       int   = 0
    trail:        deque = field(default_factory=lambda: deque(maxlen=60))
    speeds:       deque = field(default_factory=lambda: deque(maxlen=10))
    count_ts:     Optional[str]   = None
    count_frame:  int   = -1
    direction_votes: deque = field(default_factory=lambda: deque(maxlen=8))
    in_zone:      bool  = False
    headway:      Optional[float] = None

    @property
    def box(self) -> np.ndarray:
        return self.kalman.get_state()

    @property
    def cx(self) -> int:
        b=self.box; return int((b[0]+b[2])/2)

    @property
    def cy(self) -> int:
        b=self.box; return int((b[1]+b[3])/2)

    @property
    def bw(self) -> int:
        b=self.box; return int(b[2]-b[0])

    @property
    def bh(self) -> int:
        b=self.box; return int(b[3]-b[1])

    @property
    def avg_speed(self) -> float:
        return float(np.mean(self.speeds)) if self.speeds else 0.0

    @property
    def vx(self) -> float:
        return self.kalman.velocity[0]

    @property
    def vy(self) -> float:
        return self.kalman.velocity[1]

    def update_trail(self):
        self.trail.append((self.cx, self.cy))

    def update_direction(self):
        if self.vy > 1.5:    self.direction_votes.append(1)
        elif self.vy < -1.5: self.direction_votes.append(-1)

    @property
    def direction_consensus(self) -> int:
        if not self.direction_votes: return 0
        return 1 if sum(self.direction_votes) >= 0 else -1

    @property
    def is_wrong_way(self) -> bool:
        return self.direction_consensus == -1 and len(self.direction_votes) >= 5


class ByteTracker:
    """
    ByteTrack-style tracker with Kalman filtering.
    Splits detections into high/low confidence for better occlusion handling.
    """

    def __init__(self, max_misses: int = 15, min_hits: int = 2):
        self.tracks: List[Track] = []
        self.max_misses   = max_misses
        self.min_hits     = min_hits
        self.iou_hi_thresh = 0.30
        self.iou_lo_thresh = 0.15
        self.conf_hi       = 0.50
        self._next_id      = 1

    def _new_track(self, det: tuple, lane: int) -> Track:
        x1,y1,x2,y2,cls_id,cat_key,conf = det
        kal = KalmanBoxTracker(np.array([x1,y1,x2,y2], dtype=np.float32))
        kal.id = self._next_id; self._next_id += 1
        t = Track(id=kal.id, kalman=kal, cls_id=cls_id,
                  cat_key=cat_key, lane=lane, conf=conf)
        t.update_trail()
        return t

    def step(self, dets: List[tuple], lane_fn) -> List[Track]:
        """
        dets  : list of (x1,y1,x2,y2, cls_id, cat_key, conf)
        lane_fn: callable(cx) → lane_number
        """
        for t in self.tracks:
            t.kalman.predict(); t.misses += 1; t.age += 1
            t.update_direction()

        if not dets:
            self.tracks = [t for t in self.tracks if t.misses <= self.max_misses]
            return [t for t in self.tracks
                    if t.kalman.hit_streak >= self.min_hits or t.counted]

        dets_hi = [d for d in dets if d[6] >= self.conf_hi]
        dets_lo = [d for d in dets if d[6] <  self.conf_hi]

        def match(det_list, track_list, iou_thresh):
            if not det_list or not track_list:
                return [], list(range(len(track_list))), list(range(len(det_list)))
            tb = np.array([t.box for t in track_list], dtype=np.float32)
            db = np.array([[d[0],d[1],d[2],d[3]] for d in det_list], dtype=np.float32)
            return _hungarian(1.0 - _iou_matrix(tb, db), 1.0 - iou_thresh)

        matched_hi, unm_t, unm_d_hi = match(dets_hi, self.tracks, self.iou_hi_thresh)
        for ti, di in matched_hi:
            d=dets_hi[di]; t=self.tracks[ti]
            t.kalman.update(np.array([d[0],d[1],d[2],d[3]], dtype=np.float32))
            t.cat_key=d[5]; t.conf=d[6]
            t.lane=lane_fn((d[0]+d[2])//2); t.misses=0; t.update_trail()

        unm_tracks = [self.tracks[i] for i in unm_t]
        matched_lo, _, _ = match(dets_lo, unm_tracks, self.iou_lo_thresh)
        for ti, di in matched_lo:
            d=dets_lo[di]; t=unm_tracks[ti]
            t.kalman.update(np.array([d[0],d[1],d[2],d[3]], dtype=np.float32))
            t.cat_key=d[5]; t.conf=d[6]
            t.lane=lane_fn((d[0]+d[2])//2); t.misses=0; t.update_trail()

        for di in unm_d_hi:
            d=dets_hi[di]; lane=lane_fn((d[0]+d[2])//2)
            self.tracks.append(self._new_track(d, lane))

        self.tracks=[t for t in self.tracks if t.misses<=self.max_misses]
        return [t for t in self.tracks
                if t.kalman.hit_streak>=self.min_hits or t.counted]


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 7 — LANE MANAGER  (White-Line Detection + Drawing)
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class LaneBoundary:
    x_top:    float
    x_bottom: float
    y_top:    int
    y_bottom: int
    detected: bool = False

    def x_at(self, y: int) -> int:
        if self.y_bottom == self.y_top: return int(self.x_top)
        t = (y - self.y_top) / (self.y_bottom - self.y_top)
        return int(self.x_top + t * (self.x_bottom - self.x_top))


@dataclass
class Lane:
    number:      int
    left:        LaneBoundary
    right:       LaneBoundary
    color:       Tuple[int,int,int] = (255, 255, 255)
    label:       str = ""
    vehicle_ids: List[int] = field(default_factory=list)

    def contains_cx(self, cx: int, cy: int) -> bool:
        return self.left.x_at(cy) <= cx <= self.right.x_at(cy)

    def center_x_at(self, y: int) -> int:
        return (self.left.x_at(y) + self.right.x_at(y)) // 2

    def width_at(self, y: int) -> int:
        return abs(self.right.x_at(y) - self.left.x_at(y))


class WhiteLineDetector:
    """
    Detects white/yellow road lane markings via HLS masking + Hough lines.
    Includes temporal smoothing over 8 frames.
    """

    def __init__(self, frame_h: int, frame_w: int,
                 roi_top_frac: float = 0.45, roi_bottom_frac: float = 0.95):
        self.H = frame_h; self.W = frame_w
        self.roi_top    = int(frame_h * roi_top_frac)
        self.roi_bottom = int(frame_h * roi_bottom_frac)
        self._history: List[List] = []
        self._max_history = 8

    def _extract_white_mask(self, frame: np.ndarray) -> np.ndarray:
        hls = cv2.cvtColor(frame, cv2.COLOR_BGR2HLS)
        mask  = cv2.inRange(hls, np.array([0,  180,  0], np.uint8),
                                  np.array([180,255, 40], np.uint8))
        mask_y= cv2.inRange(hls, np.array([15,  80,100], np.uint8),
                                  np.array([35, 255,255], np.uint8))
        combined = cv2.bitwise_or(mask, mask_y)
        k = cv2.getStructuringElement(cv2.MORPH_RECT,(5,5))
        return cv2.morphologyEx(combined, cv2.MORPH_CLOSE, k)

    def _apply_roi(self, mask: np.ndarray) -> np.ndarray:
        roi = np.zeros_like(mask)
        pts = np.array([
            [int(self.W*.05), self.roi_bottom], [int(self.W*.42), self.roi_top],
            [int(self.W*.58), self.roi_top],    [int(self.W*.95), self.roi_bottom],
        ], dtype=np.int32)
        cv2.fillPoly(roi, [pts], 255)
        return cv2.bitwise_and(mask, roi)

    def _hough_lines(self, edges):
        lines = cv2.HoughLinesP(edges, 1, np.pi/180, 40, minLineLength=60, maxLineGap=80)
        left, right = [], []
        if lines is None: return left, right
        cx = self.W / 2
        for line in lines:
            x1,y1,x2,y2 = line[0]
            if x2 == x1: continue
            slope = (y2-y1)/(x2-x1)
            if abs(slope) < 0.35: continue
            if slope < 0 and x1 < cx and x2 < cx:        left.append(line[0])
            elif slope > 0 and x1 > cx*.5 and x2 > cx*.5: right.append(line[0])
        return left, right

    def _fit_line(self, lines) -> Optional[Tuple[float,float]]:
        if not lines: return None
        pts = []
        for x1,y1,x2,y2 in lines: pts.extend([(x1,y1),(x2,y2)])
        pts = np.array(pts, dtype=np.float32)
        vx,vy,x0,y0 = cv2.fitLine(pts, cv2.DIST_L2, 0, 0.01, 0.01)
        if abs(vx[0]) < 1e-6: return None
        slope = vy[0]/vx[0]; intercept = y0[0]-slope*x0[0]
        return float(slope), float(intercept)

    def _smooth(self, new_lines):
        self._history.append(new_lines)
        if len(self._history) > self._max_history: self._history.pop(0)
        out = []
        for i in range(len(new_lines)):
            vals = [h[i] for h in self._history if h[i] is not None]
            if vals:
                out.append((float(np.mean([v[0] for v in vals])),
                            float(np.mean([v[1] for v in vals]))))
            else:
                out.append(None)
        return out

    def _line_to_boundary(self, slope: float, intercept: float) -> LaneBoundary:
        def x_at_y(y): return (y-intercept)/slope if abs(slope)>1e-6 else 0
        return LaneBoundary(x_top=x_at_y(self.roi_top), x_bottom=x_at_y(self.roi_bottom),
                            y_top=self.roi_top, y_bottom=self.roi_bottom, detected=True)

    def detect(self, frame: np.ndarray, n_lanes: int = 2
               ) -> Optional[List[LaneBoundary]]:
        white   = self._extract_white_mask(frame)
        roi     = self._apply_roi(white)
        edges   = cv2.Canny(roi, 50, 150)
        left_l, right_l = self._hough_lines(edges)
        fit_l = self._fit_line(left_l); fit_r = self._fit_line(right_l)
        if n_lanes == 2:
            smoothed = self._smooth([fit_l, fit_r])
            results  = [self._line_to_boundary(*f) if f else None for f in smoothed]
            if any(r is not None for r in results): return results
        return None


class LaneManager:
    """
    Enterprise lane manager supporting auto / manual / uniform modes.
    Auto mode runs Hough white-line detection every 15 frames and falls back to uniform.
    Renders real-road-style solid edge lines + dashed interior dividers.
    """

    _COLORS = [(255,255,255),(200,230,255),(210,255,200),(255,230,180),(255,200,255)]

    def __init__(self, frame_w: int, frame_h: int, n_lanes: int,
                 direction: str = "NS", mode: str = "auto",
                 manual_zones: Optional[List] = None, roi_top_frac: float = 0.45):
        self.W=frame_w; self.H=frame_h; self.n_lanes=n_lanes
        self.direction=direction; self.mode=mode
        self.roi_top=int(frame_h*roi_top_frac)
        self.detector=WhiteLineDetector(frame_h, frame_w, roi_top_frac)
        self.lanes: List[Lane]=[]
        self._frame_count=0; self._detect_every=15

        if mode=="manual" and manual_zones:
            self._build_from_zones(manual_zones)
        else:
            self._build_uniform()

    def _build_uniform(self):
        lw = self.W / self.n_lanes
        self.lanes = []
        for i in range(self.n_lanes):
            left  = LaneBoundary(x_top=i*lw,     x_bottom=i*lw,
                                  y_top=self.roi_top, y_bottom=self.H)
            right = LaneBoundary(x_top=(i+1)*lw, x_bottom=(i+1)*lw,
                                  y_top=self.roi_top, y_bottom=self.H)
            color=self._COLORS[i%len(self._COLORS)]
            self.lanes.append(Lane(number=i+1, left=left, right=right,
                                   color=color, label=f"Carril {i+1}"))

    def _build_from_zones(self, zones):
        self.lanes=[]
        for i, zone in enumerate(zones):
            pts=np.array(zone, dtype=np.float32)
            left =LaneBoundary(x_top=float(pts[0,0]),x_bottom=float(pts[3,0]),
                                y_top=int(pts[0,1]),  y_bottom=int(pts[3,1]))
            right=LaneBoundary(x_top=float(pts[1,0]),x_bottom=float(pts[2,0]),
                                y_top=int(pts[1,1]),  y_bottom=int(pts[2,1]))
            color=self._COLORS[i%len(self._COLORS)]
            self.lanes.append(Lane(number=i+1,left=left,right=right,
                                   color=color,label=f"Carril {i+1}"))

    def _update_from_detection(self, boundaries):
        detected=[b for b in boundaries if b is not None]
        if len(detected)<2: return
        detected.sort(key=lambda b: b.x_bottom)
        if len(detected)>=self.n_lanes+1:
            for i in range(self.n_lanes):
                self.lanes[i].left=detected[i]; self.lanes[i].right=detected[i+1]
        elif len(detected)==2 and self.n_lanes>=2:
            ol=detected[0]; or_=detected[-1]
            st=(or_.x_top-ol.x_top)/self.n_lanes
            sb=(or_.x_bottom-ol.x_bottom)/self.n_lanes
            for i in range(self.n_lanes):
                self.lanes[i].left =LaneBoundary(x_top=ol.x_top+i*st,
                    x_bottom=ol.x_bottom+i*sb,y_top=ol.y_top,y_bottom=ol.y_bottom,detected=True)
                self.lanes[i].right=LaneBoundary(x_top=ol.x_top+(i+1)*st,
                    x_bottom=ol.x_bottom+(i+1)*sb,y_top=ol.y_top,y_bottom=ol.y_bottom,detected=True)

    def update(self, frame: np.ndarray):
        self._frame_count += 1
        if self.mode=="auto" and self._frame_count%self._detect_every==0:
            boundaries=self.detector.detect(frame, self.n_lanes)
            if boundaries: self._update_from_detection(boundaries)

    def get_lane(self, cx: int, cy: int = 0) -> int:
        for lane in self.lanes:
            if lane.contains_cx(cx, cy): return lane.number
        mid_y=(self.roi_top+self.H)//2
        dists=[abs(cx-lane.center_x_at(mid_y)) for lane in self.lanes]
        return self.lanes[int(np.argmin(dists))].number

    def draw(self, frame: np.ndarray, lane_counts: Optional[dict]=None):
        H,W=frame.shape[:2]; overlay=frame.copy()
        for i,lane in enumerate(self.lanes):
            mid_y=(self.roi_top+H)//2
            if True:  # overlay
                pts=np.array([
                    [lane.left.x_at(self.roi_top),  self.roi_top],
                    [lane.right.x_at(self.roi_top), self.roi_top],
                    [lane.right.x_at(H), H], [lane.left.x_at(H), H],
                ], dtype=np.int32)
                color_dim=tuple(max(0,int(c*.15)) for c in lane.color)
                cv2.fillPoly(overlay,[pts],color_dim)
            is_outer_left  = (i==0)
            is_outer_right = (i==self.n_lanes-1)
            self._draw_solid_line(frame,lane.left,(255,255,255),3 if is_outer_left else 2)
            if not is_outer_left:
                self._draw_dashed_line(frame,lane.left,(255,255,255),2,30,20)
            self._draw_solid_line(frame,lane.right,(255,255,255),3 if is_outer_right else 2)
            if not is_outer_right:
                self._draw_dashed_line(frame,lane.right,(255,255,255),2,30,20)

        cv2.addWeighted(overlay,0.25,frame,0.75,0,frame)

        for lane in self.lanes:
            mid_y=self.roi_top+(H-self.roi_top)//3
            badge_x=lane.center_x_at(mid_y)
            count=sum(lane_counts.get(lane.number,{}).values()) if lane_counts else 0
            self._draw_lane_badge(frame,lane,badge_x,mid_y,count)

    @staticmethod
    def _draw_solid_line(frame, boundary: LaneBoundary, color, thickness):
        p1=(int(boundary.x_at(boundary.y_top)),boundary.y_top)
        p2=(int(boundary.x_at(boundary.y_bottom)),boundary.y_bottom)
        cv2.line(frame,p1,p2,color,thickness,cv2.LINE_AA)

    @staticmethod
    def _draw_dashed_line(frame, boundary: LaneBoundary, color, thickness, dash, gap):
        y=boundary.y_top; drawing=True
        while y<boundary.y_bottom:
            yn=min(y+(dash if drawing else gap),boundary.y_bottom)
            if drawing:
                cv2.line(frame,(int(boundary.x_at(y)),y),
                         (int(boundary.x_at(yn)),yn),color,thickness,cv2.LINE_AA)
            y=yn; drawing=not drawing

    @staticmethod
    def _draw_lane_badge(frame, lane: Lane, cx: int, cy: int, count: int):
        label=f"Carril {lane.number}"; sub=f"{count} veh"
        font=cv2.FONT_HERSHEY_DUPLEX; fs=0.52; th=1
        (tw,th2),_=cv2.getTextSize(label,font,fs,th)
        (sw,_),_  =cv2.getTextSize(sub,font,.38,1)
        pad=6; bw=max(tw,sw)+pad*2; bh=th2+16+pad*2
        x0=cx-bw//2; y0=cy-bh//2
        bg=tuple(max(0,int(c*.4)) for c in lane.color)
        cv2.rectangle(frame,(x0-1,y0-1),(x0+bw+1,y0+bh+1),(0,0,0),-1)
        cv2.rectangle(frame,(x0,y0),(x0+bw,y0+bh),bg,-1)
        cv2.rectangle(frame,(x0,y0),(x0+bw,y0+bh),lane.color,1)
        cv2.putText(frame,label,(x0+(bw-tw)//2,y0+th2+pad),
                    font,fs,lane.color,th,cv2.LINE_AA)
        cv2.putText(frame,sub,(x0+(bw-sw)//2,y0+bh-pad+2),
                    font,.38,(200,255,200),1,cv2.LINE_AA)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 8 — HUD ENGINE
# ══════════════════════════════════════════════════════════════════════════════

C_WHITE=(255,255,255); C_GOLD=(0,210,255); C_GREEN=(0,220,80)
C_RED=(0,0,220); C_ORANGE=(0,140,255); C_PANEL_BG=(12,12,12)


def _put_text(img, text, pos, scale=0.45, color=C_WHITE, thickness=1, bg=None):
    font=cv2.FONT_HERSHEY_DUPLEX
    if bg is not None:
        (tw,th),bl=cv2.getTextSize(text,font,scale,thickness)
        x,y=pos
        cv2.rectangle(img,(x-2,y-th-2),(x+tw+2,y+bl+1),bg,-1)
    cv2.putText(img,text,pos,font,scale,color,thickness,cv2.LINE_AA)


def _alpha_rect(img, x1, y1, x2, y2, color, alpha=0.70):
    sub=img[y1:y2,x1:x2]
    cv2.addWeighted(np.full_like(sub,color),1-alpha,sub,alpha,0,sub)
    img[y1:y2,x1:x2]=sub


class HUDEngine:
    """
    Renders enterprise heads-up display:
    top-left category panel · top-right lane stats · bottom bar ·
    counting lines · vehicle boxes + trails + speed tags · heatmap
    """

    def __init__(self, cfg_direction: str, cfg_classification: str,
                 cfg_camera_id: str, cfg_site_name: str,
                 cfg_show_speed: bool=True, cfg_show_trails: bool=True,
                 cfg_show_wrong_way: bool=True, cfg_show_heatmap: bool=False):
        self.direction   = cfg_direction
        self.system      = cfg_classification
        self.camera_id   = cfg_camera_id
        self.site_name   = cfg_site_name
        self.show_speed  = cfg_show_speed
        self.show_trails = cfg_show_trails
        self.show_wrong  = cfg_show_wrong_way
        self.show_hmap   = cfg_show_heatmap
        self._heatmap: Optional[np.ndarray] = None

    def draw_count_lines(self, frame, line_ys: List[int]):
        H,W=frame.shape[:2]
        styles=[((0,180,180),1),((0,255,255),2),((0,180,180),1)]
        for ly,(col,th) in zip(line_ys,styles):
            for x in range(0,W,24):
                cv2.line(frame,(x,ly),(min(x+14,W),ly),col,th,cv2.LINE_AA)
        mid=line_ys[len(line_ys)//2]
        _put_text(frame,"◀ LÍNEAS DE CONTEO ▶",(W//2-100,mid-8),
                  scale=0.42,color=(0,255,255),bg=(0,0,0))

    def draw_cat_panel(self, frame, cats, cat_counts, total, fps):
        n_rows=len(cats)+5; ph=24+n_rows*22; pw=240
        _alpha_rect(frame,4,4,pw+4,ph+4,C_PANEL_BG,alpha=0.55)
        cv2.rectangle(frame,(4,4),(pw+4,ph+4),(60,60,60),1)
        y=26
        _put_text(frame,f"TOTAL VEHÍCULOS: {total}",(10,y),scale=0.65,color=C_GOLD,thickness=2)
        y+=22
        _put_text(frame,f"Sistema: {self.system}  Dir: {self.direction}",
                  (10,y),scale=0.38,color=(160,255,160))
        y+=18; cv2.line(frame,(10,y),(pw-2,y),(60,60,60),1); y+=10
        for key,info in cats.items():
            cnt=cat_counts.get(key,0); color=info["color"]
            bgr=(int(color[2]),int(color[1]),int(color[0]))
            pct=(cnt/max(total,1))*100
            bar_w=int((pw-120)*(cnt/max(total,1)))
            cv2.rectangle(frame,(10,y-10),(10+bar_w,y-1),
                          tuple(max(0,int(c*.5)) for c in bgr),-1)
            name_s=info["name"][:20]+("…" if len(info["name"])>20 else "")
            _put_text(frame,f"{key}",(10,y),scale=0.44,color=bgr)
            _put_text(frame,f"{cnt:>4}  {pct:>5.1f}%",(pw-110,y),scale=0.40,color=C_WHITE)
            y+=22
        y+=4; cv2.line(frame,(10,y),(pw-2,y),(60,60,60),1); y+=14
        _put_text(frame,f"FPS: {fps:>5.1f}",(10,y),scale=0.40,color=(180,180,180))

    def draw_lane_panel(self, frame, lane_counts, density_stats):
        H,W=frame.shape[:2]; lane_ns=sorted(lane_counts.keys())
        if not lane_ns: return
        rows=len(lane_ns)+3; ph=24+rows*28; pw=210; x0=W-pw-4
        _alpha_rect(frame,x0,4,W-4,ph+4,C_PANEL_BG,alpha=0.55)
        cv2.rectangle(frame,(x0,4),(W-4,ph+4),(60,60,60),1)
        y=26; _put_text(frame,"POR CARRIL",(x0+8,y),scale=0.55,color=C_GOLD)
        y+=22
        for ln in lane_ns:
            ltotal=sum(lane_counts[ln].values()); ds=density_stats.get(ln,{})
            vph=ds.get("flow_vph_1m",0); cong=ds.get("congestion","")
            c_col={"LIBRE":C_GREEN,"MODERADO":(0,220,220),
                   "CONGESTIONADO":C_ORANGE,"SATURADO":C_RED}.get(cong,C_WHITE)
            _put_text(frame,f"Carril {ln}:",(x0+8,y),scale=0.46,color=C_WHITE)
            _put_text(frame,f"{ltotal:>4} veh",(x0+90,y),scale=0.46,color=C_WHITE)
            y+=16
            _put_text(frame,f"  {vph:>5.0f} veh/h",(x0+8,y),scale=0.38,color=(160,200,255))
            _put_text(frame,cong,(x0+130,y),scale=0.34,color=c_col)
            y+=20; cv2.line(frame,(x0+6,y),(x0+pw-8,y),(40,40,40),1); y+=8

    def draw_bottom_bar(self, frame, fps, frame_n, total):
        H,W=frame.shape[:2]; bar_h=26
        _alpha_rect(frame,0,H-bar_h,W,H,C_PANEL_BG,alpha=0.60)
        ts=datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        _put_text(frame,ts,(W//2-90,H-8),scale=0.44,color=(200,200,200))
        _put_text(frame,f"FPS:{fps:.1f}  F:{frame_n}  TOT:{total}",
                  (10,H-8),scale=0.40,color=(150,150,150))
        _put_text(frame,f"CAM: {self.camera_id}  {self.site_name}",
                  (W-290,H-8),scale=0.40,color=(150,150,150))

    def update_heatmap(self, frame, tracks):
        H,W=frame.shape[:2]
        if self._heatmap is None: self._heatmap=np.zeros((H,W),dtype=np.float32)
        for tr in tracks:
            if 0<=tr.cx<W and 0<=tr.cy<H:
                cv2.circle(self._heatmap,(tr.cx,tr.cy),20,1.0,-1)
        self._heatmap*=0.97

    def draw_heatmap(self, frame):
        if self._heatmap is None: return
        h_norm=cv2.normalize(self._heatmap,None,0,255,cv2.NORM_MINMAX).astype(np.uint8)
        h_col=cv2.applyColorMap(h_norm,cv2.COLORMAP_JET)
        cv2.addWeighted(frame,0.65,h_col,0.35,0,frame)

    def draw_vehicle(self, frame, tr, cat_info, speed_kmh):
        x1,y1,x2,y2=[int(v) for v in tr.box]
        color=cat_info.get("color",(200,200,200))
        bgr=(int(color[2]),int(color[1]),int(color[0]))
        th=3 if tr.counted else 2
        cv2.rectangle(frame,(x1,y1),(x2,y2),bgr,th)
        corner=10
        for cx_,cy_ in [(x1,y1),(x2,y1),(x1,y2),(x2,y2)]:
            dx=1 if cx_==x1 else -1; dy=1 if cy_==y1 else -1
            cv2.line(frame,(cx_,cy_),(cx_+dx*corner,cy_),C_WHITE,2)
            cv2.line(frame,(cx_,cy_),(cx_,cy_+dy*corner),C_WHITE,2)
        lbl=f"{cat_info['label']}  L{tr.lane}  #{tr.id}"
        _put_text(frame,lbl,(x1,y1-4),scale=0.42,color=C_WHITE,bg=bgr)
        if self.show_speed and speed_kmh>0.5:
            col=C_GREEN if speed_kmh<60 else (C_ORANGE if speed_kmh<90 else C_RED)
            _put_text(frame,f"{speed_kmh:.0f}km/h",(x1,y2+14),
                      scale=0.38,color=col,bg=(0,0,0))
        if self.show_trails:
            pts=list(tr.trail)
            for i in range(1,len(pts)):
                frac=i/max(len(pts)-1,1)
                cv2.line(frame,pts[i-1],pts[i],
                         tuple(int(c*frac) for c in bgr),1,cv2.LINE_AA)
            cv2.circle(frame,(tr.cx,tr.cy),4,bgr,-1)
        if self.show_wrong and tr.is_wrong_way:
            cv2.rectangle(frame,(x1-2,y1-2),(x1+110,y1-20),(0,0,180),-1)
            _put_text(frame,"⚠ CONTRAFLUJO",(x1,y1-6),scale=0.44,color=C_WHITE)

    def draw_all(self, frame, active_tracks, cats, cat_counts, lane_counts,
                 density_stats, fps, frame_n, speed_map, count_lines):
        total=sum(cat_counts.values())
        if self.show_hmap:
            self.update_heatmap(frame,active_tracks); self.draw_heatmap(frame)
        for tr in active_tracks:
            info=cats.get(tr.cat_key,{"label":tr.cat_key,"color":(200,200,200),"icon":""})
            self.draw_vehicle(frame,tr,info,speed_map.get(tr.id,0.0))
        self.draw_count_lines(frame,count_lines)
        self.draw_cat_panel(frame,cats,cat_counts,total,fps)
        self.draw_lane_panel(frame,lane_counts,density_stats)
        self.draw_bottom_bar(frame,fps,frame_n,total)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 9 — EXCEL REPORT  (5 sheets)
# ══════════════════════════════════════════════════════════════════════════════

_EXCEL_DARK_BLUE="1F3864"; _EXCEL_MED_BLUE="2E75B6"
_EXCEL_LIGHT_BLU="D6E4F7"; _EXCEL_GRAY_HDR="F2F2F2"
_CAT_COLORS_XL={
    "CAT I":"C9B3FF","CAT II":"A8F0C0","CAT III":"A0D8FF",
    "CAT IV":"FFE680","CAT V":"FFB366","CAT VI":"FF9090","CAT VII":"FF6060",
    "C1":"A8F0C0","C2P":"FFE680","C2G":"A0D8FF","C3":"FFB366","C4":"FF9090","C5":"FF6060",
}


def _xl_hdr(ws, cell_ref, text, fill_hex=_EXCEL_DARK_BLUE, font_color="FFFFFF",
            bold=True, size=11, center=True):
    from openpyxl.styles import Font, PatternFill, Alignment
    c=ws[cell_ref]; c.value=text
    c.font=Font(color=font_color,bold=bold,size=size)
    c.fill=PatternFill("solid",fgColor=fill_hex)
    if center: c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    return c


def write_excel(path: str, summary: dict, lane_counts: dict, events: List[dict],
                speed_events: List[dict], system: str, direction: str,
                camera_id: str, video_source: str, n_lanes: int,
                site_name: str="Peaje Niquía", operator: str=""):
    if not HAVE_EXCEL:
        print("[WARN] openpyxl not installed — skipping Excel."); return
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook()
    brd_s=Side(style="thin",color="AAAAAA")
    brd=Border(left=brd_s,right=brd_s,top=brd_s,bottom=brd_s)
    ctr=Alignment(horizontal="center",vertical="center")
    lft=Alignment(horizontal="left",vertical="center")

    # ── Sheet 1: Resumen Ejecutivo ──────────────────────────────────────────
    ws=wb.active; ws.title="Resumen Ejecutivo"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:H1")
    _xl_hdr(ws,"A1",f"INFORME DE CONTEO VEHICULAR — {site_name.upper()}",size=14)
    ws.row_dimensions[1].height=32
    ws.merge_cells("A2:H2")
    _xl_hdr(ws,"A2",
            f"Sistema {system}  |  Dirección {direction}  |  "
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            _EXCEL_MED_BLUE,size=10)

    meta=[("Cámara / Fuente:",camera_id or str(video_source)),
          ("Sistema de clasificación:",system),("Dirección de flujo:",direction),
          ("Número de carriles:",str(n_lanes)),("Operador:",operator or "—"),
          ("Fecha de generación:",datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
          ("Total vehículos:",str(sum(summary.values())))]
    for r,(k,v) in enumerate(meta,start=4):
        ws.cell(r,1,k).font=Font(bold=True,size=10); ws.cell(r,1).alignment=lft
        ws.cell(r,2,v).alignment=lft; ws.cell(r,2).font=Font(size=10)
        if r%2==0:
            for c in range(1,3): ws.cell(r,c).fill=PatternFill("solid",fgColor=_EXCEL_LIGHT_BLU)

    grand_total=sum(summary.values()); r_start=13
    for ci,h in enumerate(["Categoría","Descripción","Conteo","% del Total","Ejes"],start=1):
        _xl_hdr(ws,ws.cell(r_start,ci).coordinate,h,_EXCEL_MED_BLUE)
    cats_sys=CLASSIFICATION_SYSTEMS.get(system,{})
    for ri,(cat,cnt) in enumerate(summary.items(),start=r_start+1):
        pct=(cnt/grand_total*100) if grand_total else 0
        axles=cats_sys.get(cat,{}).get("axles","—")
        for ci,val in enumerate([cat,cat_name(cat,system),cnt,f"{pct:.1f}%",axles],start=1):
            c2=ws.cell(ri,ci,val); c2.border=brd
            c2.alignment=ctr if ci!=2 else lft
            c2.fill=PatternFill("solid",fgColor=_CAT_COLORS_XL.get(cat,"FFFFFF"))
            if ci==3: c2.font=Font(bold=True)
    r_tot=r_start+len(summary)+1
    for ci,val in enumerate(["TOTAL","",grand_total,"100%"],start=1):
        c2=ws.cell(r_tot,ci,val); c2.font=Font(bold=True,size=11 if ci==1 else 10)
        c2.fill=PatternFill("solid",fgColor=_EXCEL_LIGHT_BLU)
        c2.border=brd; c2.alignment=ctr
    for col,w in zip("ABCDE",[14,40,10,14,8]):
        ws.column_dimensions[col].width=w

    try:
        from openpyxl.chart import BarChart, Reference
        chart=BarChart(); chart.type="col"; chart.grouping="clustered"
        chart.title="Distribución por Categoría"; chart.width=18; chart.height=12
        dr=Reference(ws,min_col=3,max_col=3,min_row=r_start,max_row=r_start+len(summary))
        cr=Reference(ws,min_col=1,max_col=1,min_row=r_start+1,max_row=r_start+len(summary))
        chart.add_data(dr,titles_from_data=True); chart.set_categories(cr)
        ws.add_chart(chart,f"G{r_start}")
    except Exception: pass

    # ── Sheet 2: Por Carril ────────────────────────────────────────────────
    ws2=wb.create_sheet("Por Carril"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:J1")
    _xl_hdr(ws2,"A1","CONTEO POR CARRIL Y CATEGORÍA",size=13)
    all_cats=sorted(summary.keys()); lane_nums=sorted(lane_counts.keys())
    hdr2=["Carril"]+all_cats+["TOTAL","% del Total"]
    for ci,h in enumerate(hdr2,start=1):
        _xl_hdr(ws2,ws2.cell(3,ci).coordinate,h,_EXCEL_MED_BLUE)
        ws2.column_dimensions[get_column_letter(ci)].width=13
    for ri,lane in enumerate(lane_nums,start=4):
        row_total=0; ws2.cell(ri,1,f"Carril {lane}").font=Font(bold=True)
        ws2.cell(ri,1).alignment=ctr
        for ci,cat in enumerate(all_cats,start=2):
            cnt=lane_counts[lane].get(cat,0)
            ws2.cell(ri,ci,cnt).alignment=ctr; ws2.cell(ri,ci).border=brd
            ws2.cell(ri,ci).fill=PatternFill("solid",fgColor=_CAT_COLORS_XL.get(cat,"F0F0F0"))
            row_total+=cnt
        pct=(row_total/grand_total*100) if grand_total else 0
        ws2.cell(ri,len(all_cats)+2,row_total).font=Font(bold=True)
        ws2.cell(ri,len(all_cats)+2).alignment=ctr
        ws2.cell(ri,len(all_cats)+3,f"{pct:.1f}%").alignment=ctr
        if ri%2==0: ws2.cell(ri,1).fill=PatternFill("solid",fgColor=_EXCEL_GRAY_HDR)
    r2_tot=4+len(lane_nums)
    ws2.cell(r2_tot,1,"TOTAL").font=Font(bold=True)
    for ci,cat in enumerate(all_cats,start=2):
        t=sum(lane_counts[l].get(cat,0) for l in lane_nums)
        ws2.cell(r2_tot,ci,t).font=Font(bold=True); ws2.cell(r2_tot,ci).alignment=ctr
        ws2.cell(r2_tot,ci).fill=PatternFill("solid",fgColor=_EXCEL_LIGHT_BLU)
    ws2.cell(r2_tot,len(all_cats)+2,grand_total).font=Font(bold=True)
    ws2.cell(r2_tot,len(all_cats)+2).alignment=ctr
    ws2.cell(r2_tot,len(all_cats)+2).fill=PatternFill("solid",fgColor=_EXCEL_LIGHT_BLU)

    # ── Sheet 3: Flujo por Minuto ──────────────────────────────────────────
    ws3=wb.create_sheet("Flujo por Minuto"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:J1")
    _xl_hdr(ws3,"A1","VOLUMEN DE TRÁFICO POR MINUTO",size=13)
    if events:
        minute_data: dict = defaultdict(lambda: defaultdict(int))
        for ev in events:
            minute_data[ev.get("time","00:00")[:5]][ev.get("category","?")] += 1
        hdr3=["Minuto"]+all_cats+["TOTAL"]
        for ci,h in enumerate(hdr3,start=1):
            _xl_hdr(ws3,ws3.cell(3,ci).coordinate,h,_EXCEL_MED_BLUE)
        for ri,(minute,cat_cnts) in enumerate(sorted(minute_data.items()),start=4):
            ws3.cell(ri,1,minute).alignment=ctr; row_total=0
            for ci,cat in enumerate(all_cats,start=2):
                cnt=cat_cnts.get(cat,0); ws3.cell(ri,ci,cnt).alignment=ctr
                ws3.cell(ri,ci).border=brd; row_total+=cnt
            ws3.cell(ri,len(all_cats)+2,row_total).font=Font(bold=True)
            ws3.cell(ri,len(all_cats)+2).alignment=ctr
            if ri%2==0:
                for ci in range(1,len(all_cats)+3):
                    ws3.cell(ri,ci).fill=PatternFill("solid",fgColor=_EXCEL_GRAY_HDR)
        for c in range(1,len(all_cats)+3):
            ws3.column_dimensions[get_column_letter(c)].width=12

    # ── Sheet 4: Velocidades ───────────────────────────────────────────────
    ws4=wb.create_sheet("Velocidades"); ws4.sheet_view.showGridLines=False
    ws4.merge_cells("A1:G1")
    _xl_hdr(ws4,"A1","REGISTRO DE VELOCIDADES",size=13)
    for ci,h in enumerate(["#","Timestamp","ID Track","Categoría","Carril","Velocidad (km/h)","Nivel"],start=1):
        _xl_hdr(ws4,ws4.cell(3,ci).coordinate,h,_EXCEL_MED_BLUE)
    for ri,ev in enumerate(speed_events[:5000],start=4):
        spd=ev.get("speed_kmh",0)
        lvl="Baja" if spd<40 else "Normal" if spd<80 else "Alta" if spd<120 else "Exceso"
        for ci,val in enumerate([ri-3,ev.get("time",""),ev.get("track_id",""),
                                  ev.get("category",""),ev.get("lane",""),round(spd,1),lvl],start=1):
            ws4.cell(ri,ci,val).alignment=ctr; ws4.cell(ri,ci).border=brd
        spd_col="00CC66" if spd<60 else "FFCC00" if spd<90 else "FF4444"
        ws4.cell(ri,6).fill=PatternFill("solid",fgColor=spd_col)
    for col,w in zip("ABCDEFG",[6,14,10,14,10,18,12]):
        ws4.column_dimensions[col].width=w

    # ── Sheet 5: Registro Detallado ────────────────────────────────────────
    ws5=wb.create_sheet("Registro Detallado"); ws5.sheet_view.showGridLines=False
    ws5.merge_cells("A1:J1")
    _xl_hdr(ws5,"A1","REGISTRO DE EVENTOS INDIVIDUALES",size=13)
    for ci,h in enumerate(["#","Timestamp","Frame","ID Track","Tipo YOLO",
                             "Categoría","Carril","Dirección","Confianza","Velocidad km/h"],start=1):
        _xl_hdr(ws5,ws5.cell(3,ci).coordinate,h,_EXCEL_MED_BLUE)
    for ri,ev in enumerate(events[:10000],start=4):
        for ci,val in enumerate([ri-3,ev.get("time",""),ev.get("frame",""),
                                   ev.get("track_id",""),ev.get("yolo_type",""),
                                   ev.get("category",""),ev.get("lane",""),
                                   ev.get("direction",""),round(ev.get("conf",0.0),2),
                                   round(ev.get("speed_kmh",0.0),1)],start=1):
            ws5.cell(ri,ci,val).alignment=ctr; ws5.cell(ri,ci).border=brd
        if ri%2==0:
            for ci in range(1,11):
                ws5.cell(ri,ci).fill=PatternFill("solid",fgColor=_EXCEL_GRAY_HDR)
    for ci,w in enumerate([6,14,8,10,14,12,10,14,12,16],start=1):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    for sheet in wb.worksheets: sheet.freeze_panes="A4"
    wb.save(path)
    print(f"  [Excel] → {path}  ({len(events)} eventos, {len(speed_events)} velocidades)")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 10 — PDF REPORT
# ══════════════════════════════════════════════════════════════════════════════

def write_pdf(path: str, summary: dict, lane_counts: dict, events: List[dict],
              speed_events: List[dict], system: str, direction: str,
              camera_id: str, video_source: str, n_lanes: int,
              site_name: str="Peaje Niquía", road_name: str="Autopista Medellín–Bogotá",
              operator: str="", notes: str=""):
    if not HAVE_PDF:
        print("[WARN] reportlab not installed — skipping PDF."); return

    _DARK_BLUE=rl_colors.HexColor("#1F3864"); _MED_BLUE=rl_colors.HexColor("#2E75B6")
    _LIGHT_BLU=rl_colors.HexColor("#D6E4F7"); _ACCENT=rl_colors.HexColor("#F5B700")
    _LT_GRAY=rl_colors.HexColor("#F5F5F5");   _MID_GRAY=rl_colors.HexColor("#CCCCCC")

    _CAT_PDF={
        "CAT I":"#C9B3FF","CAT II":"#A8F0C0","CAT III":"#A0D8FF",
        "CAT IV":"#FFE680","CAT V":"#FFB366","CAT VI":"#FF9090","CAT VII":"#FF6060",
        "C1":"#A8F0C0","C2P":"#FFE680","C2G":"#A0D8FF","C3":"#FFB366","C4":"#FF9090","C5":"#FF6060",
    }

    def _tbl_style(header_rows=1):
        return TableStyle([
            ("BACKGROUND",(0,0),(-1,header_rows-1),_DARK_BLUE),
            ("TEXTCOLOR",(0,0),(-1,header_rows-1),rl_colors.white),
            ("FONTNAME",(0,0),(-1,header_rows-1),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),
            ("GRID",(0,0),(-1,-1),0.4,_MID_GRAY),
            ("ROWBACKGROUNDS",(0,header_rows),(-1,-1),[_LT_GRAY,rl_colors.white]),
        ])

    def _bar_chart(labels, values, w=14*cm, h=7*cm):
        d=Drawing(w,h); bc=VerticalBarChart()
        bc.x=40; bc.y=30; bc.width=w-60; bc.height=h-50
        bc.data=[values]; bc.categoryAxis.categoryNames=labels
        bc.categoryAxis.labels.boxAnchor="ne"; bc.categoryAxis.labels.angle=30
        bc.categoryAxis.labels.fontSize=7; bc.valueAxis.forceZero=1
        bc.bars[0].fillColor=_MED_BLUE; d.add(bc); return d

    def HR():
        return HRFlowable(width="100%",thickness=1,color=_MID_GRAY,spaceAfter=8)

    doc=SimpleDocTemplate(path,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,
                          topMargin=2.5*cm,bottomMargin=2.5*cm)
    styles=getSampleStyleSheet(); story=[]

    def _ps(name,parent="Normal",**kw):
        return ParagraphStyle(name,parent=styles[parent],**kw)

    title_st=_ps("CT",parent="Title",fontSize=20,spaceAfter=6,textColor=_DARK_BLUE,alignment=TA_CENTER)
    sub_st  =_ps("CS",parent="Normal",fontSize=11,spaceAfter=4,textColor=_MED_BLUE,alignment=TA_CENTER)
    h1_st   =_ps("H1",parent="Heading1",fontSize=13,textColor=_DARK_BLUE,spaceBefore=14,spaceAfter=6)
    h2_st   =_ps("H2",parent="Heading2",fontSize=11,textColor=_MED_BLUE,spaceBefore=10,spaceAfter=4)
    body_st =_ps("B",parent="Normal",fontSize=9.5,leading=14,spaceAfter=5,alignment=TA_JUSTIFY)
    bullet_st=_ps("BL",parent="Normal",fontSize=9.5,leading=13,spaceAfter=3,leftIndent=12,firstLineIndent=-8)
    meta_k  =_ps("MK",parent="Normal",fontSize=9,fontName="Helvetica-Bold")
    meta_v  =_ps("MV",parent="Normal",fontSize=9)

    grand_total=sum(summary.values())

    # Cover
    story+=[Spacer(1,2*cm),
            Paragraph("INFORME TÉCNICO PERICIAL",title_st),
            Paragraph("Conteo y Clasificación Vehicular por Carril",title_st),HR(),
            Paragraph(site_name,sub_st),Paragraph(road_name,sub_st),Spacer(1,1*cm)]

    meta_rows=[["Cámara / Fuente:",camera_id or str(video_source)],
               ["Sistema de clasificación:",system],["Dirección de flujo:",direction],
               ["Número de carriles:",str(n_lanes)],["Total vehículos:",str(grand_total)],
               ["Operador:",operator or "N/D"],
               ["Fecha de generación:",datetime.now().strftime("%Y-%m-%d %H:%M:%S")]]
    if notes: meta_rows.append(["Notas:",notes])
    meta_tbl=Table([[Paragraph(k,meta_k),Paragraph(v,meta_v)] for k,v in meta_rows],
                   colWidths=[5.5*cm,11*cm])
    meta_tbl.setStyle(TableStyle([
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[_LIGHT_BLU,rl_colors.white]),
        ("GRID",(0,0),(-1,-1),0.4,_MID_GRAY),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story+=[meta_tbl,PageBreak()]

    # 1. Executive summary
    all_cats=sorted(summary.keys())
    story.append(Paragraph("1. Resumen Ejecutivo por Categoría Vehicular",h1_st)); story.append(HR())
    cat_rows=[["Categoría","Descripción","Conteo","% del Total"]]
    for cat,cnt in sorted(summary.items()):
        pct=(cnt/grand_total*100) if grand_total else 0
        cat_rows.append([cat,cat_name(cat,system),str(cnt),f"{pct:.1f}%"])
    cat_rows.append(["TOTAL","",str(grand_total),"100%"])
    cat_tbl=Table(cat_rows,colWidths=[2.5*cm,9*cm,2.5*cm,3*cm])
    cs=_tbl_style()
    for i,(cat,_) in enumerate(sorted(summary.items()),start=1):
        cs.add("BACKGROUND",(0,i),(-1,i),rl_colors.HexColor(_CAT_PDF.get(cat,"#FFFFFF")))
    cs.add("BACKGROUND",(0,-1),(-1,-1),_LIGHT_BLU); cs.add("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")
    cs.add("ALIGN",(2,0),(-1,-1),"CENTER"); cat_tbl.setStyle(cs)
    story+=[cat_tbl,Spacer(1,.5*cm)]
    if grand_total>0:
        story.append(Paragraph("Distribución por categoría vehicular:",h2_st))
        story+=[ _bar_chart(sorted(summary.keys()),[summary[k] for k in sorted(summary.keys())]),
                 Spacer(1,.5*cm)]

    # 2. Per-lane
    story.append(Paragraph("2. Conteo por Carril y Categoría Vehicular",h1_st)); story.append(HR())
    lane_nums=sorted(lane_counts.keys())
    lane_rows=[["Carril"]+all_cats+["TOTAL"]]
    for lane in lane_nums:
        row=[f"Carril {lane}"]; rt=0
        for cat in all_cats:
            cnt=lane_counts[lane].get(cat,0); row.append(str(cnt)); rt+=cnt
        row.append(str(rt)); lane_rows.append(row)
    tot_row=["TOTAL"]
    for cat in all_cats: tot_row.append(str(sum(lane_counts[l].get(cat,0) for l in lane_nums)))
    tot_row.append(str(grand_total)); lane_rows.append(tot_row)
    col_w=[2*cm]+[14.5*cm/max(len(all_cats)+1,1)]*(len(all_cats)+1)
    lane_tbl=Table(lane_rows,colWidths=col_w); ls=_tbl_style()
    ls.add("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")
    ls.add("BACKGROUND",(0,-1),(-1,-1),_LIGHT_BLU); ls.add("ALIGN",(1,0),(-1,-1),"CENTER")
    lane_tbl.setStyle(ls); story+=[lane_tbl,Spacer(1,.5*cm)]

    # 3. Speed stats
    if speed_events:
        speeds=[ev.get("speed_kmh",0) for ev in speed_events if ev.get("speed_kmh",0)>1]
        if speeds:
            story.append(Paragraph("3. Estadísticas de Velocidad",h1_st)); story.append(HR())
            spd_rows=[["Indicador","Valor"],
                      ["Velocidad media",f"{statistics.mean(speeds):.1f} km/h"],
                      ["Velocidad mediana",f"{statistics.median(speeds):.1f} km/h"],
                      ["Velocidad mínima",f"{min(speeds):.1f} km/h"],
                      ["Velocidad máxima",f"{max(speeds):.1f} km/h"],
                      ["Desviación estándar",f"{statistics.stdev(speeds) if len(speeds)>1 else 0:.1f} km/h"],
                      ["Registros totales",str(len(speeds))]]
            spd_tbl=Table(spd_rows,colWidths=[7*cm,5*cm]); spd_tbl.setStyle(_tbl_style())
            story+=[spd_tbl,Spacer(1,.3*cm)]

    # 4. Flow
    if events:
        by_min: dict=defaultdict(int)
        for ev in events: by_min[ev.get("time","00:00")[:5]]+=1
        if by_min:
            story.append(Paragraph("4. Análisis de Flujo Temporal",h1_st)); story.append(HR())
            peak_min=max(by_min,key=by_min.get)
            avg_flow=sum(by_min.values())/max(len(by_min),1)
            flow_rows=[["Indicador","Valor"],
                       ["Minuto pico",f"{peak_min} ({by_min[peak_min]} veh/min)"],
                       ["Flujo promedio",f"{avg_flow:.1f} veh/min  ({avg_flow*60:.0f} veh/h)"],
                       ["Minutos registrados",str(len(by_min))],
                       ["Total vehículos",str(sum(by_min.values()))]]
            flow_tbl=Table(flow_rows,colWidths=[7*cm,9*cm]); flow_tbl.setStyle(_tbl_style())
            story+=[flow_tbl,PageBreak()]

    # 5. Methodology
    story.append(Paragraph("5. Metodología",h1_st)); story.append(HR())
    story.append(Paragraph(
        f"El conteo y clasificación vehicular se realizó mediante análisis automatizado "
        f"de video utilizando el modelo YOLOv8, complementado con ByteTrack y filtrado "
        f"de Kalman para trayectorias suavizadas. Los vehículos fueron clasificados "
        f"según el sistema <b>{system}</b> mediante heurísticas de relación de aspecto "
        f"y área del bounding box. El conteo se efectuó con tres líneas virtuales "
        f"transversales al 33%, 50% y 67% de la altura del fotograma. Los carriles fueron "
        f"detectados automáticamente via transformación de Hough sobre máscara HLS "
        f"(blanco/amarillo), con suavizado temporal de 8 fotogramas, cayendo en franjas "
        f"equidistantes cuando la detección falla. La velocidad se estimó del desplazamiento "
        f"en píxeles por fotograma convertido a km/h con el factor px/metro calibrado.",
        body_st))
    story.append(Spacer(1,.5*cm))

    # 6. Limitations
    story.append(Paragraph("6. Limitaciones y Recomendaciones",h1_st)); story.append(HR())
    for lim in [
        "La clasificación de camiones pesados (CAT V–VII / C3–C5) se basa en heurísticas de "
        "relación de aspecto; YOLOv8 no detecta directamente el número de ejes. Integrar "
        "clasificadores de ejes por inducción mejoraría la precisión.",
        "La detección de carriles por visión es sensible a lluvia, sombras y deterioro "
        "de pintura vial. Se recomienda calibración geométrica con homografía.",
        "Vehículos en solapamiento pueden generar pérdida temporal de tracking "
        "(subestimación estimada 3–5%).",
        "La estimación de velocidad requiere calibración precisa del factor px/metro.",
        "Se recomienda validar ≥500 vehículos contra revisión manual para certificar "
        "precisión ≥95% ante ANI/INVIAS.",
        "Para condiciones nocturnas, usar cámaras IR y modelos YOLO entrenados para esas condiciones.",
    ]:
        story.append(Paragraph(f"• {lim}",bullet_st))

    story+=[Spacer(1,.5*cm),HR()]
    story.append(Paragraph(
        f"Informe generado automáticamente el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
        f"| Sistema: {system} | Cámara: {camera_id or 'N/A'} | Dirección: {direction} | "
        f"{site_name} — {road_name}",
        _ps("Ft",parent="Normal",fontSize=7.5,textColor=_MID_GRAY,alignment=TA_CENTER)))
    doc.build(story)
    print(f"  [PDF]   → {path}  ({grand_total} vehículos)")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 11 — HTML REPORT
# ══════════════════════════════════════════════════════════════════════════════

def write_html(path: str, summary: dict, lane_counts: dict, events: List[dict],
               speed_events: List[dict], system: str, direction: str,
               camera_id: str, video_source: str, n_lanes: int,
               site_name: str="Peaje Niquía", road_name: str="Autopista Medellín–Bogotá"):
    cats_sys=CLASSIFICATION_SYSTEMS.get(system,{})
    all_cats=sorted(summary.keys()); lane_nums=sorted(lane_counts.keys())
    grand_total=sum(summary.values())

    by_min: dict=defaultdict(int)
    for ev in events: by_min[ev.get("time","00:00")[:5]]+=1

    by_cat_json=json.dumps({k:summary.get(k,0) for k in all_cats})
    cat_names_json=json.dumps([cat_name(c,system) for c in all_cats])
    cat_colors_json=json.dumps([cats_sys.get(c,{}).get("hex","#888888") for c in all_cats])

    lane_datasets=[]
    for cat in all_cats:
        data=[lane_counts[l].get(cat,0) for l in lane_nums]
        lane_datasets.append({"label":cat,"data":data,"backgroundColor":cats_sys.get(cat,{}).get("hex","#888888")})
    lane_ds_json=json.dumps(lane_datasets)
    lane_labels_json=json.dumps([f"Carril {l}" for l in lane_nums])

    flow_times=sorted(by_min.keys()); flow_vals=[by_min[t] for t in flow_times]
    flow_t_json=json.dumps(flow_times); flow_v_json=json.dumps(flow_vals)

    speeds=[ev.get("speed_kmh",0) for ev in speed_events if ev.get("speed_kmh",0)>1]
    avg_speed=f"{sum(speeds)/len(speeds):.1f}" if speeds else "N/D"
    max_speed=f"{max(speeds):.1f}" if speeds else "N/D"

    recent_events=events[-200:]
    rows_html=""
    for ev in recent_events:
        cat_hex=cats_sys.get(ev.get("category",""),{}).get("hex","#888")
        rows_html+=(f"<tr><td>{ev.get('time','')}</td><td>{ev.get('frame','')}</td>"
                    f"<td>{ev.get('track_id','')}</td>"
                    f"<td><span class='badge' style='background:{cat_hex}'>{ev.get('category','')}</span></td>"
                    f"<td>{ev.get('lane','')}</td><td>{ev.get('yolo_type','')}</td>"
                    f"<td>{ev.get('direction','')}</td><td>{round(ev.get('speed_kmh',0),1)}</td></tr>")

    html=f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Informe Vehicular — {site_name}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  :root {{--dark-blue:#1F3864;--med-blue:#2E75B6;--light-blue:#D6E4F7;
          --accent:#F5B700;--bg:#F0F4F8;--card:#ffffff;--text:#1a1a2e}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--text)}}
  header{{background:linear-gradient(135deg,var(--dark-blue),var(--med-blue));
          color:white;padding:24px 32px;display:flex;justify-content:space-between;align-items:center}}
  header h1{{font-size:1.5rem;font-weight:700}}
  header p{{font-size:.85rem;opacity:.85;margin-top:4px}}
  .badge-header{{background:rgba(255,255,255,.15);padding:8px 16px;border-radius:20px;font-size:.8rem;text-align:right}}
  .container{{max-width:1400px;margin:0 auto;padding:24px 20px}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-bottom:24px}}
  .kpi-card{{background:var(--card);border-radius:12px;padding:20px;
             box-shadow:0 2px 12px rgba(0,0,0,.07);border-left:4px solid var(--med-blue);transition:transform .2s}}
  .kpi-card:hover{{transform:translateY(-2px)}}
  .kpi-card .val{{font-size:2rem;font-weight:800;color:var(--dark-blue)}}
  .kpi-card .lbl{{font-size:.78rem;color:#666;margin-top:4px}}
  .charts-grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px}}
  @media(max-width:768px){{.charts-grid{{grid-template-columns:1fr}}}}
  .card{{background:var(--card);border-radius:12px;padding:20px;box-shadow:0 2px 12px rgba(0,0,0,.07);margin-bottom:20px}}
  .card h3{{font-size:.95rem;color:var(--dark-blue);margin-bottom:16px;font-weight:600;
            padding-bottom:8px;border-bottom:2px solid var(--light-blue)}}
  .full-width{{grid-column:1/-1}}
  table{{width:100%;border-collapse:collapse;font-size:.82rem}}
  th{{background:var(--dark-blue);color:white;padding:9px 10px;text-align:left;font-weight:600}}
  td{{padding:8px 10px;border-bottom:1px solid #eee}}
  tr:nth-child(even) td{{background:var(--light-blue)}}
  tr:hover td{{background:#e8efff}}
  .badge{{display:inline-block;padding:2px 8px;border-radius:10px;color:white;
          font-size:.75rem;font-weight:600;text-shadow:0 1px 1px rgba(0,0,0,.3)}}
  .progress-bar{{height:8px;background:#e0e7ff;border-radius:4px;overflow:hidden}}
  .progress-fill{{height:100%;border-radius:4px}}
  .section-title{{font-size:1.1rem;font-weight:700;color:var(--dark-blue);
                  margin:28px 0 16px;padding-left:12px;border-left:4px solid var(--accent)}}
  footer{{text-align:center;font-size:.75rem;color:#888;padding:24px;margin-top:32px}}
</style>
</head>
<body>
<header>
  <div>
    <h1>📊 Informe de Conteo Vehicular</h1>
    <p>{site_name} &mdash; {road_name}</p>
    <p>Sistema: <strong>{system}</strong> &nbsp;|&nbsp; Dirección: <strong>{direction}</strong>
       &nbsp;|&nbsp; Cámara: <strong>{camera_id or video_source}</strong></p>
  </div>
  <div class="badge-header">Generado<br><strong>{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</strong></div>
</header>
<div class="container">
  <div class="kpi-grid">
    <div class="kpi-card" style="border-color:#F5B700"><div class="val">{grand_total:,}</div><div class="lbl">Total Vehículos</div></div>
    <div class="kpi-card" style="border-color:#00C853"><div class="val">{n_lanes}</div><div class="lbl">Carriles Monitoreados</div></div>
    <div class="kpi-card" style="border-color:#2E75B6"><div class="val">{len(all_cats)}</div><div class="lbl">Categorías ({system})</div></div>
    <div class="kpi-card" style="border-color:#E53935"><div class="val">{avg_speed}</div><div class="lbl">Velocidad Prom. (km/h)</div></div>
    <div class="kpi-card" style="border-color:#9C27B0"><div class="val">{max_speed}</div><div class="lbl">Velocidad Máx. (km/h)</div></div>
    <div class="kpi-card" style="border-color:#FF6F00"><div class="val">{len(events):,}</div><div class="lbl">Eventos Registrados</div></div>
  </div>
  <div class="charts-grid">
    <div class="card"><h3>Distribución por Categoría</h3><canvas id="pieChart" height="250"></canvas></div>
    <div class="card"><h3>Conteo por Carril y Categoría</h3><canvas id="laneChart" height="250"></canvas></div>
  </div>
  <div class="card full-width"><h3>Flujo Vehicular por Minuto</h3><canvas id="flowChart" height="120"></canvas></div>
  <p class="section-title">Detalle por Categoría</p>
  <div class="card">
    <table>
      <thead><tr><th>Categoría</th><th>Descripción</th><th>Conteo</th><th>% Total</th><th>Distribución</th></tr></thead>
      <tbody>
"""
    for cat in all_cats:
        cnt=summary.get(cat,0); pct=(cnt/grand_total*100) if grand_total else 0
        hex_=cats_sys.get(cat,{}).get("hex","#888")
        html+=(f"<tr><td><span class='badge' style='background:{hex_}'>{cat}</span></td>"
               f"<td>{cat_name(cat,system)}</td><td><strong>{cnt:,}</strong></td><td>{pct:.1f}%</td>"
               f"<td><div class='progress-bar'><div class='progress-fill' "
               f"style='width:{pct}%;background:{hex_}'></div></div></td></tr>")
    html+=(f"</tbody></table></div>"
           f"<p class='section-title'>Registro de Eventos (últimos {len(recent_events)})</p>"
           f"<div class='card' style='overflow-x:auto'><table><thead><tr>"
           f"<th>Hora</th><th>Frame</th><th>Track ID</th><th>Categoría</th>"
           f"<th>Carril</th><th>Tipo YOLO</th><th>Dirección</th><th>Vel. km/h</th>"
           f"</tr></thead><tbody>{rows_html}</tbody></table></div>")
    html+=(f"</div>\n<footer>Informe generado automáticamente · {site_name} · {road_name}<br>"
           f"Sistema {system} · Cámara {camera_id or video_source} · "
           f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</footer>\n"
           f"""<script>
new Chart(document.getElementById('pieChart'),{{type:'doughnut',data:{{
  labels:{cat_names_json},datasets:[{{data:Object.values({by_cat_json}),
  backgroundColor:{cat_colors_json},borderWidth:2}}]}},
  options:{{responsive:true,plugins:{{legend:{{position:'right'}}}}}}
}});
new Chart(document.getElementById('laneChart'),{{type:'bar',data:{{
  labels:{lane_labels_json},datasets:{lane_ds_json}}},
  options:{{responsive:true,scales:{{x:{{stacked:true}},y:{{stacked:true,beginAtZero:true}}}},
  plugins:{{legend:{{position:'bottom'}}}}}}
}});
new Chart(document.getElementById('flowChart'),{{type:'line',data:{{
  labels:{flow_t_json},datasets:[{{label:'Vehículos/minuto',data:{flow_v_json},
  borderColor:'#2E75B6',backgroundColor:'rgba(46,117,182,0.15)',fill:true,tension:0.4,pointRadius:3}}]}},
  options:{{responsive:true,scales:{{y:{{beginAtZero:true,title:{{display:true,text:'Veh/min'}}}}}},
  plugins:{{legend:{{display:false}}}}}}
}});
</script></body></html>""")
    with open(path,"w",encoding="utf-8") as f: f.write(html)
    print(f"  [HTML]  → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 12 — MAIN PROCESSING LOOP
# ══════════════════════════════════════════════════════════════════════════════

def run(source,
        output_path:     Optional[str]   = None,
        show:            bool             = True,
        n_lanes:         int              = 2,
        direction:       str              = "NS",
        system:          str              = "ANI",
        report_prefix:   Optional[str]    = None,
        camera_id:       Optional[str]    = None,
        yolo_model:      str              = "yolov8n.pt",
        yolo_conf:       float            = 0.30,
        yolo_iou:        float            = 0.45,
        yolo_imgsz:      int              = 640,
        inference_skip:  int              = 1,
        px_per_meter:    float            = 12.0,
        enable_speed:    bool             = True,
        lane_mode:       str              = "auto",
        roi_top_frac:    float            = 0.45,
        count_line_fracs: List[float]     = None,
        site_name:       str              = "Peaje Niquía",
        road_name:       str              = "Autopista Medellín–Bogotá",
        operator:        str              = "",
        notes:           str              = "",
        show_heatmap:    bool             = False,
        show_wrong_way:  bool             = True,
        ):
    if count_line_fracs is None:
        count_line_fracs = [0.33, 0.50, 0.67]

    cats = CLASSIFICATION_SYSTEMS[system]
    cam_str = camera_id or "CAM-01"

    print(f"\n{'='*65}")
    print(f"  ENTERPRISE VEHICLE COUNTER — {site_name}")
    print(f"  Sistema: {system}  |  Dirección: {direction}  |  Carriles: {n_lanes}")
    print(f"  Modelo: {yolo_model}  |  Conf: {yolo_conf}  |  Salto: {inference_skip}")
    print(f"{'='*65}")

    print("[INFO] Cargando modelo YOLO…")
    model = YOLO(yolo_model)

    cap = cv2.VideoCapture(source)
    if not cap.isOpened():
        raise SystemExit(f"[ERROR] No se puede abrir: {source}")

    W   = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    H   = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    FPS = cap.get(cv2.CAP_PROP_FPS) or 25.0
    TOTAL_FRAMES = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    LINES = [int(H * f) for f in count_line_fracs]

    print(f"[INFO] Resolución: {W}×{H} @ {FPS:.0f} fps  |  Frames: {TOTAL_FRAMES}")
    print(f"[INFO] Líneas de conteo: y={LINES}")

    # Init components
    lane_mgr  = LaneManager(W, H, n_lanes, direction, mode=lane_mode,
                             roi_top_frac=roi_top_frac)
    tracker   = ByteTracker(max_misses=15, min_hits=2)
    speed_est = SpeedEstimator(FPS, px_per_meter) if enable_speed else None
    density   = DensityAnalyzer(n_lanes, FPS)
    hud       = HUDEngine(direction, system, cam_str, site_name,
                          cfg_show_speed=enable_speed, cfg_show_trails=True,
                          cfg_show_wrong_way=show_wrong_way, cfg_show_heatmap=show_heatmap)

    writer = None
    if output_path:
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        writer = cv2.VideoWriter(output_path, fourcc, FPS, (W, H))
        print(f"[INFO] Guardando video anotado → {output_path}")

    cat_counts:  Dict[str, int]           = defaultdict(int)
    lane_counts: Dict[int, Dict[str,int]] = defaultdict(lambda: defaultdict(int))
    events:       List[dict]              = []
    speed_events: List[dict]              = []
    speed_map:    Dict[int, float]        = {}

    frame_n = 0; fps_v = 0.0; fps_t = time.time(); fps_c = 0
    t0 = time.time(); last_progress = -1

    print("[INFO] Procesando… Presione Q para salir.\n")

    while True:
        ok, frame = cap.read()
        if not ok: break
        frame_n += 1; fps_c += 1

        # Progress
        if TOTAL_FRAMES > 0:
            pct = int(frame_n / TOTAL_FRAMES * 100)
            if pct % 5 == 0 and pct != last_progress:
                last_progress = pct
                elapsed = time.time() - t0
                eta = (elapsed / frame_n) * (TOTAL_FRAMES - frame_n) if frame_n else 0
                print(f"  {pct:3d}%  frame={frame_n}/{TOTAL_FRAMES}  "
                      f"ETA:{eta:.0f}s  veh:{sum(cat_counts.values())}")

        # Update lane detection
        lane_mgr.update(frame)

        # YOLO inference (with skip)
        dets: List[tuple] = []
        if frame_n % max(inference_skip, 1) == 0:
            res = model(frame, classes=list(YOLO_VEHICLE_CLASSES.keys()),
                        conf=yolo_conf, iou=yolo_iou, imgsz=yolo_imgsz, verbose=False)
            if res and res[0].boxes is not None:
                for box, cid, conf_val in zip(
                    res[0].boxes.xyxy.cpu().numpy(),
                    res[0].boxes.cls.cpu().numpy().astype(int),
                    res[0].boxes.conf.cpu().numpy(),
                ):
                    x1,y1,x2,y2 = box.astype(int)
                    bw=x2-x1; bh=y2-y1
                    if bw*bh < 900: continue
                    cx_d=(x1+x2)//2; cy_d=(y1+y2)//2
                    yolo_type = YOLO_VEHICLE_CLASSES.get(cid,"car")
                    cat_key   = classify_vehicle(yolo_type,bw,bh,system,float(conf_val))
                    dets.append((x1,y1,x2,y2,int(cid),cat_key,float(conf_val)))

        # Track
        active = tracker.step(dets, lambda cx: lane_mgr.get_lane(cx, 0))

        # Update density active sets
        lane_active: Dict[int, set] = defaultdict(set)
        for tr in active: lane_active[tr.lane].add(tr.id)
        for ln in range(1, n_lanes+1): density.update_active(ln, lane_active[ln])

        # Speed estimation
        if speed_est:
            for tr in active:
                speed_est.update(tr.id, tr.vx, tr.vy)
                spd = speed_est.get_kmh(tr.id)
                speed_map[tr.id] = spd
                if spd > 1.0 and tr.counted:
                    tr.speeds.append(spd)

        # Count crossing
        for tr in active:
            if not tr.counted and len(tr.trail) >= 2:
                prev_cy = tr.trail[-2][1]
                curr_cy = tr.trail[-1][1]
                for line_i, ly in enumerate(LINES):
                    crossed = ((prev_cy < ly <= curr_cy) or (prev_cy > ly >= curr_cy))
                    if crossed and abs(curr_cy - prev_cy) >= 4:
                        tr.counted = True; tr.counted_line = line_i
                        tr.count_ts = datetime.now().strftime("%H:%M:%S")
                        tr.count_frame = frame_n
                        cat_counts[tr.cat_key] += 1
                        lane_counts[tr.lane][tr.cat_key] += 1
                        density.record_count(tr.lane)
                        total = sum(cat_counts.values())
                        yolo_type = YOLO_VEHICLE_CLASSES.get(tr.cls_id,"vehicle")
                        spd = speed_map.get(tr.id, 0.0)
                        ev = {
                            "time":       tr.count_ts,
                            "frame":      frame_n,
                            "track_id":   tr.id,
                            "yolo_type":  yolo_type,
                            "category":   tr.cat_key,
                            "lane":       tr.lane,
                            "direction":  direction,
                            "conf":       round(tr.conf, 3),
                            "speed_kmh":  round(spd, 1),
                        }
                        events.append(ev)
                        if spd > 1.0:
                            speed_events.append(ev)
                        print(f"  [{tr.count_ts}]  {tr.cat_key:<10}  "
                              f"Carril {tr.lane}  id={tr.id:4d}  "
                              f"TOTAL={total}  {spd:.0f}km/h")
                        break

        # Draw everything
        density_stats = density.get_all_stats()
        hud.draw_all(frame, active, cats, dict(cat_counts), lane_counts,
                     density_stats, fps_v, frame_n, speed_map, LINES)
        lane_mgr.draw(frame, lane_counts)

        # FPS counter
        now = time.time()
        if now - fps_t >= 1.0:
            fps_v = fps_c / (now - fps_t); fps_c = 0; fps_t = now

        if writer: writer.write(frame)
        if show:
            cv2.imshow(f"Vehicle Counter — {site_name}  [Q=salir]", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                print("[INFO] Detenido por el usuario.")
                break

    cap.release()
    if writer: writer.release()
    cv2.destroyAllWindows()

    elapsed = time.time() - t0
    grand_total = sum(cat_counts.values())

    # Console summary
    print(f"\n{'='*55}")
    print(f"  CONTEO FINAL — {site_name}")
    print(f"  Sistema: {system}  |  Dirección: {direction}")
    print(f"{'='*55}")
    print(f"  Total vehículos : {grand_total}")
    for cat in sorted(cats.keys()):
        cnt=cat_counts.get(cat,0)
        print(f"    {cat:<10}  {cat_name(cat,system):<38}: {cnt}")
    print(f"  Frames procesados : {frame_n}")
    print(f"  Tiempo elapsed    : {elapsed:.1f} s")
    if elapsed > 0 and grand_total > 0:
        print(f"  Veh/minuto        : {grand_total/(elapsed/60):.1f}")
    print(f"\n  Por carril:")
    for lane in sorted(lane_counts.keys()):
        lt=sum(lane_counts[lane].values())
        print(f"    Carril {lane}: {lt} vehículos")
        for cat,cnt in sorted(lane_counts[lane].items()):
            if cnt: print(f"      {cat:<10}: {cnt}")
    print(f"{'='*55}\n")

    # Generate reports
    prefix   = report_prefix or "informe_vehicular"
    ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")
    src_str  = str(source)

    # CSV
    csv_path = f"{prefix}_{ts_str}_log.csv"
    if events:
        with open(csv_path,"w",newline="",encoding="utf-8") as f:
            w=csv.DictWriter(f,fieldnames=events[0].keys())
            w.writeheader(); w.writerows(events)
        print(f"  [CSV]   → {csv_path}")

    # JSON
    json_path = f"{prefix}_{ts_str}_summary.json"
    with open(json_path,"w",encoding="utf-8") as f:
        json.dump({
            "meta": {
                "system":system,"direction":direction,"n_lanes":n_lanes,
                "camera_id":cam_str,"video_source":src_str,
                "site_name":site_name,"road_name":road_name,
                "generated":datetime.now().isoformat(),
                "frames":frame_n,"elapsed_s":round(elapsed,1),
            },
            "total":grand_total,
            "by_category":dict(cat_counts),
            "by_lane":{str(k):dict(v) for k,v in lane_counts.items()},
            "density":density.get_all_stats(),
        },f,indent=2,ensure_ascii=False)
    print(f"  [JSON]  → {json_path}")

    # Excel
    excel_path = f"{prefix}_{ts_str}.xlsx"
    write_excel(excel_path, dict(cat_counts), lane_counts, events, speed_events,
                system, direction, cam_str, src_str, n_lanes, site_name, operator)

    # PDF
    pdf_path = f"{prefix}_{ts_str}.pdf"
    write_pdf(pdf_path, dict(cat_counts), lane_counts, events, speed_events,
              system, direction, cam_str, src_str, n_lanes, site_name, road_name, operator, notes)

    # HTML
    html_path = f"{prefix}_{ts_str}.html"
    write_html(html_path, dict(cat_counts), lane_counts, events, speed_events,
               system, direction, cam_str, src_str, n_lanes, site_name, road_name)

    return dict(cat_counts), dict(lane_counts)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 13 — BATCH PROCESSOR
# ══════════════════════════════════════════════════════════════════════════════

def run_batch(video_list_file: str, default_system: str = "ANI"):
    """
    Process multiple videos from a text file.
    Each line: video_path,lanes,direction,camera_id
    Example:
        /videos/cam1_NS.mp4,3,NS,CAM1-NS
        /videos/cam2_SN.mp4,2,SN,CAM2-SN
    """
    with open(video_list_file) as f:
        lines=[l.strip() for l in f if l.strip() and not l.startswith("#")]

    all_results=[]; combined: Dict[str,int]=defaultdict(int)
    for line in lines:
        parts=line.split(",")
        video     =parts[0].strip()
        lanes_n   =int(parts[1].strip()) if len(parts)>1 else 2
        direc     =parts[2].strip()      if len(parts)>2 else "NS"
        cam_id    =parts[3].strip()      if len(parts)>3 else None
        prefix    =Path(video).stem

        print(f"\n{'#'*62}\n  PROCESANDO: {video}\n{'#'*62}")
        try:
            cats,lanes_r=run(source=video,show=False,n_lanes=lanes_n,direction=direc,
                             system=default_system,report_prefix=prefix,camera_id=cam_id)
            all_results.append({"video":video,"camera":cam_id,"direction":direc,
                                 "counts":cats,"lane_counts":lanes_r})
            for cat,cnt in cats.items(): combined[cat]+=cnt
        except Exception as e:
            print(f"[ERROR] Falló {video}: {e}")

    if all_results:
        print(f"\n{'='*62}\n  RESUMEN CONSOLIDADO BATCH\n{'='*62}")
        for cat,cnt in sorted(combined.items()):
            print(f"  {cat:<10}: {cnt}")
        print(f"  TOTAL : {sum(combined.values())}")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Enterprise Vehicle Counter — Colombia (ANI/INVIAS/COL)",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=__doc__,
    )

    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--video",  help="Ruta al archivo de video")
    src.add_argument("--camera", type=int, help="Índice de webcam (usualmente 0)")
    src.add_argument("--batch",  help="Archivo de texto con lista de videos para procesar")

    ap.add_argument("--output",         default=None,          help="Ruta del video anotado de salida")
    ap.add_argument("--lanes",          type=int, default=2,   help="Número de carriles (default: 2)")
    ap.add_argument("--direction",      choices=["NS","SN"],   default="NS",
                    help="NS = Norte→Sur  |  SN = Sur→Norte")
    ap.add_argument("--classification", choices=["ANI","INVIAS","COL"], default="ANI",
                    help="Sistema de clasificación vehicular")
    ap.add_argument("--report",         default="informe_vehicular",
                    help="Prefijo para archivos de reporte")
    ap.add_argument("--camera-id",      default=None,          help="ID de cámara (ej: CAM1-NS)")
    ap.add_argument("--model",          default="yolov8n.pt",  help="Modelo YOLOv8 (n/s/m/l/x)")
    ap.add_argument("--conf",           type=float, default=0.30, help="Confianza mínima YOLO")
    ap.add_argument("--skip",           type=int,   default=1,  help="Procesar cada N fotogramas")
    ap.add_argument("--px-per-meter",   type=float, default=12.0, help="Píxeles por metro (calibración)")
    ap.add_argument("--lane-mode",      choices=["auto","uniform","manual"], default="auto",
                    help="Modo de detección de carriles")
    ap.add_argument("--site",           default="Peaje Niquía", help="Nombre del sitio")
    ap.add_argument("--road",           default="Autopista Medellín–Bogotá", help="Nombre de la vía")
    ap.add_argument("--operator",       default="",             help="Nombre del operador")
    ap.add_argument("--notes",          default="",             help="Notas adicionales para el reporte")
    ap.add_argument("--no-preview",     action="store_true",    help="Deshabilitar ventana de previsualización")
    ap.add_argument("--no-speed",       action="store_true",    help="Deshabilitar estimación de velocidad")
    ap.add_argument("--heatmap",        action="store_true",    help="Mostrar mapa de calor de trayectorias")
    ap.add_argument("--no-wrong-way",   action="store_true",    help="Deshabilitar alerta de contraflujo")

    args = ap.parse_args()

    if args.batch:
        run_batch(args.batch, default_system=args.classification)
    else:
        source = args.video if args.video else args.camera
        run(
            source         = source,
            output_path    = args.output,
            show           = not args.no_preview,
            n_lanes        = args.lanes,
            direction      = args.direction,
            system         = args.classification,
            report_prefix  = args.report,
            camera_id      = args.camera_id,
            yolo_model     = args.model,
            yolo_conf      = args.conf,
            inference_skip = args.skip,
            px_per_meter   = args.px_per_meter,
            enable_speed   = not args.no_speed,
            lane_mode      = args.lane_mode,
            site_name      = args.site,
            road_name      = args.road,
            operator       = args.operator,
            notes          = args.notes,
            show_heatmap   = args.heatmap,
            show_wrong_way = not args.no_wrong_way,
        )
