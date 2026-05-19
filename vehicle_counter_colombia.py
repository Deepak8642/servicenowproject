"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         VEHICLE COUNTER — NIQUÍA TOLL STATION — FULL IMPLEMENTATION         ║
║  Supports: ANI (CAT I–VII)  |  INVIAS (CAT I–V)  |  Colombian (C1–C5)      ║
║  Features: Lane detection · Per-lane counts · CSV/Excel/PDF reports          ║
║            Dual-carriageway support · Expert report generation               ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALL:
    pip install ultralytics opencv-python numpy openpyxl reportlab

RUN (basic):
    python vehicle_counter_colombia.py --video footage.mp4

RUN (full options):
    python vehicle_counter_colombia.py \\
        --video footage.mp4 \\
        --output annotated.mp4 \\
        --lanes 3 \\
        --direction NS \\
        --classification ANI \\
        --report my_report \\
        --camera-id CAM1 \\
        --no-preview

CLASSIFICATION OPTIONS:
    --classification ANI      →  CAT I to CAT VII  (default, preferred)
    --classification INVIAS   →  CAT I to CAT V
    --classification COL      →  C1 to C5 (Colombian standard)

DIRECTION OPTIONS:
    --direction NS   →  North-South  (3 lanes, toward Medellín)
    --direction SN   →  South-North  (2 lanes, toward Copacabana)
"""

import argparse
import csv
import json
import os
import sys
import time
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np

# ── Optional imports (reports) ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAVE_EXCEL = True
except ImportError:
    HAVE_EXCEL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAVE_PDF = True
except ImportError:
    HAVE_PDF = False

try:
    from ultralytics import YOLO
except ImportError:
    raise SystemExit("ERROR: Run:  pip install ultralytics opencv-python numpy")


# ══════════════════════════════════════════════════════════════════════════════
#  CLASSIFICATION SYSTEMS
# ══════════════════════════════════════════════════════════════════════════════

# YOLO classes → vehicle type
YOLO_VEHICLE_CLASSES = {2: "car", 3: "motorcycle", 5: "bus", 7: "truck"}

# Approximate vehicle dimensions in metres (length × width)
# Used to estimate category from bounding box aspect ratio + size
# (pixel-based heuristic; improves with known camera calibration)

# ─────────────────────────────────────────────────
#  ANI Classification (CAT I – VII)
# ─────────────────────────────────────────────────
ANI_CATEGORIES = {
    "CAT I":   {
        "label": "CAT I",
        "name":  "Motocicletas y bicicletas",
        "yolo":  ["motorcycle"],
        "axles": 2,
        "color": (180, 120, 255),
    },
    "CAT II":  {
        "label": "CAT II",
        "name":  "Automóviles, camperos, camionetas y microbuses",
        "yolo":  ["car"],
        "axles": 2,
        "color": (0, 220, 80),
    },
    "CAT III": {
        "label": "CAT III",
        "name":  "Buses",
        "yolo":  ["bus"],
        "axles": 2,
        "color": (0, 160, 255),
    },
    "CAT IV":  {
        "label": "CAT IV",
        "name":  "Camiones de 2 ejes",
        "yolo":  ["truck"],
        "axles": 2,
        "color": (255, 200, 0),
    },
    "CAT V":   {
        "label": "CAT V",
        "name":  "Camiones de 3 ejes",
        "yolo":  ["truck"],
        "axles": 3,
        "color": (255, 120, 0),
    },
    "CAT VI":  {
        "label": "CAT VI",
        "name":  "Camiones de 4 ejes",
        "yolo":  ["truck"],
        "axles": 4,
        "color": (220, 50, 50),
    },
    "CAT VII": {
        "label": "CAT VII",
        "name":  "Camiones de 5 o más ejes",
        "yolo":  ["truck"],
        "axles": 5,
        "color": (140, 0, 0),
    },
}

# ─────────────────────────────────────────────────
#  INVIAS Classification (CAT I – V)
# ─────────────────────────────────────────────────
INVIAS_CATEGORIES = {
    "CAT I":  {
        "label": "CAT I",
        "name":  "Motos",
        "yolo":  ["motorcycle"],
        "color": (180, 120, 255),
    },
    "CAT II": {
        "label": "CAT II",
        "name":  "Automóviles y camperos",
        "yolo":  ["car"],
        "color": (0, 220, 80),
    },
    "CAT III":{
        "label": "CAT III",
        "name":  "Buses y busetas",
        "yolo":  ["bus"],
        "color": (0, 160, 255),
    },
    "CAT IV": {
        "label": "CAT IV",
        "name":  "Camiones livianos (2 ejes)",
        "yolo":  ["truck"],
        "color": (255, 200, 0),
    },
    "CAT V":  {
        "label": "CAT V",
        "name":  "Camiones pesados (3+ ejes)",
        "yolo":  ["truck"],
        "color": (220, 50, 50),
    },
}

# ─────────────────────────────────────────────────
#  Colombian Standard Classification (C1 – C5)
# ─────────────────────────────────────────────────
COL_CATEGORIES = {
    "C1":  {
        "label": "C1",
        "name":  "Automóviles, microbuses y motos",
        "yolo":  ["car", "motorcycle"],
        "color": (0, 220, 80),
    },
    "C2P": {
        "label": "C2P",
        "name":  "Camión 2 ejes pequeño",
        "yolo":  ["truck"],
        "color": (255, 200, 0),
    },
    "C2G": {
        "label": "C2G",
        "name":  "Camión 2 ejes grande / bus",
        "yolo":  ["bus", "truck"],
        "color": (0, 160, 255),
    },
    "C3":  {
        "label": "C3",
        "name":  "Camión 3 ejes",
        "yolo":  ["truck"],
        "color": (255, 120, 0),
    },
    "C4":  {
        "label": "C4",
        "name":  "Camión 4 ejes",
        "yolo":  ["truck"],
        "color": (220, 50, 50),
    },
    "C5":  {
        "label": "C5",
        "name":  "Camión 5+ ejes",
        "yolo":  ["truck"],
        "color": (140, 0, 0),
    },
}

CLASSIFICATION_SYSTEMS = {
    "ANI":    ANI_CATEGORIES,
    "INVIAS": INVIAS_CATEGORIES,
    "COL":    COL_CATEGORIES,
}

# Map YOLO type + bounding-box area → category key
# (heuristic; refined by aspect ratio for trucks)
def classify_vehicle(yolo_type: str, box_w: int, box_h: int,
                     system: str = "ANI") -> str:
    cats = CLASSIFICATION_SYSTEMS[system]
    area = box_w * box_h
    aspect = box_w / max(box_h, 1)

    if system == "ANI":
        if yolo_type == "motorcycle":
            return "CAT I"
        if yolo_type == "car":
            return "CAT II"
        if yolo_type == "bus":
            return "CAT III"
        if yolo_type == "truck":
            # Heuristic: longer trucks → more axles
            if aspect > 3.5 or area > 80000:
                return "CAT VII"
            if aspect > 2.8 or area > 55000:
                return "CAT VI"
            if aspect > 2.2 or area > 35000:
                return "CAT V"
            return "CAT IV"

    elif system == "INVIAS":
        if yolo_type == "motorcycle":
            return "CAT I"
        if yolo_type == "car":
            return "CAT II"
        if yolo_type == "bus":
            return "CAT III"
        if yolo_type == "truck":
            if aspect > 2.5 or area > 40000:
                return "CAT V"
            return "CAT IV"

    elif system == "COL":
        if yolo_type in ("car", "motorcycle"):
            return "C1"
        if yolo_type == "bus":
            return "C2G"
        if yolo_type == "truck":
            if aspect > 3.5 or area > 80000:
                return "C5"
            if aspect > 2.8 or area > 55000:
                return "C4"
            if aspect > 2.2 or area > 35000:
                return "C3"
            if aspect > 1.5:
                return "C2G"
            return "C2P"

    # fallback
    for key, cat in cats.items():
        if yolo_type in cat["yolo"]:
            return key
    return list(cats.keys())[0]


# ══════════════════════════════════════════════════════════════════════════════
#  LANE MANAGER
# ══════════════════════════════════════════════════════════════════════════════

class LaneManager:
    """Divides frame width into N equal lanes and assigns vehicles to lanes."""

    def __init__(self, frame_width: int, frame_height: int,
                 n_lanes: int, direction: str = "NS"):
        self.W         = frame_width
        self.H         = frame_height
        self.n_lanes   = n_lanes
        self.direction = direction
        self.lane_w    = frame_width / n_lanes
        # Boundaries: list of x coordinates separating lanes
        self.boundaries = [int(i * self.lane_w) for i in range(1, n_lanes)]

    def get_lane(self, cx: int) -> int:
        """Return 1-indexed lane number for a centroid x."""
        for i, bx in enumerate(self.boundaries):
            if cx < bx:
                return i + 1
        return self.n_lanes

    def draw(self, frame: np.ndarray):
        """Draw lane dividers and labels onto frame."""
        for bx in self.boundaries:
            cv2.line(frame, (bx, 0), (bx, self.H),
                     (255, 255, 0), 1, cv2.LINE_AA)
        lane_w = self.W // self.n_lanes
        for i in range(self.n_lanes):
            lx = i * lane_w + lane_w // 2
            label = f"Lane {i+1}"
            cv2.putText(frame, label, (lx - 28, 20),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55,
                        (255, 255, 0), 1, cv2.LINE_AA)

    def lane_label(self, lane_n: int) -> str:
        if self.direction == "NS":
            return f"Carril {lane_n} N→S"
        return f"Carril {lane_n} S→N"


# ══════════════════════════════════════════════════════════════════════════════
#  IoU TRACKER
# ══════════════════════════════════════════════════════════════════════════════

class Track:
    _nid = 1

    def __init__(self, box, cls_id, cat_key, lane):
        self.id      = Track._nid; Track._nid += 1
        self.box     = np.array(box, dtype=float)
        self.cls_id  = cls_id
        self.cat_key = cat_key
        self.lane    = lane
        self.age     = 1
        self.misses  = 0
        self.counted = False
        self.trail   = deque(maxlen=50)
        cx = int((box[0] + box[2]) / 2)
        cy = int((box[1] + box[3]) / 2)
        self.trail.append((cx, cy))

    def update(self, box, cat_key, lane):
        self.box     = 0.6 * np.array(box) + 0.4 * self.box
        self.cat_key = cat_key
        self.lane    = lane
        self.age    += 1
        self.misses  = 0
        cx = int((self.box[0] + self.box[2]) / 2)
        cy = int((self.box[1] + self.box[3]) / 2)
        self.trail.append((cx, cy))

    @property
    def cx(self): return int((self.box[0] + self.box[2]) / 2)
    @property
    def cy(self): return int((self.box[1] + self.box[3]) / 2)
    @property
    def bw(self): return int(self.box[2] - self.box[0])
    @property
    def bh(self): return int(self.box[3] - self.box[1])


def _iou(a, b):
    ix1 = max(a[0], b[0]); iy1 = max(a[1], b[1])
    ix2 = min(a[2], b[2]); iy2 = min(a[3], b[3])
    inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
    if not inter:
        return 0.0
    ua = ((a[2]-a[0])*(a[3]-a[1]) + (b[2]-b[0])*(b[3]-b[1]) - inter)
    return inter / ua if ua else 0.0


class Tracker:
    def __init__(self):
        self.tracks = []

    def step(self, dets):
        used_t = set(); used_d = set()
        for di, d in enumerate(dets):
            best_iou, best_ti = 0.20, -1
            for ti, t in enumerate(self.tracks):
                if ti in used_t:
                    continue
                s = _iou(d[:4], t.box)
                if s > best_iou:
                    best_iou, best_ti = s, ti
            if best_ti >= 0:
                self.tracks[best_ti].update(d[:4], d[5], d[6])
                used_t.add(best_ti); used_d.add(di)
        for di, d in enumerate(dets):
            if di not in used_d:
                self.tracks.append(Track(d[:4], d[4], d[5], d[6]))
        for ti, t in enumerate(self.tracks):
            if ti not in used_t:
                t.misses += 1
        self.tracks = [t for t in self.tracks if t.misses <= 12]
        return [t for t in self.tracks if t.age >= 2]


# ══════════════════════════════════════════════════════════════════════════════
#  DRAWING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def put_label(img, txt, pos, bg, fg=(255, 255, 255), sc=0.46, th=1):
    f = cv2.FONT_HERSHEY_SIMPLEX
    (tw, th2), bl = cv2.getTextSize(txt, f, sc, th)
    x, y = pos
    cv2.rectangle(img, (x-2, y-th2-3), (x+tw+2, y+bl+1), bg, -1)
    cv2.putText(img, txt, (x, y), f, sc, fg, th, cv2.LINE_AA)


def draw_hud(frame, lane_counts, cat_counts, fps, count_lines,
             system, cats, direction):
    H, W = frame.shape[:2]
    total = sum(cat_counts.values())

    # Draw counting lines
    for i, ly in enumerate(count_lines):
        thick = 2 if i == 1 else 1
        col   = (0, 255, 255) if i == 1 else (0, 180, 180)
        cv2.line(frame, (0, ly), (W, ly), col, thick, cv2.LINE_AA)
    cv2.putText(frame, "LINEAS DE CONTEO",
                (W // 2 - 80, count_lines[1] - 8),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1, cv2.LINE_AA)

    # ── Left panel: category counts ──────────────────────────────────────
    rows   = 3 + len(cats)
    ph     = 18 + rows * 24
    ov     = frame.copy()
    cv2.rectangle(ov, (5, 5), (250, ph), (10, 10, 10), -1)
    cv2.addWeighted(ov, 0.65, frame, 0.35, 0, frame)

    y = 28
    cv2.putText(frame, f"TOTAL: {total}", (12, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.75,
                (255, 255, 255), 2, cv2.LINE_AA)
    y += 26
    cv2.putText(frame, f"Sistema: {system}", (12, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.44,
                (180, 255, 180), 1, cv2.LINE_AA)
    y += 20
    cv2.putText(frame, f"Dirección: {direction}", (12, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.44,
                (180, 255, 180), 1, cv2.LINE_AA)
    y += 22
    for key in cats:
        cnt   = cat_counts.get(key, 0)
        color = cats[key]["color"]
        bgr   = (int(color[2]), int(color[1]), int(color[0]))
        cv2.putText(frame, f"  {key:<10}: {cnt}", (12, y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.46,
                    bgr, 1, cv2.LINE_AA)
        y += 22

    # ── Right panel: per-lane counts ─────────────────────────────────────
    lane_keys = sorted(lane_counts.keys())
    rw  = 180
    rh  = 18 + len(lane_keys) * 24 + 24
    rx  = W - rw - 5
    ov2 = frame.copy()
    cv2.rectangle(ov2, (rx, 5), (W - 5, rh), (10, 10, 10), -1)
    cv2.addWeighted(ov2, 0.65, frame, 0.35, 0, frame)

    ry = 28
    cv2.putText(frame, "POR CARRIL", (rx + 8, ry),
                cv2.FONT_HERSHEY_SIMPLEX, 0.55,
                (255, 255, 255), 1, cv2.LINE_AA)
    ry += 22
    for lk in lane_keys:
        ltotal = sum(lane_counts[lk].values())
        cv2.putText(frame, f"  {lk}: {ltotal}", (rx + 8, ry),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.46,
                    (160, 220, 255), 1, cv2.LINE_AA)
        ry += 22

    # ── Bottom: FPS + timestamp ──────────────────────────────────────────
    ts = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    put_label(frame, ts,       (W // 2 - 90, H - 12),
              bg=(20, 20, 20), sc=0.46)
    put_label(frame, f"FPS: {fps:.1f}", (12, H - 12),
              bg=(20, 20, 20), sc=0.42)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(path, summary, lane_counts, events,
                system, direction, camera_id, video_source):
    if not HAVE_EXCEL:
        print("[WARN] openpyxl not installed — skipping Excel report.")
        return

    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    hdr_font   = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13)
    bold_font  = Font(bold=True)
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")

    CAT_COLORS = {
        "CAT I":   "C9B3FF", "CAT II":  "80FF9B", "CAT III": "80C8FF",
        "CAT IV":  "FFE680", "CAT V":   "FFB366", "CAT VI":  "FF8080",
        "CAT VII": "FF6666",
        "C1":  "80FF9B", "C2P": "FFE680", "C2G": "80C8FF",
        "C3":  "FFB366", "C4":  "FF8080", "C5":  "FF6666",
    }

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"

    ws.merge_cells("A1:F1")
    ws["A1"] = "INFORME DE CONTEO VEHICULAR – PEAJE NIQUÍA"
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A1"].alignment = center

    meta = [
        ("Cámara / Video:", camera_id or video_source),
        ("Sistema de clasificación:", system),
        ("Dirección:", direction),
        ("Generado:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for r, (k, v) in enumerate(meta, start=3):
        ws.cell(r, 1, k).font = bold_font
        ws.cell(r, 2, v)

    # Category totals table
    ws["A8"] = "Categoría";    ws["A8"].font = hdr_font; ws["A8"].fill = hdr_fill
    ws["B8"] = "Descripción";  ws["B8"].font = hdr_font; ws["B8"].fill = hdr_fill
    ws["C8"] = "Total";        ws["C8"].font = hdr_font; ws["C8"].fill = hdr_fill
    ws["D8"] = "% del total";  ws["D8"].font = hdr_font; ws["D8"].fill = hdr_fill

    grand_total = sum(summary.values())
    for r, (cat, cnt) in enumerate(summary.items(), start=9):
        ws.cell(r, 1, cat).font   = bold_font
        ws.cell(r, 2, _cat_name(cat, system))
        ws.cell(r, 3, cnt).alignment = center
        pct = (cnt / grand_total * 100) if grand_total else 0
        ws.cell(r, 4, f"{pct:.1f}%").alignment = center
        fill_color = CAT_COLORS.get(cat, "FFFFFF")
        for c in range(1, 5):
            ws.cell(r, c).fill   = PatternFill("solid", fgColor=fill_color)
            ws.cell(r, c).border = border

    ws.cell(9 + len(summary), 1, "TOTAL").font = bold_font
    ws.cell(9 + len(summary), 3, grand_total).font = bold_font
    ws.cell(9 + len(summary), 3).alignment = center

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14

    # ── Sheet 2: Per-lane breakdown ──────────────────────────────────────
    ws2 = wb.create_sheet("Por Carril")
    ws2["A1"] = "CONTEO POR CARRIL Y CATEGORÍA"
    ws2["A1"].font = Font(bold=True, size=13, color="1F3864")

    all_cats = sorted(summary.keys())
    lane_nums = sorted(lane_counts.keys())

    # Header row
    ws2.cell(3, 1, "Carril").font = hdr_font
    ws2.cell(3, 1).fill = hdr_fill
    for ci, cat in enumerate(all_cats, start=2):
        ws2.cell(3, ci, cat).font = hdr_font
        ws2.cell(3, ci).fill = hdr_fill
        ws2.cell(3, ci).alignment = center
    ws2.cell(3, len(all_cats) + 2, "TOTAL").font = hdr_font
    ws2.cell(3, len(all_cats) + 2).fill = hdr_fill
    ws2.cell(3, len(all_cats) + 2).alignment = center

    for ri, lane in enumerate(lane_nums, start=4):
        ws2.cell(ri, 1, f"Carril {lane}").font = bold_font
        row_total = 0
        for ci, cat in enumerate(all_cats, start=2):
            cnt = lane_counts[lane].get(cat, 0)
            ws2.cell(ri, ci, cnt).alignment = center
            row_total += cnt
            fill_color = CAT_COLORS.get(cat, "F0F0F0")
            ws2.cell(ri, ci).fill = PatternFill("solid", fgColor=fill_color)
        ws2.cell(ri, len(all_cats) + 2, row_total).font = bold_font
        ws2.cell(ri, len(all_cats) + 2).alignment = center

    for c in range(1, len(all_cats) + 3):
        ws2.column_dimensions[get_column_letter(c)].width = 13

    # ── Sheet 3: Raw events log ──────────────────────────────────────────
    ws3 = wb.create_sheet("Registro Detallado")
    ws3["A1"] = "REGISTRO DE EVENTOS INDIVIDUALES"
    ws3["A1"].font = Font(bold=True, size=13)

    headers = ["#", "Timestamp", "Frame", "ID Track",
               "Tipo YOLO", "Categoría", "Carril", "Dirección"]
    for ci, h in enumerate(headers, start=1):
        ws3.cell(2, ci, h).font = hdr_font
        ws3.cell(2, ci).fill = hdr_fill
        ws3.cell(2, ci).alignment = center

    for ri, ev in enumerate(events, start=3):
        row = [ri - 2, ev["time"], ev["frame"], ev["track_id"],
               ev["yolo_type"], ev["category"], ev["lane"], ev["direction"]]
        for ci, val in enumerate(row, start=1):
            ws3.cell(ri, ci, val).alignment = center

    col_widths = [6, 22, 8, 10, 14, 12, 10, 16]
    for ci, w in enumerate(col_widths, start=1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 4: Hourly breakdown ────────────────────────────────────────
    if events:
        ws4 = wb.create_sheet("Resumen por Hora")
        ws4["A1"] = "VOLUMEN DE TRÁFICO POR HORA"
        ws4["A1"].font = Font(bold=True, size=13)

        hourly = defaultdict(lambda: defaultdict(int))
        for ev in events:
            hour_key = ev["time"][:5]  # "HH:MM" → group by minute for short vids
            # For multi-hour videos use: ev["time"][:2] + ":00"
            hourly[hour_key][ev["category"]] += 1

        ws4.cell(3, 1, "Hora/Min").font = hdr_font
        ws4.cell(3, 1).fill = hdr_fill
        for ci, cat in enumerate(all_cats, start=2):
            ws4.cell(3, ci, cat).font = hdr_font
            ws4.cell(3, ci).fill = hdr_fill
            ws4.cell(3, ci).alignment = center
        ws4.cell(3, len(all_cats) + 2, "TOTAL").font = hdr_font
        ws4.cell(3, len(all_cats) + 2).fill = hdr_fill

        for ri, (hour, cats_cnt) in enumerate(sorted(hourly.items()), start=4):
            ws4.cell(ri, 1, hour).font = bold_font
            row_total = 0
            for ci, cat in enumerate(all_cats, start=2):
                cnt = cats_cnt.get(cat, 0)
                ws4.cell(ri, ci, cnt).alignment = center
                row_total += cnt
            ws4.cell(ri, len(all_cats) + 2, row_total).font = bold_font
            ws4.cell(ri, len(all_cats) + 2).alignment = center

        for c in range(1, len(all_cats) + 3):
            ws4.column_dimensions[get_column_letter(c)].width = 13

    wb.save(path)
    print(f"  [Excel] → {path}")


def _cat_name(cat, system):
    cats = CLASSIFICATION_SYSTEMS.get(system, {})
    return cats.get(cat, {}).get("name", cat)


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPERT REPORT
# ══════════════════════════════════════════════════════════════════════════════

def write_pdf(path, summary, lane_counts, events,
              system, direction, camera_id, video_source, n_lanes):
    if not HAVE_PDF:
        print("[WARN] reportlab not installed — skipping PDF report.")
        return

    doc = SimpleDocTemplate(path, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title block
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"],
                                 fontSize=16, spaceAfter=4,
                                 textColor=colors.HexColor("#1F3864"),
                                 alignment=TA_CENTER)
    sub_style   = ParagraphStyle("SubStyle", parent=styles["Normal"],
                                 fontSize=10, spaceAfter=2,
                                 alignment=TA_CENTER,
                                 textColor=colors.HexColor("#555555"))
    body_style  = ParagraphStyle("BodyStyle", parent=styles["Normal"],
                                 fontSize=10, spaceAfter=6, leading=14)
    bold_style  = ParagraphStyle("BoldStyle", parent=styles["Normal"],
                                 fontSize=10, fontName="Helvetica-Bold")

    story.append(Paragraph("INFORME TÉCNICO PERICIAL", title_style))
    story.append(Paragraph("Conteo y Clasificación Vehicular por Carril", title_style))
    story.append(Paragraph("Peaje Niquía – Autopista Medellín-Bogotá", sub_style))
    story.append(HRFlowable(width="100%", thickness=2,
                             color=colors.HexColor("#1F3864")))
    story.append(Spacer(1, 10))

    # Metadata table
    meta_data = [
        ["Cámara / Fuente de video:", camera_id or str(video_source)],
        ["Sistema de clasificación:", system],
        ["Dirección de flujo:", direction],
        ["Número de carriles:", str(n_lanes)],
        ["Total vehículos contados:", str(sum(summary.values()))],
        ["Fecha de generación:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    meta_table = Table(meta_data, colWidths=[7*cm, 10*cm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,0), (-1,-1),
         [colors.HexColor("#EEF2FF"), colors.white]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 16))

    # ── 1. Summary by category ───────────────────────────────────────────
    story.append(Paragraph("1. Resumen por Categoría Vehicular", styles["Heading2"]))
    story.append(Spacer(1, 4))

    grand_total = sum(summary.values())
    cat_rows = [["Categoría", "Descripción", "Conteo", "% del Total"]]
    for cat, cnt in summary.items():
        pct = (cnt / grand_total * 100) if grand_total else 0
        cat_rows.append([cat, _cat_name(cat, system), str(cnt), f"{pct:.1f}%"])
    cat_rows.append(["TOTAL", "", str(grand_total), "100%"])

    cat_table = Table(cat_rows, colWidths=[2.5*cm, 9*cm, 2.5*cm, 3*cm])
    cat_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  colors.HexColor("#1F3864")),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTNAME",      (0,-1),  (-1,-1),  "Helvetica-Bold"),
        ("BACKGROUND",    (0,-1),  (-1,-1),  colors.HexColor("#D0D8EE")),
        ("ROWBACKGROUNDS",(0, 1),  (-1,-2),
         [colors.HexColor("#F5F7FF"), colors.white]),
        ("GRID",          (0, 0),  (-1,-1),  0.5, colors.grey),
        ("ALIGN",         (2, 0),  (-1,-1),  "CENTER"),
        ("FONTSIZE",      (0, 0),  (-1,-1),  10),
        ("TOPPADDING",    (0, 0),  (-1,-1),  5),
        ("BOTTOMPADDING", (0, 0),  (-1,-1),  5),
    ]))
    story.append(cat_table)
    story.append(Spacer(1, 16))

    # ── 2. Per-lane breakdown ────────────────────────────────────────────
    story.append(Paragraph("2. Conteo por Carril y Categoría", styles["Heading2"]))
    story.append(Spacer(1, 4))

    all_cats  = sorted(summary.keys())
    lane_nums = sorted(lane_counts.keys())

    lane_header = ["Carril"] + all_cats + ["TOTAL"]
    lane_rows   = [lane_header]
    for lane in lane_nums:
        row = [f"Carril {lane}"]
        row_total = 0
        for cat in all_cats:
            cnt = lane_counts[lane].get(cat, 0)
            row.append(str(cnt))
            row_total += cnt
        row.append(str(row_total))
        lane_rows.append(row)

    # Totals row
    totals_row = ["TOTAL"]
    for cat in all_cats:
        totals_row.append(str(sum(lane_counts[l].get(cat, 0) for l in lane_nums)))
    totals_row.append(str(grand_total))
    lane_rows.append(totals_row)

    col_w = [2.5*cm] + [14.5*cm / max(len(all_cats)+1, 1)] * (len(all_cats)+1)
    lane_table = Table(lane_rows, colWidths=col_w)
    lane_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  colors.HexColor("#1F3864")),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTNAME",      (0,-1),  (-1,-1),  "Helvetica-Bold"),
        ("BACKGROUND",    (0,-1),  (-1,-1),  colors.HexColor("#D0D8EE")),
        ("ROWBACKGROUNDS",(0, 1),  (-1,-2),
         [colors.HexColor("#F5F7FF"), colors.white]),
        ("GRID",          (0, 0),  (-1,-1),  0.5, colors.grey),
        ("ALIGN",         (1, 0),  (-1,-1),  "CENTER"),
        ("FONTSIZE",      (0, 0),  (-1,-1), 9),
        ("TOPPADDING",    (0, 0),  (-1,-1),  5),
        ("BOTTOMPADDING", (0, 0),  (-1,-1),  5),
    ]))
    story.append(lane_table)
    story.append(Spacer(1, 16))

    # ── 3. Methodology ──────────────────────────────────────────────────
    story.append(Paragraph("3. Metodología", styles["Heading2"]))
    methodology = (
        f"El conteo y clasificación vehicular se realizó mediante análisis de "
        f"video utilizando el modelo de detección de objetos YOLOv8 (You Only Look Once), "
        f"complementado con un algoritmo de seguimiento de trayectorias basado en "
        f"Intersección sobre Unión (IoU). Los vehículos fueron clasificados según "
        f"el sistema <b>{system}</b> a partir de la tipología detectada y las dimensiones "
        f"relativas del vehículo en el plano de imagen. "
        f"El conteo se realizó mediante tres líneas virtuales transversales posicionadas "
        f"al 33%, 50% y 67% de la altura del fotograma, registrando el cruce de cada "
        f"trayectoria una única vez para evitar conteos duplicados. "
        f"Los carriles fueron definidos como franjas verticales equidistantes a lo largo "
        f"del ancho del fotograma, asignando cada vehículo al carril correspondiente "
        f"según la coordenada x de su centroide al momento del cruce."
    )
    story.append(Paragraph(methodology, body_style))
    story.append(Spacer(1, 16))

    # ── 4. Limitations ──────────────────────────────────────────────────
    story.append(Paragraph("4. Limitaciones y Consideraciones", styles["Heading2"]))
    limitations = [
        "La clasificación de camiones pesados (categorías de múltiples ejes) se basa "
        "en heurísticas de relación de aspecto y área de la caja de detección, dado "
        "que el modelo YOLOv8 no distingue el número de ejes directamente.",
        "La precisión de asignación de carril depende del ángulo y posición de la "
        "cámara. Cámaras con perspectiva pronunciada pueden requerir calibración adicional.",
        "Vehículos parcialmente ocluidos o en situaciones de solapamiento pueden "
        "generar conteos ligeramente inferiores al real.",
        "Se recomienda validar una muestra del conteo automático contra revisión manual "
        "para certificar la precisión ante la ANI.",
    ]
    for lim in limitations:
        story.append(Paragraph(f"• {lim}", body_style))
    story.append(Spacer(1, 20))

    # Footer
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    footer_text = (
        f"Informe generado automáticamente el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
        f"| Sistema: {system} | Cámara: {camera_id or 'N/A'} | Dirección: {direction}"
    )
    story.append(Paragraph(footer_text,
                            ParagraphStyle("Footer", parent=styles["Normal"],
                                           fontSize=8, textColor=colors.grey,
                                           alignment=TA_CENTER)))

    doc.build(story)
    print(f"  [PDF]   → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN PROCESSING LOOP
# ══════════════════════════════════════════════════════════════════════════════

def run(source, output_path=None, show=True,
        n_lanes=2, direction="NS", system="ANI",
        report_prefix=None, camera_id=None):

    cats = CLASSIFICATION_SYSTEMS[system]

    print(f"\n{'='*62}")
    print(f"  VEHICLE COUNTER — PEAJE NIQUÍA")
    print(f"  Sistema: {system}  |  Dirección: {direction}  |  Carriles: {n_lanes}")
    print(f"{'='*62}")
    print("[INFO] Loading YOLOv8n… (~6 MB download on first run)")
    model = YOLO("yolov8n.pt")

    cap = cv2.VideoCapture(source)
    if not cap.isOpened():
        raise SystemExit(f"[ERROR] Cannot open: {source}")

    W   = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    H   = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    FPS = cap.get(cv2.CAP_PROP_FPS) or 25.0
    TOTAL_FRAMES = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # Three counting lines at 33%, 50%, 67% of height
    LINES = [int(H * f) for f in (0.33, 0.50, 0.67)]

    lane_mgr = LaneManager(W, H, n_lanes, direction)
    print(f"[INFO] Resolution: {W}×{H} @ {FPS:.0f} fps  |"
          f"  Total frames: {TOTAL_FRAMES}"
          f"  |  Count lines: y={LINES}")

    writer = None
    if output_path:
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        writer = cv2.VideoWriter(output_path, fourcc, FPS, (W, H))
        print(f"[INFO] Saving annotated video → {output_path}")

    tracker   = Tracker()
    # cat_counts[category_key] = int
    cat_counts  = defaultdict(int)
    # lane_counts[lane_n][category_key] = int
    lane_counts = defaultdict(lambda: defaultdict(int))
    events      = []          # raw detection log

    frame_n = 0
    fps_v = 0.0; fps_t = time.time(); fps_c = 0
    t0    = time.time()
    last_progress = -1

    print("[INFO] Processing… Press Q to quit.\n")

    while True:
        ok, frame = cap.read()
        if not ok:
            break
        frame_n += 1; fps_c += 1

        # Progress indicator (every 5%)
        if TOTAL_FRAMES > 0:
            pct = int(frame_n / TOTAL_FRAMES * 100)
            if pct % 5 == 0 and pct != last_progress:
                last_progress = pct
                elapsed = time.time() - t0
                eta = (elapsed / frame_n) * (TOTAL_FRAMES - frame_n) if frame_n else 0
                print(f"  Progress: {pct:3d}%  frame={frame_n}/{TOTAL_FRAMES}"
                      f"  ETA: {eta:.0f}s  vehicles: {sum(cat_counts.values())}")

        # ── YOLO Detection ───────────────────────────────────────────────
        res = model(frame,
                    classes=list(YOLO_VEHICLE_CLASSES.keys()),
                    conf=0.25, iou=0.45,
                    imgsz=640, verbose=False)

        dets = []
        if res and res[0].boxes is not None:
            for box, cid in zip(
                res[0].boxes.xyxy.cpu().numpy(),
                res[0].boxes.cls.cpu().numpy().astype(int)
            ):
                x1, y1, x2, y2 = box.astype(int)
                bw = x2 - x1; bh = y2 - y1
                if bw * bh < 900:           # skip tiny detections
                    continue
                cx_d = (x1 + x2) // 2
                yolo_type = YOLO_VEHICLE_CLASSES.get(cid, "car")
                cat_key   = classify_vehicle(yolo_type, bw, bh, system)
                lane      = lane_mgr.get_lane(cx_d)
                # dets: (x1,y1,x2,y2, yolo_cls_id, cat_key, lane)
                dets.append((x1, y1, x2, y2, cid, cat_key, lane))

        # ── Track ────────────────────────────────────────────────────────
        active = tracker.step(dets)

        # ── Count + Draw ─────────────────────────────────────────────────
        for tr in active:
            x1, y1, x2, y2 = [int(v) for v in tr.box]
            cat_key  = tr.cat_key
            cat_info = cats.get(cat_key, {"label": cat_key,
                                           "color": (200, 200, 200)})
            color    = cat_info["color"]
            bgr      = (int(color[2]), int(color[1]), int(color[0]))

            # Count when centroid crosses any of the 3 lines
            if not tr.counted and len(tr.trail) >= 2:
                prev_cy = tr.trail[-2][1]
                curr_cy = tr.trail[-1][1]
                for ly in LINES:
                    crossed = (
                        (prev_cy < ly <= curr_cy) or
                        (prev_cy > ly >= curr_cy)
                    )
                    if crossed and abs(curr_cy - prev_cy) >= 4:
                        tr.counted = True
                        cat_counts[cat_key]        += 1
                        lane_counts[tr.lane][cat_key] += 1
                        total = sum(cat_counts.values())
                        ts    = datetime.now().strftime("%H:%M:%S")
                        yolo_type = YOLO_VEHICLE_CLASSES.get(tr.cls_id, "vehicle")
                        events.append({
                            "time":       ts,
                            "frame":      frame_n,
                            "track_id":   tr.id,
                            "yolo_type":  yolo_type,
                            "category":   cat_key,
                            "lane":       tr.lane,
                            "direction":  direction,
                        })
                        print(f"  [{ts}]  {cat_key:<10}  "
                              f"Lane {tr.lane}  id={tr.id:4d}  "
                              f"TOTAL={total}")
                        break

            # Bounding box
            col_draw = bgr if tr.counted else (0, 140, 255)
            cv2.rectangle(frame, (x1, y1), (x2, y2), col_draw, 2)
            lbl = f"{cat_info['label']} L{tr.lane} #{tr.id}"
            put_label(frame, lbl, (x1, y1 - 2), bg=col_draw)

            # Movement trail
            pts = list(tr.trail)
            for i in range(1, len(pts)):
                cv2.line(frame, pts[i-1], pts[i], (80, 200, 255), 1)
            cv2.circle(frame, (tr.cx, tr.cy), 4, col_draw, -1)

        # ── Draw lanes + HUD ─────────────────────────────────────────────
        lane_mgr.draw(frame)
        draw_hud(frame, lane_counts, cat_counts, fps_v,
                 LINES, system, cats, direction)

        # ── FPS counter ──────────────────────────────────────────────────
        now = time.time()
        if now - fps_t >= 1.0:
            fps_v = fps_c / (now - fps_t)
            fps_c = 0
            fps_t = now

        if writer:
            writer.write(frame)
        if show:
            cv2.imshow("Vehicle Counter — Peaje Niquía  [Q=quit]", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                print("[INFO] Stopped by user.")
                break

    # ── Cleanup ──────────────────────────────────────────────────────────
    cap.release()
    if writer:
        writer.release()
    cv2.destroyAllWindows()

    elapsed    = time.time() - t0
    grand_total = sum(cat_counts.values())

    # ── Console final report ─────────────────────────────────────────────
    print(f"\n{'='*52}")
    print(f"  CONTEO FINAL DE VEHÍCULOS")
    print(f"  Sistema: {system}  |  Dirección: {direction}")
    print(f"{'='*52}")
    print(f"  Total vehículos : {grand_total}")
    for cat in sorted(cats.keys()):
        cnt = cat_counts.get(cat, 0)
        print(f"    {cat:<10}  {_cat_name(cat, system):<38}: {cnt}")
    print(f"  Frames          : {frame_n}")
    print(f"  Tiempo          : {elapsed:.1f} s")
    if elapsed > 0 and grand_total > 0:
        print(f"  Veh/minuto      : {grand_total / (elapsed / 60):.1f}")
    print(f"\n  Por carril:")
    for lane in sorted(lane_counts.keys()):
        lt = sum(lane_counts[lane].values())
        print(f"    Carril {lane}: {lt} vehículos")
        for cat, cnt in sorted(lane_counts[lane].items()):
            if cnt:
                print(f"      {cat:<10}: {cnt}")
    print(f"{'='*52}\n")

    # ── Generate reports ─────────────────────────────────────────────────
    prefix = report_prefix or "informe_vehicular"
    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")

    # CSV raw log
    csv_path = f"{prefix}_{ts_str}_log.csv"
    if events:
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=events[0].keys())
            w.writeheader(); w.writerows(events)
        print(f"  [CSV]   → {csv_path}")

    # JSON summary
    json_path = f"{prefix}_{ts_str}_summary.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "meta": {
                "system": system,
                "direction": direction,
                "n_lanes": n_lanes,
                "camera_id": camera_id,
                "video_source": str(source),
                "generated": datetime.now().isoformat(),
                "frames": frame_n,
                "elapsed_s": round(elapsed, 1),
            },
            "total": grand_total,
            "by_category": dict(cat_counts),
            "by_lane": {str(k): dict(v) for k, v in lane_counts.items()},
        }, f, indent=2, ensure_ascii=False)
    print(f"  [JSON]  → {json_path}")

    # Excel
    excel_path = f"{prefix}_{ts_str}.xlsx"
    write_excel(excel_path, dict(cat_counts), lane_counts, events,
                system, direction, camera_id, source)

    # PDF
    pdf_path = f"{prefix}_{ts_str}.pdf"
    write_pdf(pdf_path, dict(cat_counts), lane_counts, events,
              system, direction, camera_id, source, n_lanes)

    return dict(cat_counts), dict(lane_counts)


# ══════════════════════════════════════════════════════════════════════════════
#  BATCH PROCESSOR (process multiple videos automatically)
# ══════════════════════════════════════════════════════════════════════════════

def run_batch(video_list_file: str, default_system="ANI"):
    """
    Process multiple videos from a text file.
    Each line format:
        video_path,lanes,direction,camera_id
    Example:
        /videos/cam1_NS.mp4,3,NS,CAM1-NS
        /videos/cam2_SN.mp4,2,SN,CAM2-SN
    """
    with open(video_list_file) as f:
        lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]

    all_results = []
    for line in lines:
        parts = line.split(",")
        video      = parts[0].strip()
        lanes      = int(parts[1].strip()) if len(parts) > 1 else 2
        direction  = parts[2].strip()      if len(parts) > 2 else "NS"
        cam_id     = parts[3].strip()      if len(parts) > 3 else None
        prefix     = Path(video).stem

        print(f"\n{'#'*62}")
        print(f"  PROCESSING: {video}")
        print(f"{'#'*62}")

        try:
            cats, lanes_r = run(
                source        = video,
                show          = False,
                n_lanes       = lanes,
                direction     = direction,
                system        = default_system,
                report_prefix = prefix,
                camera_id     = cam_id,
            )
            all_results.append({
                "video": video,
                "camera": cam_id,
                "direction": direction,
                "counts": cats,
                "lane_counts": lanes_r,
            })
        except Exception as e:
            print(f"[ERROR] Failed to process {video}: {e}")

    # Consolidated summary
    if all_results:
        print(f"\n{'='*62}")
        print("  CONSOLIDATED BATCH SUMMARY")
        print(f"{'='*62}")
        combined = defaultdict(int)
        for r in all_results:
            for cat, cnt in r["counts"].items():
                combined[cat] += cnt
        for cat, cnt in sorted(combined.items()):
            print(f"  {cat:<10}: {cnt}")
        print(f"  TOTAL : {sum(combined.values())}")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Vehicle Counter — Peaje Niquía — Full Colombian Classification",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=__doc__,
    )

    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--video",  help="Path to video file")
    src.add_argument("--camera", type=int, help="Webcam index (usually 0)")
    src.add_argument("--batch",  help="Text file listing videos to process in batch")

    ap.add_argument("--output",  default=None,
                    help="Path for annotated output video (optional)")
    ap.add_argument("--lanes",   type=int, default=2,
                    help="Number of lanes (NS=3, SN=2 for Niquía)")
    ap.add_argument("--direction", choices=["NS", "SN"], default="NS",
                    help="NS = North→South (Medellín) | SN = South→North (Copacabana)")
    ap.add_argument("--classification", choices=["ANI", "INVIAS", "COL"],
                    default="ANI",
                    help="Vehicle classification system")
    ap.add_argument("--report",  default="informe_vehicular",
                    help="Base name for output reports (default: informe_vehicular)")
    ap.add_argument("--camera-id", default=None,
                    help="Camera identifier for reports (e.g. CAM1-NS)")
    ap.add_argument("--no-preview", action="store_true",
                    help="Disable live preview window (faster, for servers)")

    args = ap.parse_args()

    if args.batch:
        run_batch(args.batch, default_system=args.classification)
    else:
        source = args.video if args.video else args.camera
        run(
            source        = source,
            output_path   = args.output,
            show          = not args.no_preview,
            n_lanes       = args.lanes,
            direction     = args.direction,
            system        = args.classification,
            report_prefix = args.report,
            camera_id     = args.camera_id,
        )
