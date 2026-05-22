"""
╔══════════════════════════════════════════════════════════════════════════════════╗
║   NIQUÍA TOLL STATION — ENTERPRISE VEHICLE COUNTER                               ║
║   Complete Production System — Single File                                       ║
║                                                                                  ║
║   Built specifically for the 4 CCTV cameras at Niquía, Autopista Medellín–Bogotá║
║   Camera geometry tuned to actual footage:                                       ║
║     CAM1 (NS): 3-lane, overhead-left mount, strong perspective, clear lines      ║
║     CAM2 (NS): 3-lane, overhead-right, busy traffic, dashed centre line          ║
║     CAM3 (SN): 2-lane, overhead-right, evening glare top-right                  ║
║     CAM4 (SN): 2-lane, wider angle, lower mount, includes hard shoulder          ║
║                                                                                  ║
║   Classification: ANI (CAT I–VII) · INVIAS (CAT I–V) · Colombian (C1–C5)       ║
║   Detection:      YOLOv8 + Kalman Filter + ByteTrack two-stage association       ║
║   Lane lines:     White-line detector — HLS + Hough + glare suppression          ║
║                   + temporal smoothing over 15 frames                            ║
║   Reports:        CSV · JSON · Excel (5 sheets) · PDF (ANI format) · HTML        ║
║   Extras:         Speed estimation · Density/flow · Heatmap · Wrong-way alert    ║
╚══════════════════════════════════════════════════════════════════════════════════╝

INSTALL:
    pip install ultralytics opencv-python numpy openpyxl reportlab

RUN — single camera:
    python niquia_vehicle_counter_FINAL.py --video cam1.mp4 --lanes 3 --direction NS --camera-id CAM1-NS

RUN — all 4 cameras batch:
    python niquia_vehicle_counter_FINAL.py --batch niquia_batch.txt

    niquia_batch.txt (one line per camera):
        /videos/cam1_NS.mp4,3,NS,CAM1-NS
        /videos/cam2_NS.mp4,3,NS,CAM2-NS
        /videos/cam3_SN.mp4,2,SN,CAM3-SN
        /videos/cam4_SN.mp4,2,SN,CAM4-SN

ALL OPTIONS:
    --video          Path to video file
    --camera         Webcam index (0)
    --batch          Text file listing multiple videos
    --output         Save annotated video
    --lanes          Number of lanes (NS=3, SN=2)
    --direction      NS or SN
    --classification ANI | INVIAS | COL  (default: ANI)
    --report         Base name for output reports
    --camera-id      Camera ID string for reports
    --model          yolov8n.pt | yolov8s.pt | yolov8m.pt (default: yolov8s.pt)
    --conf           YOLO confidence threshold (default: 0.30)
    --skip           Process every Nth frame (default: 1)
    --px-per-meter   Pixels per metre for speed (default: 12)
    --lane-mode      auto | uniform (default: auto)
    --site           Site name for reports
    --road           Road name for reports
    --operator       Operator name
    --notes          Extra notes for PDF
    --no-preview     Headless mode
    --no-speed       Disable speed estimation
    --heatmap        Trajectory heatmap overlay
    --no-wrong-way   Disable wrong-way alert
"""

from __future__ import annotations
import argparse, csv, json, math, os, sys, time
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import cv2
import numpy as np

try:
    from ultralytics import YOLO
except ImportError:
    raise SystemExit("ERROR: pip install ultralytics opencv-python numpy")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAVE_EXCEL = True
except ImportError:
    HAVE_EXCEL = False

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, PageBreak)
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    HAVE_PDF = True
except ImportError:
    HAVE_PDF = False


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — CLASSIFICATION SYSTEMS
# ══════════════════════════════════════════════════════════════════════════════

YOLO_VEHICLE_CLASSES: Dict[int, str] = {
    2: "car", 3: "motorcycle", 5: "bus", 7: "truck",
}

ANI_CATEGORIES: Dict[str, dict] = {
    "CAT I":   {"label":"CAT I",   "name":"Motocicletas y bicicletas",
                "desc":"Vehículos de dos ruedas motorizados",
                "yolo":["motorcycle"],"axles":2,"color":(180,120,255),"hex":"#B478FF"},
    "CAT II":  {"label":"CAT II",  "name":"Automóviles, camperos, camionetas y microbuses",
                "desc":"Vehículos livianos de pasajeros hasta 8 asientos",
                "yolo":["car"],"axles":2,"color":(0,220,80),"hex":"#00DC50"},
    "CAT III": {"label":"CAT III", "name":"Buses",
                "desc":"Vehículos de transporte masivo de pasajeros",
                "yolo":["bus"],"axles":2,"color":(0,160,255),"hex":"#00A0FF"},
    "CAT IV":  {"label":"CAT IV",  "name":"Camiones de 2 ejes",
                "desc":"Vehículos de carga con dos ejes simples",
                "yolo":["truck"],"axles":2,"color":(255,200,0),"hex":"#FFC800"},
    "CAT V":   {"label":"CAT V",   "name":"Camiones de 3 ejes",
                "desc":"Vehículos de carga con tres ejes",
                "yolo":["truck"],"axles":3,"color":(255,120,0),"hex":"#FF7800"},
    "CAT VI":  {"label":"CAT VI",  "name":"Camiones de 4 ejes",
                "desc":"Vehículos de carga con cuatro ejes",
                "yolo":["truck"],"axles":4,"color":(220,50,50),"hex":"#DC3232"},
    "CAT VII": {"label":"CAT VII", "name":"Camiones de 5 o más ejes",
                "desc":"Tracto-camiones y vehículos articulados pesados",
                "yolo":["truck"],"axles":5,"color":(140,0,0),"hex":"#8C0000"},
}

INVIAS_CATEGORIES: Dict[str, dict] = {
    "CAT I":  {"label":"CAT I",  "name":"Motos","desc":"Motocicletas y ciclomotores",
               "yolo":["motorcycle"],"color":(180,120,255),"hex":"#B478FF"},
    "CAT II": {"label":"CAT II", "name":"Automóviles y camperos",
               "desc":"Vehículos livianos de hasta 8 pasajeros",
               "yolo":["car"],"color":(0,220,80),"hex":"#00DC50"},
    "CAT III":{"label":"CAT III","name":"Buses y busetas",
               "desc":"Vehículos de transporte público colectivo",
               "yolo":["bus"],"color":(0,160,255),"hex":"#00A0FF"},
    "CAT IV": {"label":"CAT IV", "name":"Camiones livianos (2 ejes)",
               "desc":"Camiones de dos ejes peso bruto ≤10.5 t",
               "yolo":["truck"],"color":(255,200,0),"hex":"#FFC800"},
    "CAT V":  {"label":"CAT V",  "name":"Camiones pesados (3+ ejes)",
               "desc":"Vehículos de carga con tres o más ejes",
               "yolo":["truck"],"color":(220,50,50),"hex":"#DC3232"},
}

COL_CATEGORIES: Dict[str, dict] = {
    "C1":  {"label":"C1",  "name":"Automóviles, microbuses y motos",
            "desc":"Vehículos livianos de pasajeros y motos",
            "yolo":["car","motorcycle"],"color":(0,220,80),"hex":"#00DC50"},
    "C2P": {"label":"C2P", "name":"Camión 2 ejes pequeño",
            "desc":"Camión ligero de dos ejes",
            "yolo":["truck"],"color":(255,200,0),"hex":"#FFC800"},
    "C2G": {"label":"C2G", "name":"Camión 2 ejes grande / bus",
            "desc":"Camión pesado de dos ejes o bus",
            "yolo":["bus","truck"],"color":(0,160,255),"hex":"#00A0FF"},
    "C3":  {"label":"C3",  "name":"Camión 3 ejes","desc":"Vehículo articulado de tres ejes",
            "yolo":["truck"],"color":(255,120,0),"hex":"#FF7800"},
    "C4":  {"label":"C4",  "name":"Camión 4 ejes","desc":"Vehículo articulado de cuatro ejes",
            "yolo":["truck"],"color":(220,50,50),"hex":"#DC3232"},
    "C5":  {"label":"C5",  "name":"Camión 5+ ejes","desc":"Tracto-camión con más de cuatro ejes",
            "yolo":["truck"],"color":(140,0,0),"hex":"#8C0000"},
}

CLASSIFICATION_SYSTEMS: Dict[str,Dict[str,dict]] = {
    "ANI": ANI_CATEGORIES, "INVIAS": INVIAS_CATEGORIES, "COL": COL_CATEGORIES,
}


def classify_vehicle(yolo_type: str, bw: int, bh: int, system: str = "ANI") -> str:
    area   = bw * bh
    aspect = bw / max(bh, 1)
    if system == "ANI":
        if yolo_type == "motorcycle": return "CAT I"
        if yolo_type == "car":        return "CAT II"
        if yolo_type == "bus":        return "CAT III"
        if yolo_type == "truck":
            if aspect > 3.5 or area > 90_000: return "CAT VII"
            if aspect > 2.8 or area > 60_000: return "CAT VI"
            if aspect > 2.2 or area > 38_000: return "CAT V"
            return "CAT IV"
    elif system == "INVIAS":
        if yolo_type == "motorcycle": return "CAT I"
        if yolo_type == "car":        return "CAT II"
        if yolo_type == "bus":        return "CAT III"
        if yolo_type == "truck":
            return "CAT V" if (aspect > 2.5 or area > 45_000) else "CAT IV"
    elif system == "COL":
        if yolo_type in ("car","motorcycle"): return "C1"
        if yolo_type == "bus":                return "C2G"
        if yolo_type == "truck":
            if aspect > 3.5 or area > 90_000: return "C5"
            if aspect > 2.8 or area > 60_000: return "C4"
            if aspect > 2.2 or area > 38_000: return "C3"
            return "C2G" if aspect > 1.5 else "C2P"
    cats = CLASSIFICATION_SYSTEMS.get(system, {})
    for key, cat in cats.items():
        if yolo_type in cat.get("yolo", []): return key
    return list(cats.keys())[0] if cats else "CAT II"


def cat_name(cat: str, system: str) -> str:
    return CLASSIFICATION_SYSTEMS.get(system,{}).get(cat,{}).get("name", cat)

def cat_color_bgr(cat: str, system: str) -> Tuple[int,int,int]:
    c = CLASSIFICATION_SYSTEMS.get(system,{}).get(cat,{}).get("color",(200,200,200))
    return (int(c[2]),int(c[1]),int(c[0]))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — SPEED ESTIMATOR
# ══════════════════════════════════════════════════════════════════════════════

class SpeedEstimator:
    def __init__(self, fps: float, px_per_meter: float, smoothing: int = 8):
        self.fps = max(fps,1.0); self.ppm = max(px_per_meter,0.1)
        self._raw: Dict[int,deque] = {}; self._sm = smoothing

    def update(self, tid: int, vx: float, vy: float):
        if tid not in self._raw: self._raw[tid] = deque(maxlen=self._sm)
        self._raw[tid].append(math.sqrt(vx**2+vy**2))

    def get_kmh(self, tid: int) -> float:
        if tid not in self._raw or not self._raw[tid]: return 0.0
        return round(float(np.mean(self._raw[tid]))/self.ppm*self.fps*3.6, 1)

    def remove(self, tid: int): self._raw.pop(tid,None)
    def reset(self): self._raw.clear()


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — DENSITY ANALYZER
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class LaneDensity:
    lane_n:   int
    flow_buf: deque = field(default_factory=lambda: deque(maxlen=3600))
    active:   set   = field(default_factory=set)
    total:    int   = 0
    peak_vph: float = 0.0
    last_ts:  Optional[float] = None
    hw_buf:   deque = field(default_factory=lambda: deque(maxlen=50))

    def record(self, ts: float):
        if self.last_ts is not None:
            hw = ts-self.last_ts
            if 0.3 < hw < 120: self.hw_buf.append(hw)
        self.last_ts=ts; self.flow_buf.append(ts); self.total+=1
        vph=self.flow_vph()
        if vph>self.peak_vph: self.peak_vph=vph

    def flow_vph(self, w: float=60.0) -> float:
        now=time.time()
        return sum(1 for t in self.flow_buf if t>=now-w)*(3600/w)

    @property
    def avg_headway(self) -> float:
        return float(np.mean(self.hw_buf)) if self.hw_buf else 0.0


class DensityAnalyzer:
    def __init__(self, n_lanes: int, fps: float=25.0):
        self.n=n_lanes; self.fps=fps
        self.lanes: Dict[int,LaneDensity]={i:LaneDensity(i) for i in range(1,n_lanes+1)}
        self.t0=time.time()

    def record(self, ln: int):
        if ln in self.lanes: self.lanes[ln].record(time.time())

    def set_active(self, ln: int, ids: set):
        if ln in self.lanes: self.lanes[ln].active=ids

    def congestion(self, ln: int) -> str:
        if ln not in self.lanes: return "?"
        occ=min(len(self.lanes[ln].active)/10.0,1.0)
        if occ<0.25: return "LIBRE"
        if occ<0.55: return "MODERADO"
        if occ<0.80: return "CONGESTIONADO"
        return "SATURADO"

    def congestion_bgr(self, ln: int) -> Tuple[int,int,int]:
        return {"LIBRE":(0,200,80),"MODERADO":(0,200,200),
                "CONGESTIONADO":(0,120,255),"SATURADO":(0,0,220)
                }.get(self.congestion(ln),(200,200,200))

    def stats(self, ln: int) -> dict:
        if ln not in self.lanes: return {}
        s=self.lanes[ln]
        return {"lane":ln,"flow_vph_1m":round(s.flow_vph(60),1),
                "flow_vph_5m":round(s.flow_vph(300),1),
                "peak_vph":round(s.peak_vph,1),
                "congestion":self.congestion(ln),
                "avg_headway_s":round(s.avg_headway,2),"total":s.total}

    def all_stats(self) -> Dict[int,dict]:
        return {ln:self.stats(ln) for ln in self.lanes}


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — KALMAN BOX TRACKER
# ══════════════════════════════════════════════════════════════════════════════

class KalmanBoxTracker:
    """7-state Kalman filter: [cx,cy,area,ratio,vcx,vcy,varea]."""
    count = 0
    def __init__(self, bbox: np.ndarray):
        KalmanBoxTracker.count += 1
        self.id  = KalmanBoxTracker.count
        self.kf  = cv2.KalmanFilter(7,4)
        self.kf.measurementMatrix   = np.eye(4,7,dtype=np.float32)
        self.kf.transitionMatrix    = np.eye(7,dtype=np.float32)
        for i in range(4): self.kf.transitionMatrix[i,i+3]=1.0
        self.kf.processNoiseCov     = np.eye(7,dtype=np.float32)*1e-2
        self.kf.measurementNoiseCov = np.eye(4,dtype=np.float32)*1e-1
        self.kf.errorCovPost        = np.eye(7,dtype=np.float32)
        self.kf.statePost           = self._to_state(bbox)
        self.time_since_update=0; self.hits=1; self.hit_streak=1

    @staticmethod
    def _to_state(b):
        w=b[2]-b[0]; h=b[3]-b[1]; x=b[0]+w/2; y=b[1]+h/2
        return np.array([[x],[y],[w*h],[w/max(h,1)],[0.],[0.],[0.]],dtype=np.float32)

    @staticmethod
    def _to_z(b):
        w=b[2]-b[0]; h=b[3]-b[1]; x=b[0]+w/2; y=b[1]+h/2
        return np.array([[x],[y],[w*h],[w/max(h,1)]],dtype=np.float32)

    @staticmethod
    def _to_box(st):
        w=math.sqrt(abs(st[2,0]*st[3,0])); h=abs(st[2,0])/max(w,1)
        return np.array([st[0,0]-w/2,st[1,0]-h/2,st[0,0]+w/2,st[1,0]+h/2],dtype=np.float32)

    def predict(self) -> np.ndarray:
        if self.time_since_update>0: self.hit_streak=0
        self.time_since_update+=1; self.kf.predict()
        return self._to_box(self.kf.statePost)

    def update(self, bbox: np.ndarray):
        self.time_since_update=0; self.hits+=1; self.hit_streak+=1
        self.kf.correct(self._to_z(bbox))

    def state(self) -> np.ndarray:
        return self._to_box(self.kf.statePost)

    @property
    def velocity(self) -> Tuple[float,float]:
        st=self.kf.statePost; return float(st[4,0]),float(st[5,0])


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 5 — TRACK OBJECT
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Track:
    id:           int
    kalman:       KalmanBoxTracker
    cls_id:       int
    cat_key:      str
    lane:         int
    conf:         float
    state:        int  = 1
    counted:      bool = False
    counted_line: int  = -1
    age:          int  = 1
    misses:       int  = 0
    trail:        deque= field(default_factory=lambda: deque(maxlen=60))
    speeds:       deque= field(default_factory=lambda: deque(maxlen=10))
    count_ts:     Optional[str]=None
    count_frame:  int  = -1
    dir_votes:    deque= field(default_factory=lambda: deque(maxlen=8))

    @property
    def box(self)  -> np.ndarray: return self.kalman.state()
    @property
    def cx(self)   -> int: b=self.box; return int((b[0]+b[2])/2)
    @property
    def cy(self)   -> int: b=self.box; return int((b[1]+b[3])/2)
    @property
    def bw(self)   -> int: b=self.box; return int(b[2]-b[0])
    @property
    def bh(self)   -> int: b=self.box; return int(b[3]-b[1])
    @property
    def vx(self)   -> float: return self.kalman.velocity[0]
    @property
    def vy(self)   -> float: return self.kalman.velocity[1]

    def update_trail(self): self.trail.append((self.cx,self.cy))

    def update_dir(self):
        if   self.vy> 1.5: self.dir_votes.append(1)
        elif self.vy<-1.5: self.dir_votes.append(-1)

    @property
    def is_wrong_way(self) -> bool:
        return len(self.dir_votes)>=5 and sum(self.dir_votes)<0

    @property
    def avg_speed(self) -> float:
        return float(np.mean(self.speeds)) if self.speeds else 0.0


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 6 — BYTE TRACKER
# ══════════════════════════════════════════════════════════════════════════════

def _iou_matrix(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    aa=(a[:,2]-a[:,0])*(a[:,3]-a[:,1]); ab=(b[:,2]-b[:,0])*(b[:,3]-b[:,1])
    ix1=np.maximum(a[:,None,0],b[None,:,0]); iy1=np.maximum(a[:,None,1],b[None,:,1])
    ix2=np.minimum(a[:,None,2],b[None,:,2]); iy2=np.minimum(a[:,None,3],b[None,:,3])
    inter=np.maximum(0,ix2-ix1)*np.maximum(0,iy2-iy1)
    union=aa[:,None]+ab[None,:]-inter
    return np.where(union>0,inter/union,0.0)


def _greedy_match(cost: np.ndarray, threshold: float):
    matched=[]; used_t=set(); used_d=set()
    for flat in np.argsort(cost,axis=None):
        ti=flat//cost.shape[1]; di=flat%cost.shape[1]
        if cost[ti,di]>=threshold: break
        if ti in used_t or di in used_d: continue
        matched.append((int(ti),int(di))); used_t.add(ti); used_d.add(di)
    unm_t=[i for i in range(cost.shape[0]) if i not in used_t]
    unm_d=[j for j in range(cost.shape[1]) if j not in used_d]
    return matched,unm_t,unm_d


class ByteTracker:
    """ByteTrack two-stage association with Kalman-filtered tracks."""
    def __init__(self, max_misses: int=15, min_hits: int=2):
        self.tracks: List[Track]=[]; self.max_misses=max_misses
        self.min_hits=min_hits; self.iou_hi=0.30; self.iou_lo=0.15
        self.conf_split=0.50; self._nid=1

    def _new(self, det: tuple, lane: int) -> Track:
        x1,y1,x2,y2,cls_id,cat_key,conf=det
        kal=KalmanBoxTracker(np.array([x1,y1,x2,y2],dtype=np.float32))
        kal.id=self._nid; self._nid+=1
        t=Track(id=kal.id,kalman=kal,cls_id=cls_id,cat_key=cat_key,lane=lane,conf=conf)
        t.update_trail(); return t

    def _match(self, dets, tracks, thresh):
        if not dets or not tracks:
            return [],[],list(range(len(dets)))
        tb=np.array([t.box for t in tracks],dtype=np.float32)
        db=np.array([[d[0],d[1],d[2],d[3]] for d in dets],dtype=np.float32)
        m,ut,ud=_greedy_match(1.0-_iou_matrix(tb,db),1.0-thresh)
        return m,ut,ud

    def step(self, dets: List[tuple], lane_fn) -> List[Track]:
        for t in self.tracks:
            t.kalman.predict(); t.misses+=1; t.age+=1; t.update_dir()
        if not dets:
            self.tracks=[t for t in self.tracks if t.misses<=self.max_misses]
            return [t for t in self.tracks if t.kalman.hit_streak>=self.min_hits or t.counted]

        hi=[d for d in dets if d[6]>=self.conf_split]
        lo=[d for d in dets if d[6]< self.conf_split]

        m1,umt1,umd1=self._match(hi,self.tracks,self.iou_hi)
        matched_ids=set()
        for ti,di in m1:
            d=hi[di]; t=self.tracks[ti]
            t.kalman.update(np.array([d[0],d[1],d[2],d[3]],dtype=np.float32))
            t.cat_key=d[5]; t.conf=d[6]
            t.lane=lane_fn((d[0]+d[2])//2,(d[1]+d[3])//2)
            t.misses=0; t.update_trail(); matched_ids.add(id(t))

        unm_tracks=[self.tracks[i] for i in umt1]
        m2,_,_=self._match(lo,unm_tracks,self.iou_lo)
        for ti,di in m2:
            d=lo[di]; t=unm_tracks[ti]
            t.kalman.update(np.array([d[0],d[1],d[2],d[3]],dtype=np.float32))
            t.cat_key=d[5]; t.conf=d[6]
            t.lane=lane_fn((d[0]+d[2])//2,(d[1]+d[3])//2)
            t.misses=0; t.update_trail(); matched_ids.add(id(t))

        for di in umd1:
            d=hi[di]; lane=lane_fn((d[0]+d[2])//2,(d[1]+d[3])//2)
            self.tracks.append(self._new(d,lane))

        self.tracks=[t for t in self.tracks if t.misses<=self.max_misses]
        return [t for t in self.tracks if t.kalman.hit_streak>=self.min_hits or t.counted]

    def reset(self):
        self.tracks.clear(); KalmanBoxTracker.count=0; self._nid=1


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 7 — ROBUST LANE BOUNDARY SYSTEM  (v2 — no Hough hallucination)
# ══════════════════════════════════════════════════════════════════════════════
#
#  ROOT CAUSE of the old bug:
#    HoughLinesP fires on power cables, shadows, car roof edges, and road
#    cracks.  Those spurious segments then get clustered into fake "lane
#    lines" at wrong angles.  The fix: stop using Hough for lane geometry.
#
#  NEW APPROACH — three layers, each more reliable than Hough:
#
#  Layer 1 — SCANLINE WHITE-PEAK DETECTION
#    Sample N horizontal scanlines in the road ROI.  On each scanline,
#    find local brightness peaks in the white-mask image.  These peaks are
#    actual road markings, not cables (cables are thin, diagonal, and faint).
#    Project peaks to a common reference y via the camera's perspective
#    transform (vanishing-point model).  Cluster the projected x-positions
#    with a simple gap-based merge to find n_lanes+1 boundary candidates.
#
#  Layer 2 — VEHICLE-TRACK BOOTSTRAPPING
#    After the first ~60 frames the ByteTracker has real vehicle trajectories.
#    Lane centres derived from where vehicles actually drive are perfect ground
#    truth.  K-means on track centroids at the reference y gives extremely
#    stable lane centres; outer edges are inferred with a fixed margin.
#    This layer activates automatically once enough tracks exist.
#
#  Layer 3 — STABLE PERSPECTIVE FALLBACK
#    If neither layer fires (night, fog, empty road), use a uniform grid with
#    correct perspective scaling.  This is ALWAYS a sensible starting point
#    and never produces the random diagonal lines seen before.
#
#  Temporal smoothing (EMA α=0.1) is applied to all boundary x-positions
#  at every update so the overlay lines never jump.
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class LaneBoundary:
    """Single lane boundary — two anchor points (top/bottom of ROI)."""
    x_top:    float
    x_bottom: float
    y_top:    int
    y_bottom: int
    detected: bool = False

    def x_at(self, y: int) -> int:
        if self.y_bottom == self.y_top: return int(self.x_top)
        t = (y - self.y_top) / (self.y_bottom - self.y_top)
        return int(self.x_top + t * (self.x_bottom - self.x_top))


class ScanlineWhiteDetector:
    """
    Finds lane-marking positions by scanning horizontal stripes inside the
    road ROI and locating white/yellow brightness peaks.

    Why this beats Hough on CCTV footage:
      • Only considers high-lightness pixels → ignores dark cables/shadows
      • Each peak is anchored to a real pixel, not a line extrapolation
      • Perspective projection maps every peak to the same reference y
        so clustering is 1-D and trivially correct
      • No angle filtering needed — horizontal scan is inherently correct
    """

    def __init__(self, H: int, W: int, roi_top_frac: float = 0.30,
                 roi_bottom_frac: float = 0.92, roi_left_frac: float = 0.44):
        self.H = H; self.W = W
        self.roi_top    = int(H * roi_top_frac)
        self.roi_bottom = int(H * roi_bottom_frac)
        # Vanishing point: assumed at horizontal centre, at roi_top
        self.vp_x = W * 0.68  # camera left-of-road: VP is right-of-centre
        self.vp_y = float(self.roi_top)

    # ── Perspective scale: how much wider the road is at row y vs at roi_top ─
    def _persp_scale(self, y: int) -> float:
        dy_bottom = self.roi_bottom - self.vp_y
        dy_now    = max(y - self.vp_y, 1.0)
        return dy_now / max(dy_bottom, 1.0)   # 0..1, 1 at bottom

    # ── Map an x-coordinate measured at row y → equivalent x at roi_bottom ──
    def _project_to_bottom(self, x: float, y: int) -> float:
        sc_now = self._persp_scale(y)
        if sc_now < 1e-4: return x
        # Road extends from vp_x ± half_road_width * scale at row y
        offset_from_centre = x - self.vp_x
        return self.vp_x + offset_from_centre / sc_now

    # ── White+yellow pixel mask ───────────────────────────────────────────────
    @staticmethod
    def _road_marking_mask(frame: np.ndarray) -> np.ndarray:
        # Suppress glare (>240 brightness → grey, kills sun reflections)
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        glare = gray > 240
        fg = frame.copy()
        fg[glare] = [100, 100, 100]

        hls = cv2.cvtColor(fg, cv2.COLOR_BGR2HLS)
        # White:  high lightness (>160), very low saturation (<45)
        white  = cv2.inRange(hls, np.array([0,  160, 0],  np.uint8),
                                  np.array([180, 255, 45], np.uint8))
        # Yellow: hue 12-35, good lightness, high saturation
        yellow = cv2.inRange(hls, np.array([12,  80, 80],  np.uint8),
                                  np.array([35,  220, 255], np.uint8))
        mask = cv2.bitwise_or(white, yellow)
        # Close dashes (connect short broken markings)
        k = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 15))
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, k)
        # Remove tiny specks (min width ~8 px)
        k2 = cv2.getStructuringElement(cv2.MORPH_RECT, (8, 1))
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, k2)
        return mask

    # ── Find peaks along a single horizontal scanline ────────────────────────
    @staticmethod
    def _scanline_peaks(row: np.ndarray, min_gap: int = 20,
                        min_width: int = 4) -> List[int]:
        """Return x-centres of white runs in a 1-D binary row."""
        peaks = []
        in_run = False; run_start = 0
        for x, v in enumerate(row):
            if v > 0 and not in_run:
                in_run = True; run_start = x
            elif v == 0 and in_run:
                run_len = x - run_start
                if run_len >= min_width:
                    cx = (run_start + x) // 2
                    # Enforce minimum gap between accepted peaks
                    if not peaks or (cx - peaks[-1]) >= min_gap:
                        peaks.append(cx)
                in_run = False
        if in_run:
            run_len = len(row) - run_start
            if run_len >= min_width:
                cx = (run_start + len(row)) // 2
                if not peaks or (cx - peaks[-1]) >= min_gap:
                    peaks.append(cx)
        return peaks

    def detect_boundary_xs(self, frame: np.ndarray,
                            n_boundaries: int) -> Optional[List[float]]:
        """
        Return n_boundaries x-positions (at roi_bottom reference) sorted left→right.
        Returns None if insufficient evidence.
        """
        mask = self._road_marking_mask(frame)
        # Apply road ROI trapezoid.
        # ROI trapezoid: excludes non-road area on the left.
        # roi_left_x controls where the road starts (default 0.44 for this camera).
        rl_bot = self.roi_left_x
        rl_top = max(self.roi_left_x - 0.02, 0.0)  # slight taper toward VP
        roi_mask = np.zeros_like(mask)
        pts = np.array([
            [int(self.W * rl_bot), self.roi_bottom],  # bottom-left
            [int(self.W * rl_top), self.roi_top],     # top-left
            [int(self.W * 0.98),   self.roi_top],     # top-right
            [int(self.W * 0.99),   self.roi_bottom],  # bottom-right
        ], dtype=np.int32)
        cv2.fillPoly(roi_mask, [pts], 255)
        mask = cv2.bitwise_and(mask, roi_mask)

        # Sample scanlines: 8 evenly spaced rows in the lower 2/3 of ROI
        scan_ys = np.linspace(
            self.roi_top + (self.roi_bottom - self.roi_top) * 0.35,
            self.roi_bottom - 10, 8, dtype=int)

        projected: List[float] = []
        for y in scan_ys:
            if y < 0 or y >= self.H: continue
            row = mask[y, :]
            peaks = self._scanline_peaks(row)
            for px in peaks:
                projected.append(self._project_to_bottom(float(px), int(y)))

        if len(projected) < n_boundaries * 2:
            return None  # Not enough evidence

        # Sort and cluster with gap threshold = 6% of frame width
        projected.sort()
        gap_thresh = self.W * 0.06
        clusters: List[List[float]] = []
        for px in projected:
            if clusters and (px - clusters[-1][-1]) < gap_thresh:
                clusters[-1].append(px)
            else:
                clusters.append([px])

        # Keep only clusters with >1 vote (reject isolated noise pixels)
        strong = [c for c in clusters if len(c) >= 2]
        if len(strong) < 2:
            return None

        centres = sorted(float(np.median(c)) for c in strong)

        # If we have exactly the right count, use them directly
        if len(centres) == n_boundaries:
            return centres

        # If too many, keep the n_boundaries most evenly-spaced
        if len(centres) > n_boundaries:
            # Pick subset maximising minimum gap (greedy)
            best: List[float] = []
            for start_idx in range(len(centres) - n_boundaries + 1):
                candidate = centres[start_idx: start_idx + n_boundaries]
                if not best:
                    best = candidate
                    continue
                min_gap_c = min(candidate[i+1]-candidate[i]
                                for i in range(len(candidate)-1))
                min_gap_b = min(best[i+1]-best[i] for i in range(len(best)-1))
                if min_gap_c > min_gap_b:
                    best = candidate
            return best if len(best) == n_boundaries else centres[:n_boundaries]

        # If too few: return what we found (LaneManager will interpolate)
        return centres


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 8 — LANE MANAGER  (camera-agnostic, hallucination-free)
# ══════════════════════════════════════════════════════════════════════════════

_LANE_COLORS = [(255,255,120),(120,220,255),(120,255,160),(255,180,120),(220,120,255)]


@dataclass
class Lane:
    number: int
    left:   LaneBoundary
    right:  LaneBoundary
    color:  Tuple[int,int,int] = (255,255,255)
    label:  str = ""

    def contains(self, cx: int, cy: int) -> bool:
        return self.left.x_at(cy) <= cx <= self.right.x_at(cy)

    def center_x(self, y: int) -> int:
        return (self.left.x_at(y) + self.right.x_at(y)) // 2

    def width(self, y: int) -> int:
        return abs(self.right.x_at(y) - self.left.x_at(y))


class LaneManager:
    """
    Camera-agnostic lane manager.

    Priority order for lane geometry:
      1. Scanline white-peak detector   (fires when bright markings visible)
      2. Vehicle-track bootstrapper     (fires after ~60 detected vehicles)
      3. Uniform perspective fallback   (always available, never wrong-direction)

    All x-positions are smoothed with EMA (α=0.08) so lane lines are stable.
    No Hough transform is used — zero hallucination risk from cables/shadows.
    """

    # EMA smoothing factor: lower = smoother but slower to adapt
    _EMA_ALPHA = 0.08

    def __init__(self, W: int, H: int, n_lanes: int,
                 direction: str = "NS", mode: str = "auto",
                 roi_top_frac: float = 0.30, roi_left_frac: float = 0.44):
        self.W = W; self.H = H; self.n_lanes = n_lanes
        self.direction = direction; self.mode = mode
        self.roi_top     = int(H * roi_top_frac)
        self.roi_bottom  = int(H * 0.92)
        self._roi_left_frac = roi_left_frac

        self.detector = ScanlineWhiteDetector(H, W, roi_top_frac, 0.92, self._roi_left_frac)
        self.lanes: List[Lane] = []

        # EMA state: boundary x-positions at roi_top and roi_bottom
        # Shape: (n_lanes+1, 2)  → each row = [x_top, x_bottom]
        self._ema_xs: Optional[np.ndarray] = None

        # Track accumulator for vehicle-based calibration
        # Stores (cx_at_ref_y) for counted/stable tracks
        self._track_xs: deque = deque(maxlen=400)
        self._track_calibrated = False
        self._track_calib_count = 0  # how many times track calib succeeded

        self._frame_n = 0
        self._scan_every = 6   # run scanline detector every N frames
        self._build_uniform()  # always start from a sane baseline

    # ── Perspective-aware reference y ────────────────────────────────────────
    @property
    def _ref_y(self) -> int:
        """Reference row for x-comparisons: lower third of ROI."""
        return int(self.roi_top + (self.roi_bottom - self.roi_top) * 0.75)

    # ── Perspective interpolation helper ─────────────────────────────────────
    def _make_boundary(self, x_bottom: float, x_top: float,
                       detected: bool = True) -> LaneBoundary:
        return LaneBoundary(
            x_top=x_top, x_bottom=x_bottom,
            y_top=self.roi_top, y_bottom=self.roi_bottom,
            detected=detected)

    # ── Build boundaries from n_lanes+1 x-positions at roi_bottom ────────────
    def _xs_bottom_to_lanes(self, xs_bottom: List[float],
                             detected: bool = True):
        """
        xs_bottom: sorted x positions at roi_bottom for n_lanes+1 boundaries.
        Derive x_top via vanishing-point perspective model.
        """
        n = self.n_lanes
        vp_x = self.W * 0.68  # vanishing point x — camera left-of-road
        # Perspective ratio: at roi_top lines are compressed toward vp_x
        scale_top = (self.roi_top  - self.detector.vp_y) / \
                    max(self.roi_bottom - self.detector.vp_y, 1.0)

        bounds = []
        for xb in xs_bottom:
            # Top position: converge toward vanishing point
            xt = vp_x + (xb - vp_x) * scale_top
            bounds.append(self._make_boundary(xb, xt, detected))

        for i in range(n):
            col = _LANE_COLORS[i % len(_LANE_COLORS)]
            if i < len(self.lanes):
                self.lanes[i].left  = bounds[i]
                self.lanes[i].right = bounds[i+1]
                self.lanes[i].color = col
            else:
                self.lanes.append(Lane(
                    number=i+1, left=bounds[i], right=bounds[i+1],
                    color=col, label=f"Carril {i+1}"))
        self.lanes = self.lanes[:n]

    # ── EMA update ────────────────────────────────────────────────────────────
    def _ema_update(self, new_xs_bottom: List[float]) -> List[float]:
        """Smooth boundary x-positions at roi_bottom with EMA."""
        arr = np.array(new_xs_bottom, dtype=np.float64)
        if self._ema_xs is None or len(self._ema_xs) != len(arr):
            self._ema_xs = arr.copy()
        else:
            self._ema_xs = (1.0 - self._EMA_ALPHA) * self._ema_xs + \
                            self._EMA_ALPHA          * arr
        return self._ema_xs.tolist()

    # ── Layer 3: uniform perspective fallback ─────────────────────────────────
    def _build_uniform(self):
        """
        Evenly-spaced lanes that respect perspective.
        Road occupies 8%…96% of frame width at roi_bottom.
        This is always a correct starting layout and is used until real
        detections improve it.
        """
        road_left  = self.W * self._roi_left_frac
        road_right = self.W * 0.99
        xs = [road_left + i * (road_right - road_left) / self.n_lanes
              for i in range(self.n_lanes + 1)]
        smoothed = self._ema_update(xs)
        self._xs_bottom_to_lanes(smoothed, detected=False)

    # ── Layer 2: vehicle-track calibration ────────────────────────────────────
    def feed_track_cx(self, cx: int, cy: int):
        """
        Call this for every confirmed (counted) vehicle centroid.
        Once we have enough points, recalibrate lane centres.
        """
        if cy <= 0 or cy >= self.H: return
        # Project to roi_bottom reference via vanishing-point model
        ref_x = self.detector._project_to_bottom(float(cx), cy)
        self._track_xs.append(ref_x)

    def _try_track_calibration(self) -> bool:
        """
        Infer lane centre positions from vehicle track accumulator.
        Returns True if calibration succeeded.
        """
        if len(self._track_xs) < self.n_lanes * 12:
            return False

        xs = np.array(sorted(self._track_xs))
        # K-means 1-D: iteratively assign points to nearest centroid
        # Initialise with uniform spacing across the data range
        lo, hi = float(xs.min()), float(xs.max())
        centres = np.linspace(lo, hi, self.n_lanes)

        for _ in range(20):
            labels = np.argmin(np.abs(xs[:, None] - centres[None, :]), axis=1)
            new_centres = np.array([
                xs[labels == k].mean() if (labels == k).any() else centres[k]
                for k in range(self.n_lanes)
            ])
            if np.allclose(new_centres, centres, atol=1.0): break
            centres = new_centres

        centres = np.sort(centres)
        if len(centres) < 2: return False

        # Estimate lane half-width from spacing between centres
        spacings = np.diff(centres)
        half_w   = float(np.median(spacings)) * 0.52

        xs_bound = [float(centres[0] - half_w)]
        for c in centres:
            xs_bound.append(float(c + half_w))
        xs_bound[0]  = max(xs_bound[0],  self.W * 0.40)
        xs_bound[-1] = min(xs_bound[-1], self.W * 0.99)

        if len(xs_bound) != self.n_lanes + 1: return False

        smoothed = self._ema_update(xs_bound)
        self._xs_bottom_to_lanes(smoothed, detected=True)
        self._track_calibrated = True
        return True

    # ── Layer 1: scanline white-peak calibration ──────────────────────────────
    def _try_scanline_calibration(self, frame: np.ndarray) -> bool:
        xs = self.detector.detect_boundary_xs(frame, self.n_lanes + 1)
        if xs is None or len(xs) < 2:
            return False

        # Validate: boundaries must be ordered and spaced reasonably
        xs_sorted = sorted(xs)
        min_gap = self.W * 0.04
        if any((xs_sorted[i+1] - xs_sorted[i]) < min_gap
               for i in range(len(xs_sorted) - 1)):
            return False

        # Pad or trim to exactly n_lanes+1 entries using uniform fallback
        while len(xs_sorted) < self.n_lanes + 1:
            # Append a boundary one lane-width past the last
            gap = xs_sorted[-1] - xs_sorted[-2] if len(xs_sorted) >= 2 else self.W * 0.2
            xs_sorted.append(min(xs_sorted[-1] + gap, self.W * 0.99))

        xs_sorted = xs_sorted[:self.n_lanes + 1]
        smoothed  = self._ema_update(xs_sorted)
        self._xs_bottom_to_lanes(smoothed, detected=True)
        return True

    # ── Public: update every frame ────────────────────────────────────────────
    def update(self, frame: np.ndarray):
        self._frame_n += 1

        if self.mode == "uniform":
            # Recalculate uniform in case EMA drifted
            road_left  = self.W * self._roi_left_frac
            road_right = self.W * 0.99  # road ends at ~99% x
            xs = [road_left + i * (road_right - road_left) / self.n_lanes
                  for i in range(self.n_lanes + 1)]
            smoothed = self._ema_update(xs)
            self._xs_bottom_to_lanes(smoothed, detected=False)
            return

        # Try track calibration every 30 frames (highest priority once warm)
        if self._frame_n % 30 == 0 and len(self._track_xs) >= self.n_lanes * 12:
            self._try_track_calibration()
            return

        # Try scanline detector every _scan_every frames
        if self._frame_n % self._scan_every == 0:
            success = self._try_scanline_calibration(frame)
            if not success and not self._track_calibrated:
                # Re-apply uniform so we don't drift from no-detection frames
                road_left  = self.W * self._roi_left_frac
                road_right = self.W * 0.99  # road ends at ~99% x
                xs = [road_left + i * (road_right - road_left) / self.n_lanes
                      for i in range(self.n_lanes + 1)]
                smoothed = self._ema_update(xs)
                self._xs_bottom_to_lanes(smoothed, detected=False)

    # ── Public: lane assignment ───────────────────────────────────────────────
    def get_lane(self, cx: int, cy: int = 0) -> int:
        """Return 1-based lane number for vehicle centroid (cx, cy)."""
        if cy <= 0: cy = self._ref_y
        # Feed into track accumulator for calibration
        self.feed_track_cx(cx, cy)
        for lane in self.lanes:
            if lane.contains(cx, cy):
                return lane.number
        # Nearest-centre fallback
        ref_y = self._ref_y
        dists = [abs(cx - lane.center_x(ref_y)) for lane in self.lanes]
        return self.lanes[int(np.argmin(dists))].number

    # ── Public: draw overlay ──────────────────────────────────────────────────
    def draw(self, frame: np.ndarray, lane_counts: Optional[dict] = None):
        H = frame.shape[0]
        overlay = frame.copy()
        for lane in self.lanes:
            pts = np.array([
                [int(lane.left.x_at(self.roi_top)),   self.roi_top],
                [int(lane.right.x_at(self.roi_top)),  self.roi_top],
                [int(lane.right.x_at(H)),              H],
                [int(lane.left.x_at(H)),               H],
            ], dtype=np.int32)
            dim = tuple(max(0, int(c * 0.16)) for c in lane.color)
            cv2.fillPoly(overlay, [pts], dim)
        cv2.addWeighted(overlay, 0.25, frame, 0.75, 0, frame)

        n = self.n_lanes
        drawn_xs = set()   # avoid redrawing the same shared boundary twice
        for i, lane in enumerate(self.lanes):
            for side in ("left", "right"):
                b    = lane.left if side == "left" else lane.right
                is_outer = (i == 0 and side == "left") or \
                           (i == n-1 and side == "right")
                xt = b.x_at(self.roi_top)
                xb = b.x_at(H)
                key = (xt, xb)
                if key in drawn_xs: continue
                drawn_xs.add(key)
                col   = (255, 255, 255) if b.detected else (140, 180, 80)
                thick = 2 if is_outer else 1
                cv2.line(frame, (xt, self.roi_top), (xb, H),
                         col, thick, cv2.LINE_AA)

        # Lane labels
        for lane in self.lanes:
            ref_y  = int(H * 0.22)
            cx_lbl = lane.center_x(ref_y)
            total  = sum(lane_counts.get(lane.number, {}).values()) \
                     if lane_counts else 0
            cv2.putText(frame, f"L{lane.number}:{total}",
                        (cx_lbl - 22, ref_y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.52,
                        lane.color, 1, cv2.LINE_AA)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 9 — HUD ENGINE
# ══════════════════════════════════════════════════════════════════════════════

C_WHITE=(255,255,255); C_BLACK=(0,0,0)
C_GREEN=(0,220,80);    C_ORANGE=(0,160,255); C_RED=(0,0,220)
C_CYAN=(255,220,0);    C_DARK=(10,10,10)

def _put(img,txt,pos,scale=0.46,color=C_WHITE,bg=None,thickness=1):
    f=cv2.FONT_HERSHEY_SIMPLEX
    (tw,th),bl=cv2.getTextSize(txt,f,scale,thickness)
    x,y=int(pos[0]),int(pos[1])
    if bg is not None:
        cv2.rectangle(img,(x-2,y-th-2),(x+tw+2,y+bl+1),bg,-1)
    cv2.putText(img,txt,(x,y),f,scale,color,thickness,cv2.LINE_AA)


class HUDEngine:
    def __init__(self,direction,system,camera_id,site_name,
                 cfg_show_speed=True,cfg_show_trails=True,
                 cfg_show_wrong=True,cfg_show_hmap=False):
        self.direction=direction; self.system=system
        self.camera_id=camera_id; self.site_name=site_name
        self.show_speed=cfg_show_speed; self.show_trails=cfg_show_trails
        self.show_wrong=cfg_show_wrong; self.show_hmap=cfg_show_hmap
        self._heatmap=None

    def _ensure_hmap(self,H,W):
        if self._heatmap is None:
            self._heatmap=np.zeros((H,W),dtype=np.float32)

    def update_heatmap(self,frame,tracks):
        H,W=frame.shape[:2]; self._ensure_hmap(H,W)
        for tr in tracks:
            cx,cy=tr.cx,tr.cy
            if 0<=cx<W and 0<=cy<H: self._heatmap[cy,cx]+=1.0

    def draw_heatmap(self,frame):
        if self._heatmap is None: return
        norm=cv2.normalize(self._heatmap,None,0,255,cv2.NORM_MINMAX)
        hm=cv2.applyColorMap(norm.astype(np.uint8),cv2.COLORMAP_JET)
        cv2.addWeighted(frame,0.65,hm,0.35,0,frame)

    def draw_count_lines(self,frame,lines_y):
        H,W=frame.shape[:2]
        for i,ly in enumerate(lines_y):
            col=(0,255,255) if i==1 else (0,200,180)
            thick=2 if i==1 else 1
            cv2.line(frame,(0,ly),(W,ly),col,thick,cv2.LINE_AA)
        mid=lines_y[len(lines_y)//2]
        cv2.putText(frame,"LINEAS DE CONTEO",(10,mid-8),
                    cv2.FONT_HERSHEY_SIMPLEX,0.45,(0,255,255),1,cv2.LINE_AA)

    def draw_cat_panel(self,frame,cats,cat_counts,total,fps):
        rows=4+len(cats); ph=18+rows*23
        ov=frame.copy()
        cv2.rectangle(ov,(4,4),(252,ph),C_DARK,-1)
        cv2.addWeighted(ov,0.70,frame,0.30,0,frame)
        y=26
        cv2.putText(frame,f"TOTAL: {total}",(10,y),
                    cv2.FONT_HERSHEY_SIMPLEX,0.80,C_WHITE,2,cv2.LINE_AA)
        y+=26
        _put(frame,f"Sistema : {self.system}",  (10,y),scale=0.42,color=(180,255,180)); y+=19
        _put(frame,f"Dir     : {self.direction}",(10,y),scale=0.42,color=(180,255,180)); y+=19
        _put(frame,f"Cam     : {self.camera_id}",(10,y),scale=0.42,color=(180,255,180)); y+=22
        for cat_key in sorted(cats.keys()):
            cnt=cat_counts.get(cat_key,0)
            col=cats[cat_key].get("color",(200,200,200))
            bgr=(int(col[2]),int(col[1]),int(col[0]))
            _put(frame,f"  {cat_key:<10}: {cnt}",(10,y),scale=0.46,color=bgr); y+=22
        _put(frame,f"FPS: {fps:.1f}",(10,y+2),scale=0.40,color=(120,120,120))

    def draw_lane_panel(self,frame,lane_counts,density_stats):
        H,W=frame.shape[:2]
        lane_nums=sorted(lane_counts.keys())
        if not lane_nums: return
        rw=185; rh=26+len(lane_nums)*52+10; rx=W-rw-4
        ov=frame.copy()
        cv2.rectangle(ov,(rx,4),(W-4,rh),C_DARK,-1)
        cv2.addWeighted(ov,0.70,frame,0.30,0,frame)
        ry=26
        cv2.putText(frame,"POR CARRIL",(rx+8,ry),
                    cv2.FONT_HERSHEY_SIMPLEX,0.54,C_WHITE,1,cv2.LINE_AA)
        ry+=22
        for ln in lane_nums:
            lt=sum(lane_counts[ln].values())
            st=density_stats.get(ln,{})
            fvph=st.get("flow_vph_1m",0)
            cong=st.get("congestion","?")
            cong_col={"LIBRE":(0,200,80),"MODERADO":(0,200,200),
                      "CONGESTIONADO":(0,100,255),"SATURADO":(0,0,220)
                      }.get(cong,(200,200,200))
            _put(frame,f" Carril {ln}: {lt}",(rx+6,ry),scale=0.48,color=C_WHITE); ry+=20
            _put(frame,f"  {fvph:.0f} veh/h",(rx+6,ry),scale=0.40,color=(160,220,255)); ry+=18
            _put(frame,f"  {cong}",           (rx+6,ry),scale=0.40,color=cong_col);     ry+=24

    def draw_bottom_bar(self,frame,fps,frame_n,total):
        H,W=frame.shape[:2]
        cv2.rectangle(frame,(0,H-22),(W,H),C_DARK,-1)
        ts=datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        _put(frame,ts,      (W//2-90,H-5),scale=0.44,color=C_WHITE)
        _put(frame,f"Frame:{frame_n}",(W-140,H-5),scale=0.40,color=(150,150,150))
        _put(frame,self.site_name,    (8,H-5),     scale=0.40,color=(150,150,150))

    def draw_vehicle(self,frame,tr,cat_info,speed_kmh):
        x1,y1,x2,y2=[int(v) for v in tr.box]
        col=cat_info.get("color",(200,200,200))
        bgr=(int(col[2]),int(col[1]),int(col[0]))
        thick=3 if tr.counted else 2
        cv2.rectangle(frame,(x1,y1),(x2,y2),bgr,thick)
        # Corner accent marks
        corner=10
        for px,py in [(x1,y1),(x2,y1),(x1,y2),(x2,y2)]:
            dx=1 if px==x1 else -1; dy=1 if py==y1 else -1
            cv2.line(frame,(px,py),(px+dx*corner,py),C_WHITE,2)
            cv2.line(frame,(px,py),(px,py+dy*corner),C_WHITE,2)
        lbl=f"{cat_info['label']}  L{tr.lane}  #{tr.id}"
        _put(frame,lbl,(x1,y1-4),scale=0.42,color=C_WHITE,bg=bgr)
        if self.show_speed and speed_kmh>0.5:
            sc=C_GREEN if speed_kmh<60 else (C_ORANGE if speed_kmh<90 else C_RED)
            _put(frame,f"{speed_kmh:.0f}km/h",(x1,y2+14),scale=0.38,color=sc,bg=C_BLACK)
        if self.show_trails:
            pts=list(tr.trail)
            for i in range(1,len(pts)):
                frac=i/max(len(pts)-1,1)
                cv2.line(frame,pts[i-1],pts[i],tuple(int(c*frac) for c in bgr),1)
            cv2.circle(frame,(tr.cx,tr.cy),4,bgr,-1)
        if self.show_wrong and tr.is_wrong_way:
            cv2.rectangle(frame,(x1-2,y1-20),(x1+120,y1-1),(0,0,180),-1)
            _put(frame,"CONTRAFLUJO",(x1,y1-5),scale=0.44,color=C_WHITE)

    def draw_all(self,frame,active_tracks,cats,cat_counts,lane_counts,
                 density_stats,fps,frame_n,speed_map,count_lines):
        total=sum(cat_counts.values())
        if self.show_hmap:
            self.update_heatmap(frame,active_tracks); self.draw_heatmap(frame)
        for tr in active_tracks:
            info=cats.get(tr.cat_key,{"label":tr.cat_key,"color":(200,200,200)})
            self.draw_vehicle(frame,tr,info,speed_map.get(tr.id,0.0))
        self.draw_count_lines(frame,count_lines)
        self.draw_cat_panel(frame,cats,cat_counts,total,fps)
        self.draw_lane_panel(frame,lane_counts,density_stats)
        self.draw_bottom_bar(frame,fps,frame_n,total)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 10 — EXCEL REPORT (5 sheets + embedded bar chart)
# ══════════════════════════════════════════════════════════════════════════════

_XL_DARK="1F3864"; _XL_MED="2E75B6"; _XL_LIGHT="D6E4F7"; _XL_GRAY="F2F2F2"
_CAT_XL={"CAT I":"C9B3FF","CAT II":"A8F0C0","CAT III":"A0D8FF",
          "CAT IV":"FFE680","CAT V":"FFB366","CAT VI":"FF9090","CAT VII":"FF6060",
          "C1":"A8F0C0","C2P":"FFE680","C2G":"A0D8FF",
          "C3":"FFB366","C4":"FF9090","C5":"FF6060"}

def _xlh(ws,ref,text,fill=_XL_DARK,fc="FFFFFF",bold=True,sz=11,center=True):
    c=ws[ref]; c.value=text
    c.font=Font(color=fc,bold=bold,size=sz)
    c.fill=PatternFill("solid",fgColor=fill)
    if center: c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    return c


def write_excel(path,summary,lane_counts,events,speed_events,
                system,direction,camera_id,video_source,n_lanes,
                site_name="Peaje Niquía",operator=""):
    if not HAVE_EXCEL:
        print("[WARN] openpyxl not installed — skipping Excel."); return

    wb=openpyxl.Workbook()
    brd=Border(left=Side(style="thin",color="AAAAAA"),right=Side(style="thin",color="AAAAAA"),
               top=Side(style="thin",color="AAAAAA"), bottom=Side(style="thin",color="AAAAAA"))
    ctr=Alignment(horizontal="center",vertical="center")
    lft=Alignment(horizontal="left",  vertical="center")
    grand_total=sum(summary.values())
    all_cats=sorted(summary.keys()); lane_nums=sorted(lane_counts.keys())

    # ── Sheet 1: Resumen Ejecutivo ────────────────────────────────────────────
    ws=wb.active; ws.title="Resumen Ejecutivo"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:H1")
    _xlh(ws,"A1",f"INFORME DE CONTEO VEHICULAR — {site_name.upper()}",sz=14)
    ws.row_dimensions[1].height=32
    ws.merge_cells("A2:H2")
    _xlh(ws,"A2",
         f"Sistema: {system}  |  Dirección: {direction}  |  "
         f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
         _XL_MED,sz=10)
    meta=[("Cámara / Fuente:",camera_id or str(video_source)),
          ("Sistema clasificación:",system),("Dirección de flujo:",direction),
          ("Número de carriles:",str(n_lanes)),("Operador:",operator or "—"),
          ("Fecha:",datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
          ("Total vehículos:",str(grand_total))]
    for r,(k,v) in enumerate(meta,start=4):
        ws.cell(r,1,k).font=Font(bold=True,size=10); ws.cell(r,1).alignment=lft
        ws.cell(r,2,v).alignment=lft; ws.cell(r,2).font=Font(size=10)
        if r%2==0:
            for c in range(1,3): ws.cell(r,c).fill=PatternFill("solid",fgColor=_XL_LIGHT)
    r0=13
    for ci,h in enumerate(["Categoría","Descripción","Conteo","% del Total","Ejes"],start=1):
        _xlh(ws,ws.cell(r0,ci).coordinate,h,_XL_MED)
    cats_sys=CLASSIFICATION_SYSTEMS.get(system,{})
    for ri,(cat,cnt) in enumerate(summary.items(),start=r0+1):
        pct=(cnt/grand_total*100) if grand_total else 0
        axles=cats_sys.get(cat,{}).get("axles","—")
        for ci,val in enumerate([cat,cat_name(cat,system),cnt,f"{pct:.1f}%",axles],start=1):
            c2=ws.cell(ri,ci,val); c2.border=brd
            c2.alignment=ctr if ci!=2 else lft
            c2.fill=PatternFill("solid",fgColor=_CAT_XL.get(cat,"FFFFFF"))
            if ci==3: c2.font=Font(bold=True)
    r_tot=r0+len(summary)+1
    for ci,val in enumerate(["TOTAL","",grand_total,"100%"],start=1):
        c2=ws.cell(r_tot,ci,val)
        c2.font=Font(bold=True,size=11 if ci==1 else 10)
        c2.fill=PatternFill("solid",fgColor=_XL_LIGHT)
        c2.border=brd; c2.alignment=ctr
    for col,w in zip("ABCDE",[14,42,10,14,8]):
        ws.column_dimensions[col].width=w
    try:
        chart=BarChart(); chart.type="col"; chart.grouping="clustered"
        chart.title="Distribución por Categoría"; chart.width=18; chart.height=12
        dr=Reference(ws,min_col=3,max_col=3,min_row=r0,max_row=r0+len(summary))
        cr=Reference(ws,min_col=1,max_col=1,min_row=r0+1,max_row=r0+len(summary))
        chart.add_data(dr,titles_from_data=True); chart.set_categories(cr)
        ws.add_chart(chart,"G13")
    except Exception: pass

    # ── Sheet 2: Por Carril ───────────────────────────────────────────────────
    ws2=wb.create_sheet("Por Carril"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:J1"); _xlh(ws2,"A1","CONTEO POR CARRIL Y CATEGORÍA",sz=13)
    hdr2=["Carril"]+all_cats+["TOTAL","% del Total"]
    for ci,h in enumerate(hdr2,start=1):
        _xlh(ws2,ws2.cell(3,ci).coordinate,h,_XL_MED)
        ws2.column_dimensions[get_column_letter(ci)].width=13
    for ri,lane in enumerate(lane_nums,start=4):
        ws2.cell(ri,1,f"Carril {lane}").font=Font(bold=True)
        ws2.cell(ri,1).alignment=ctr; row_total=0
        for ci,cat in enumerate(all_cats,start=2):
            cnt=lane_counts[lane].get(cat,0)
            ws2.cell(ri,ci,cnt).alignment=ctr; ws2.cell(ri,ci).border=brd
            ws2.cell(ri,ci).fill=PatternFill("solid",fgColor=_CAT_XL.get(cat,"F0F0F0"))
            row_total+=cnt
        pct=(row_total/grand_total*100) if grand_total else 0
        ws2.cell(ri,len(all_cats)+2,row_total).font=Font(bold=True)
        ws2.cell(ri,len(all_cats)+2).alignment=ctr
        ws2.cell(ri,len(all_cats)+3,f"{pct:.1f}%").alignment=ctr
    r2t=4+len(lane_nums)
    ws2.cell(r2t,1,"TOTAL").font=Font(bold=True)
    for ci,cat in enumerate(all_cats,start=2):
        t=sum(lane_counts[l].get(cat,0) for l in lane_nums)
        ws2.cell(r2t,ci,t).font=Font(bold=True); ws2.cell(r2t,ci).alignment=ctr
        ws2.cell(r2t,ci).fill=PatternFill("solid",fgColor=_XL_LIGHT)
    ws2.cell(r2t,len(all_cats)+2,grand_total).font=Font(bold=True)
    ws2.cell(r2t,len(all_cats)+2).alignment=ctr
    ws2.cell(r2t,len(all_cats)+2).fill=PatternFill("solid",fgColor=_XL_LIGHT)

    # ── Sheet 3: Flujo por Minuto ─────────────────────────────────────────────
    ws3=wb.create_sheet("Flujo por Minuto"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:J1"); _xlh(ws3,"A1","VOLUMEN DE TRÁFICO POR MINUTO",sz=13)
    if events:
        md: dict=defaultdict(lambda: defaultdict(int))
        for ev in events:
            md[ev.get("time","00:00")[:5]][ev.get("category","?")] += 1
        hdr3=["Minuto"]+all_cats+["TOTAL"]
        for ci,h in enumerate(hdr3,start=1):
            _xlh(ws3,ws3.cell(3,ci).coordinate,h,_XL_MED)
        for ri,(minute,cc) in enumerate(sorted(md.items()),start=4):
            ws3.cell(ri,1,minute).alignment=ctr; row_total=0
            for ci,cat in enumerate(all_cats,start=2):
                cnt=cc.get(cat,0); ws3.cell(ri,ci,cnt).alignment=ctr
                ws3.cell(ri,ci).border=brd; row_total+=cnt
            ws3.cell(ri,len(all_cats)+2,row_total).font=Font(bold=True)
            ws3.cell(ri,len(all_cats)+2).alignment=ctr
            if ri%2==0:
                for ci in range(1,len(all_cats)+3):
                    ws3.cell(ri,ci).fill=PatternFill("solid",fgColor=_XL_GRAY)
        for c in range(1,len(all_cats)+3):
            ws3.column_dimensions[get_column_letter(c)].width=12

    # ── Sheet 4: Velocidades ──────────────────────────────────────────────────
    ws4=wb.create_sheet("Velocidades"); ws4.sheet_view.showGridLines=False
    ws4.merge_cells("A1:G1"); _xlh(ws4,"A1","REGISTRO DE VELOCIDADES",sz=13)
    for ci,h in enumerate(["#","Timestamp","ID Track","Categoría",
                             "Carril","Velocidad (km/h)","Nivel"],start=1):
        _xlh(ws4,ws4.cell(3,ci).coordinate,h,_XL_MED)
    for ri,ev in enumerate(speed_events[:5000],start=4):
        spd=ev.get("speed_kmh",0)
        lvl="Baja" if spd<40 else "Normal" if spd<80 else "Alta" if spd<120 else "Exceso"
        for ci,val in enumerate([ri-3,ev.get("time",""),ev.get("track_id",""),
                                   ev.get("category",""),ev.get("lane",""),
                                   round(spd,1),lvl],start=1):
            ws4.cell(ri,ci,val).alignment=ctr; ws4.cell(ri,ci).border=brd
        sc="00CC66" if spd<60 else "FFCC00" if spd<90 else "FF4444"
        ws4.cell(ri,6).fill=PatternFill("solid",fgColor=sc)
    for col,w in zip("ABCDEFG",[6,14,10,14,10,18,12]):
        ws4.column_dimensions[col].width=w

    # ── Sheet 5: Registro Detallado ───────────────────────────────────────────
    ws5=wb.create_sheet("Registro Detallado"); ws5.sheet_view.showGridLines=False
    ws5.merge_cells("A1:J1"); _xlh(ws5,"A1","REGISTRO DE EVENTOS INDIVIDUALES",sz=13)
    for ci,h in enumerate(["#","Timestamp","Frame","ID Track","Tipo YOLO",
                             "Categoría","Carril","Dirección","Confianza","Vel km/h"],start=1):
        _xlh(ws5,ws5.cell(3,ci).coordinate,h,_XL_MED)
    for ri,ev in enumerate(events[:10000],start=4):
        for ci,val in enumerate([ri-3,ev.get("time",""),ev.get("frame",""),
                                   ev.get("track_id",""),ev.get("yolo_type",""),
                                   ev.get("category",""),ev.get("lane",""),
                                   ev.get("direction",""),round(ev.get("conf",0),3),
                                   round(ev.get("speed_kmh",0),1)],start=1):
            ws5.cell(ri,ci,val).alignment=ctr; ws5.cell(ri,ci).border=brd
        if ri%2==0:
            for ci in range(1,11):
                ws5.cell(ri,ci).fill=PatternFill("solid",fgColor=_XL_GRAY)
    for ci,w in enumerate([6,14,8,10,14,12,10,14,12,16],start=1):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    for sheet in wb.worksheets: sheet.freeze_panes="A4"
    wb.save(path)
    print(f"  [Excel] -> {path}  ({len(events)} eventos, {len(speed_events)} velocidades)")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 11 — PDF REPORT (ANI expert format, Spanish)
# ══════════════════════════════════════════════════════════════════════════════

def write_pdf(path,summary,lane_counts,events,speed_events,
              system,direction,camera_id,video_source,n_lanes,
              site_name="Peaje Niquía",road_name="Autopista Medellín–Bogotá",
              operator="",notes=""):
    if not HAVE_PDF:
        print("[WARN] reportlab not installed — skipping PDF."); return

    DARK=rl_colors.HexColor("#1F3864"); MED=rl_colors.HexColor("#2E75B6")
    LIGHT=rl_colors.HexColor("#D6E4F7"); MG=rl_colors.HexColor("#CCCCCC")
    LTG=rl_colors.HexColor("#F5F5F5")
    _CAT_PDF={"CAT I":"#C9B3FF","CAT II":"#A8F0C0","CAT III":"#A0D8FF",
              "CAT IV":"#FFE680","CAT V":"#FFB366","CAT VI":"#FF9090","CAT VII":"#FF6060",
              "C1":"#A8F0C0","C2P":"#FFE680","C2G":"#A0D8FF",
              "C3":"#FFB366","C4":"#FF9090","C5":"#FF6060"}

    def ts_base():
        return TableStyle([
            ("BACKGROUND",(0,0),(-1,0),DARK),
            ("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),
            ("GRID",(0,0),(-1,-1),0.4,MG),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[LTG,rl_colors.white]),
        ])

    def HR(): return HRFlowable(width="100%",thickness=1,color=MG,spaceAfter=8)

    doc=SimpleDocTemplate(path,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,
                          topMargin=2.5*cm,bottomMargin=2.5*cm)
    styles=getSampleStyleSheet(); story=[]

    def ps(name,parent="Normal",**kw):
        return ParagraphStyle(name,parent=styles[parent],**kw)

    title_s=ps("t",parent="Title",fontSize=20,spaceAfter=4,textColor=DARK,alignment=TA_CENTER)
    sub_s  =ps("s",parent="Normal",fontSize=11,spaceAfter=3,textColor=MED,alignment=TA_CENTER)
    h1_s   =ps("h1",parent="Heading1",fontSize=13,textColor=DARK,spaceBefore=14,spaceAfter=6)
    h2_s   =ps("h2",parent="Heading2",fontSize=11,textColor=MED,spaceBefore=10,spaceAfter=4)
    body_s =ps("b",parent="Normal",fontSize=9.5,leading=14,spaceAfter=5,alignment=TA_JUSTIFY)
    bul_s  =ps("bl",parent="Normal",fontSize=9.5,leading=13,
                spaceAfter=3,leftIndent=12,firstLineIndent=-8)
    mk_s   =ps("mk",parent="Normal",fontSize=9,fontName="Helvetica-Bold")
    mv_s   =ps("mv",parent="Normal",fontSize=9)

    grand_total=sum(summary.values())
    all_cats=sorted(summary.keys()); lane_nums=sorted(lane_counts.keys())

    # Cover
    story+=[Spacer(1,1.5*cm),
            Paragraph("INFORME TÉCNICO PERICIAL",title_s),
            Paragraph("Conteo y Clasificación Vehicular por Carril",title_s),
            HR(),Paragraph(site_name,sub_s),Paragraph(road_name,sub_s),Spacer(1,0.8*cm)]
    meta_rows=[["Cámara / Fuente:",      camera_id or str(video_source)],
               ["Sistema clasificación:",system],
               ["Dirección de flujo:",   direction],
               ["Número de carriles:",   str(n_lanes)],
               ["Total vehículos:",      str(grand_total)],
               ["Operador:",             operator or "N/D"],
               ["Fecha generación:",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")]]
    if notes: meta_rows.append(["Notas:",notes])
    mt=Table([[Paragraph(k,mk_s),Paragraph(v,mv_s)] for k,v in meta_rows],
             colWidths=[5.5*cm,11*cm])
    mt.setStyle(TableStyle([
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[LIGHT,rl_colors.white]),
        ("GRID",(0,0),(-1,-1),0.4,MG),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story+=[mt,PageBreak()]

    # 1. Category summary
    story.append(Paragraph("1. Resumen Ejecutivo por Categoría Vehicular",h1_s))
    story.append(HR())
    cat_rows=[["Categoría","Descripción","Conteo","% del Total"]]
    for cat,cnt in sorted(summary.items()):
        pct=(cnt/grand_total*100) if grand_total else 0
        cat_rows.append([cat,cat_name(cat,system),str(cnt),f"{pct:.1f}%"])
    cat_rows.append(["TOTAL","",str(grand_total),"100%"])
    ct=Table(cat_rows,colWidths=[2.5*cm,9*cm,2.5*cm,3*cm])
    ts=ts_base()
    for ri,cat in enumerate(sorted(summary.keys()),start=1):
        ts.add("BACKGROUND",(0,ri),(-1,ri),rl_colors.HexColor(_CAT_PDF.get(cat,"#FFFFFF")))
    ts.add("BACKGROUND",(0,-1),(-1,-1),LIGHT)
    ts.add("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")
    ct.setStyle(ts); story+=[ct,Spacer(1,14)]

    # Bar chart
    try:
        vals=[summary.get(c,0) for c in all_cats]
        d=Drawing(16*cm,8*cm); bc=VerticalBarChart()
        bc.x=40; bc.y=30; bc.width=14*cm-60; bc.height=8*cm-60
        bc.data=[vals]; bc.categoryAxis.categoryNames=all_cats
        bc.categoryAxis.labels.boxAnchor="ne"; bc.categoryAxis.labels.angle=30
        bc.categoryAxis.labels.fontSize=7; bc.valueAxis.forceZero=1
        bc.bars[0].fillColor=MED; d.add(bc); story+=[d,Spacer(1,14)]
    except Exception: pass

    # 2. Per-lane
    story.append(Paragraph("2. Conteo por Carril y Categoría",h1_s)); story.append(HR())
    lh=["Carril"]+all_cats+["TOTAL"]
    lr=[lh]
    for lane in lane_nums:
        row=[f"Carril {lane}"]; rt=0
        for cat in all_cats:
            cnt=lane_counts[lane].get(cat,0); row.append(str(cnt)); rt+=cnt
        row.append(str(rt)); lr.append(row)
    tr_row=["TOTAL"]
    for cat in all_cats:
        tr_row.append(str(sum(lane_counts[l].get(cat,0) for l in lane_nums)))
    tr_row.append(str(grand_total)); lr.append(tr_row)
    cw=[2.5*cm]+[14.5*cm/max(len(all_cats)+1,1)]*(len(all_cats)+1)
    lt2=ts_base()
    lt2.add("BACKGROUND",(0,-1),(-1,-1),LIGHT)
    lt2.add("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")
    story+=[Table(lr,colWidths=cw,style=lt2),Spacer(1,14)]

    # 3. 15-min intervals
    if events:
        story.append(Paragraph("3. Resumen de Flujo por Intervalo de 15 Minutos",h1_s))
        story.append(HR())
        iv_data: dict=defaultdict(lambda: defaultdict(int))
        for ev in events:
            ts2=ev.get("time","00:00")
            h2,m2=int(ts2[:2]),int(ts2[3:5]) if len(ts2)>=5 else 0
            b15=f"{h2:02d}:{(m2//15)*15:02d}"
            iv_data[b15][ev.get("category","?")] += 1
        iv_h=["Intervalo"]+all_cats+["TOTAL"]; iv_r=[iv_h]
        for iv,cc in sorted(iv_data.items()):
            row=[iv]; rt=0
            for cat in all_cats:
                cnt=cc.get(cat,0); row.append(str(cnt)); rt+=cnt
            row.append(str(rt)); iv_r.append(row)
        cw2=[2*cm]+[14*cm/max(len(all_cats)+1,1)]*(len(all_cats)+1)
        story+=[Table(iv_r,colWidths=cw2,style=ts_base()),Spacer(1,14)]

    # 4. Methodology
    story.append(PageBreak())
    story.append(Paragraph("4. Metodología",h1_s)); story.append(HR())
    story.append(Paragraph(
        f"El conteo y clasificación vehicular se realizó mediante análisis automatizado de "
        f"video utilizando el modelo de detección de objetos YOLOv8, complementado con el "
        f"algoritmo de seguimiento ByteTrack con filtro de Kalman para mantener identidad "
        f"de vehículos a través de los fotogramas. Los vehículos se clasificaron según el "
        f"sistema <b>{system}</b> a partir de la tipología detectada y las dimensiones relativas "
        f"en el plano de imagen. La detección de carriles se realizó mediante análisis del "
        f"espacio de color HLS para aislar marcas viales blancas y amarillas, seguido de "
        f"detección Canny y transformada de Hough probabilística, con suavizado temporal "
        f"sobre 15 fotogramas para estabilidad. Los carriles discontinuos (líneas de puntos) "
        f"se conectan mediante cierre morfológico antes de la detección.",body_s))

    cfg_rows=[
        ["Detector YOLO:","YOLOv8 — automóvil, motocicleta, bus, camión"],
        ["Tracker:","ByteTrack con filtro de Kalman (7 estados)"],
        ["Líneas de conteo:","3 líneas a 42%, 57% y 72% de la altura del fotograma"],
        ["Lógica anti-duplicado:","Cada track_id se cuenta exactamente una vez"],
        ["Detección de carriles:","HLS + Canny + Hough + suavizado 15 frames"],
        ["Supresión de glare:","Píxeles > 230 de brillo reemplazados por gris neutro"],
    ]
    ctbl=Table(cfg_rows,colWidths=[5*cm,11.5*cm])
    ctbl.setStyle(TableStyle([
        ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[LIGHT,rl_colors.white]),
        ("GRID",(0,0),(-1,-1),0.3,MG),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    story+=[ctbl,Spacer(1,14)]

    # 5. Limitations
    story.append(Paragraph("5. Limitaciones y Consideraciones Técnicas",h1_s))
    story.append(HR())
    for lim in [
        "La clasificación de vehículos pesados (3+ ejes) usa heurísticas de relación de "
        "aspecto y área. Se recomienda revisión manual de una muestra de camiones para "
        "certificación ante la ANI.",
        "El encandilamiento solar en algunos ángulos de cámara puede reducir temporalmente "
        "la precisión de detección de líneas viales. El algoritmo incluye supresión de glare "
        "pero períodos de sol directo pueden causar pérdidas momentáneas.",
        "Vehículos en situación de solapamiento pueden generar un conteo ligeramente inferior "
        "al real durante períodos de alta densidad vehicular.",
        "La estimación de velocidad es orientativa. Para medición legal se requiere "
        "calibración métrica precisa de cada cámara.",
        "Se recomienda validar una muestra de >=200 eventos contra revisión manual para "
        "certificar la precisión ante la ANI.",
    ]:
        story.append(Paragraph(f"• {lim}",bul_s))

    story+=[Spacer(1,20),HR(),
            Paragraph(
                f"Informe generado automáticamente — "
                f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
                f"Sistema: {system} | Cámara: {camera_id or 'N/A'} | Dir: {direction}",
                ps("ft",parent="Normal",fontSize=7,
                   textColor=rl_colors.grey,alignment=TA_CENTER))]
    doc.build(story)
    print(f"  [PDF]   -> {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 12 — HTML REPORT
# ══════════════════════════════════════════════════════════════════════════════

def write_html(path,summary,lane_counts,events,speed_events,
               system,direction,camera_id,video_source,n_lanes,
               site_name="Peaje Niquía",road_name="Autopista Medellín–Bogotá"):
    grand_total=sum(summary.values())
    all_cats=sorted(summary.keys()); lane_nums=sorted(lane_counts.keys())
    ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cat_html=""
    for cat,cnt in sorted(summary.items()):
        pct=(cnt/grand_total*100) if grand_total else 0
        hex_c=CLASSIFICATION_SYSTEMS.get(system,{}).get(cat,{}).get("hex","#ccc")
        cat_html+=(f'<tr style="background:{hex_c}22"><td><b>{cat}</b></td>'
                   f'<td>{cat_name(cat,system)}</td>'
                   f'<td class="num">{cnt}</td><td class="num">{pct:.1f}%</td></tr>\n')
    cat_html+=f'<tr class="total"><td><b>TOTAL</b></td><td></td><td class="num"><b>{grand_total}</b></td><td class="num">100%</td></tr>'

    lane_html="<tr><th>Carril</th>"+"".join(f"<th>{c}</th>" for c in all_cats)+"<th>TOTAL</th></tr>\n"
    for ln in lane_nums:
        row=f"<tr><td><b>Carril {ln}</b></td>"; rt=0
        for cat in all_cats:
            cnt=lane_counts[ln].get(cat,0)
            hx=CLASSIFICATION_SYSTEMS.get(system,{}).get(cat,{}).get("hex","#ccc")
            row+=f'<td class="num" style="background:{hx}33">{cnt}</td>'; rt+=cnt
        row+=f'<td class="num"><b>{rt}</b></td></tr>\n'; lane_html+=row
    lane_html+=f'<tr class="total"><td><b>TOTAL</b></td>'
    for cat in all_cats:
        lane_html+=f'<td class="num"><b>{sum(lane_counts[l].get(cat,0) for l in lane_nums)}</b></td>'
    lane_html+=f'<td class="num"><b>{grand_total}</b></td></tr>'

    ev_html=""
    for ev in events[-200:]:
        ev_html+=(f'<tr><td>{ev.get("time","")}</td><td>{ev.get("frame","")}</td>'
                  f'<td>{ev.get("track_id","")}</td><td><b>{ev.get("category","")}</b></td>'
                  f'<td>{ev.get("lane","")}</td><td>{ev.get("yolo_type","")}</td>'
                  f'<td>{ev.get("speed_kmh",0):.1f}</td></tr>\n')

    html=f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Informe Vehicular — {site_name}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#1a202c}}
.hdr{{background:linear-gradient(135deg,#1F3864,#2E75B6);color:white;padding:32px 40px}}
.hdr h1{{font-size:24px;font-weight:700;margin-bottom:6px}}
.hdr p{{font-size:12px;opacity:.85;margin:2px 0}}
.wrap{{max-width:1100px;margin:24px auto;padding:0 20px}}
.kpi{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:14px;margin-bottom:24px}}
.kpi-c{{background:white;border-radius:10px;padding:18px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
.kpi-c .v{{font-size:32px;font-weight:700;color:#1F3864}}
.kpi-c .l{{font-size:11px;color:#666;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}}
.card{{background:white;border-radius:10px;padding:22px;margin-bottom:22px;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
.card h2{{font-size:15px;font-weight:700;color:#1F3864;margin-bottom:14px;
          padding-bottom:8px;border-bottom:2px solid #D6E4F7}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#1F3864;color:white;padding:8px 10px;text-align:center;font-weight:600}}
td{{padding:7px 10px;border-bottom:1px solid #e8ecf0;text-align:left}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
tr:hover td{{background:#f7f9fc}}
tr.total td{{background:#D6E4F7!important;font-weight:700}}
.ft{{text-align:center;padding:18px;color:#999;font-size:11px}}
</style></head><body>
<div class="hdr">
  <h1>Informe Técnico Pericial — Conteo Vehicular</h1>
  <p>{site_name} | {road_name}</p>
  <p>Sistema: {system} &nbsp;|&nbsp; Dirección: {direction} &nbsp;|&nbsp; Cámara: {camera_id or video_source} &nbsp;|&nbsp; Carriles: {n_lanes}</p>
  <p>Generado: {ts}</p>
</div>
<div class="wrap">
  <div class="kpi">
    <div class="kpi-c"><div class="v">{grand_total}</div><div class="l">Total Vehículos</div></div>
    <div class="kpi-c"><div class="v">{n_lanes}</div><div class="l">Carriles</div></div>
    <div class="kpi-c"><div class="v">{direction}</div><div class="l">Dirección</div></div>
    <div class="kpi-c"><div class="v">{system}</div><div class="l">Clasificación</div></div>
    <div class="kpi-c"><div class="v">{len(events)}</div><div class="l">Eventos Registrados</div></div>
  </div>
  <div class="card"><h2>1. Resumen por Categoría Vehicular</h2>
    <table><tr><th>Categoría</th><th>Descripción</th><th>Conteo</th><th>% del Total</th></tr>
    {cat_html}</table></div>
  <div class="card"><h2>2. Conteo por Carril y Categoría</h2>
    <table>{lane_html}</table></div>
  <div class="card"><h2>3. Últimos 200 Eventos Detectados</h2>
    <table><tr><th>Hora</th><th>Frame</th><th>ID Track</th><th>Categoría</th>
    <th>Carril</th><th>Tipo YOLO</th><th>Vel km/h</th></tr>
    {ev_html}</table></div>
</div>
<div class="ft">Informe generado automáticamente — {ts} — {site_name} — Sistema {system}</div>
</body></html>"""

    with open(path,"w",encoding="utf-8") as f: f.write(html)
    print(f"  [HTML]  -> {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 13 — MAIN PROCESSING LOOP
# ══════════════════════════════════════════════════════════════════════════════

def run(
    source,
    output_path      = None,
    show             = True,
    n_lanes          = 2,
    direction        = "NS",
    system           = "ANI",
    report_prefix    = None,
    camera_id        = None,
    yolo_model       = "yolov8s.pt",
    yolo_conf        = 0.30,
    yolo_iou         = 0.45,
    yolo_imgsz       = 640,
    inference_skip   = 1,
    px_per_meter     = 12.0,
    enable_speed     = True,
    lane_mode        = "auto",
    count_line_fracs = None,
    roi_top_frac     = 0.28,
    roi_left_frac    = 0.44,   # x-fraction where road starts (0=left edge, 0.44=default)
    site_name        = "Peaje Niquía",
    road_name        = "Autopista Medellín–Bogotá",
    operator         = "",
    notes            = "",
    show_heatmap     = False,
    show_wrong_way   = True,
):
    """Main processing entry point. Processes one video and generates all reports."""
    if count_line_fracs is None:
        # 3 counting lines tuned for Niquía angled overhead cameras.
        # Placed at 42%, 57%, 72% of frame height where vehicles are
        # fully visible and lane assignment is most accurate.
        count_line_fracs = [0.42, 0.57, 0.72]

    cats    = CLASSIFICATION_SYSTEMS[system]
    cam_str = camera_id or "CAM-01"

    print(f"\n{'='*65}")
    print(f"  NIQUÍA ENTERPRISE VEHICLE COUNTER")
    print(f"  Site:    {site_name}  |  Road: {road_name}")
    print(f"  System:  {system}     |  Direction: {direction}  |  Lanes: {n_lanes}")
    print(f"  Model:   {yolo_model} |  Conf: {yolo_conf}  |  Skip: {inference_skip}")
    print(f"{'='*65}")

    print("[INFO] Loading YOLO model...")
    model = YOLO(yolo_model)

    cap = cv2.VideoCapture(source)
    if not cap.isOpened():
        raise SystemExit(f"[ERROR] Cannot open: {source}")

    W            = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    H            = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    FPS          = cap.get(cv2.CAP_PROP_FPS) or 25.0
    TOTAL_FRAMES = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    LINES        = [int(H*f) for f in count_line_fracs]

    print(f"[INFO] Resolution: {W}x{H} @ {FPS:.0f} fps  |  Total frames: {TOTAL_FRAMES}")
    print(f"[INFO] Count lines y={LINES}  |  Lane mode: {lane_mode}")

    # Initialise all components
    lane_mgr  = LaneManager(W,H,n_lanes,direction,mode=lane_mode,
                       roi_top_frac=roi_top_frac,roi_left_frac=roi_left_frac)
    tracker   = ByteTracker(max_misses=15,min_hits=2)
    speed_est = SpeedEstimator(FPS,px_per_meter) if enable_speed else None
    density   = DensityAnalyzer(n_lanes,FPS)
    hud       = HUDEngine(direction,system,cam_str,site_name,
                          cfg_show_speed=enable_speed,cfg_show_trails=True,
                          cfg_show_wrong=show_wrong_way,cfg_show_hmap=show_heatmap)

    writer = None
    if output_path:
        fourcc=cv2.VideoWriter_fourcc(*"mp4v")
        writer=cv2.VideoWriter(output_path,fourcc,FPS,(W,H))
        print(f"[INFO] Saving annotated video -> {output_path}")

    cat_counts:   Dict[str,int]           = defaultdict(int)
    lane_counts:  Dict[int,Dict[str,int]] = defaultdict(lambda: defaultdict(int))
    events:       List[dict]              = []
    speed_events: List[dict]              = []
    speed_map:    Dict[int,float]         = {}

    frame_n=0; fps_v=0.0; fps_t=time.time(); fps_c=0
    t0=time.time(); last_pct=-1

    print("[INFO] Processing... Press Q to quit.\n")

    while True:
        ok,frame=cap.read()
        if not ok: break
        frame_n+=1; fps_c+=1

        # Progress
        if TOTAL_FRAMES>0:
            pct=int(frame_n/TOTAL_FRAMES*100)
            if pct%5==0 and pct!=last_pct:
                last_pct=pct
                eta=(time.time()-t0)/frame_n*(TOTAL_FRAMES-frame_n) if frame_n else 0
                print(f"  {pct:3d}%  frame={frame_n}/{TOTAL_FRAMES}  "
                      f"ETA:{eta:.0f}s  veh:{sum(cat_counts.values())}")

        # Update white-line lane detection
        lane_mgr.update(frame)

        # YOLO detection
        dets: List[tuple]=[]
        if frame_n % max(inference_skip,1) == 0:
            res=model(frame,
                      classes=list(YOLO_VEHICLE_CLASSES.keys()),
                      conf=yolo_conf,iou=yolo_iou,
                      imgsz=yolo_imgsz,verbose=False)
            if res and res[0].boxes is not None:
                for box,cid,conf_val in zip(
                    res[0].boxes.xyxy.cpu().numpy(),
                    res[0].boxes.cls.cpu().numpy().astype(int),
                    res[0].boxes.conf.cpu().numpy(),
                ):
                    x1,y1,x2,y2=box.astype(int)
                    bw=x2-x1; bh=y2-y1
                    if bw*bh<900: continue
                    yolo_type=YOLO_VEHICLE_CLASSES.get(cid,"car")
                    cat_key=classify_vehicle(yolo_type,bw,bh,system)
                    dets.append((x1,y1,x2,y2,int(cid),cat_key,float(conf_val)))

        # Track
        active=tracker.step(dets,lane_mgr.get_lane)

        # Density active sets
        lane_active: Dict[int,set]=defaultdict(set)
        for tr in active: lane_active[tr.lane].add(tr.id)
        for ln in range(1,n_lanes+1): density.set_active(ln,lane_active[ln])

        # Speed estimation
        if speed_est:
            for tr in active:
                speed_est.update(tr.id,tr.vx,tr.vy)
                spd=speed_est.get_kmh(tr.id)
                speed_map[tr.id]=spd
                if spd>1.0 and tr.counted: tr.speeds.append(spd)

        # Count crossings
        for tr in active:
            if not tr.counted and len(tr.trail)>=2:
                prev_cy=tr.trail[-2][1]; curr_cy=tr.trail[-1][1]
                for li,ly in enumerate(LINES):
                    crossed=((prev_cy<ly<=curr_cy) or (prev_cy>ly>=curr_cy))
                    if crossed and abs(curr_cy-prev_cy)>=4:
                        tr.counted=True; tr.counted_line=li
                        tr.count_ts=datetime.now().strftime("%H:%M:%S")
                        tr.count_frame=frame_n
                        cat_counts[tr.cat_key]+=1
                        lane_counts[tr.lane][tr.cat_key]+=1
                        density.record(tr.lane)
                        spd=speed_map.get(tr.id,0.0)
                        total=sum(cat_counts.values())
                        yolo_type=YOLO_VEHICLE_CLASSES.get(tr.cls_id,"vehicle")
                        ev={"time":tr.count_ts,"frame":frame_n,
                            "track_id":tr.id,"yolo_type":yolo_type,
                            "category":tr.cat_key,"lane":tr.lane,
                            "direction":direction,
                            "conf":round(tr.conf,3),"speed_kmh":round(spd,1)}
                        events.append(ev)
                        if spd>1.0: speed_events.append(ev)
                        print(f"  [{tr.count_ts}]  {tr.cat_key:<10}  "
                              f"Carril {tr.lane}  id={tr.id:4d}  "
                              f"TOTAL={total}  {spd:.0f}km/h")
                        break

        # Draw HUD + lanes
        density_stats=density.all_stats()
        hud.draw_all(frame,active,cats,dict(cat_counts),lane_counts,
                     density_stats,fps_v,frame_n,speed_map,LINES)
        lane_mgr.draw(frame,lane_counts)

        # FPS counter
        now=time.time()
        if now-fps_t>=1.0:
            fps_v=fps_c/(now-fps_t); fps_c=0; fps_t=now

        if writer: writer.write(frame)
        if show:
            cv2.imshow(f"Vehicle Counter — {site_name}  [Q=salir]",frame)
            if cv2.waitKey(1)&0xFF==ord("q"):
                print("[INFO] Stopped by user."); break

    cap.release()
    if writer: writer.release()
    cv2.destroyAllWindows()

    elapsed=time.time()-t0
    grand_total=sum(cat_counts.values())

    # Console final summary
    print(f"\n{'='*60}")
    print(f"  CONTEO FINAL — {site_name}")
    print(f"  Sistema: {system}  |  Dirección: {direction}")
    print(f"{'='*60}")
    print(f"  Total vehículos  : {grand_total}")
    for cat in sorted(cats.keys()):
        cnt=cat_counts.get(cat,0)
        print(f"    {cat:<10}  {cat_name(cat,system):<40}: {cnt}")
    print(f"\n  Por carril:")
    for lane in sorted(lane_counts.keys()):
        lt=sum(lane_counts[lane].values())
        print(f"    Carril {lane}: {lt} vehículos")
        for cat,cnt in sorted(lane_counts[lane].items()):
            if cnt: print(f"      {cat:<10}: {cnt}")
    print(f"\n  Frames: {frame_n}  |  Tiempo: {elapsed:.1f}s", end="")
    if elapsed>0 and grand_total>0:
        print(f"  |  Veh/min: {grand_total/(elapsed/60):.1f}")
    else:
        print()
    print(f"{'='*60}\n")

    # Generate all reports
    prefix  = report_prefix or "informe_vehicular"
    ts_str  = datetime.now().strftime("%Y%m%d_%H%M%S")
    src_str = str(source)

    csv_path=f"{prefix}_{ts_str}_log.csv"
    if events:
        with open(csv_path,"w",newline="",encoding="utf-8") as f:
            w=csv.DictWriter(f,fieldnames=events[0].keys())
            w.writeheader(); w.writerows(events)
        print(f"  [CSV]   -> {csv_path}")

    json_path=f"{prefix}_{ts_str}_summary.json"
    with open(json_path,"w",encoding="utf-8") as f:
        json.dump({
            "meta":{"system":system,"direction":direction,"n_lanes":n_lanes,
                    "camera_id":cam_str,"video_source":src_str,
                    "site_name":site_name,"road_name":road_name,
                    "generated":datetime.now().isoformat(),
                    "frames":frame_n,"elapsed_s":round(elapsed,1)},
            "total":grand_total,
            "by_category":dict(cat_counts),
            "by_lane":{str(k):dict(v) for k,v in lane_counts.items()},
            "density":density.all_stats(),
        },f,indent=2,ensure_ascii=False)
    print(f"  [JSON]  -> {json_path}")

    excel_path=f"{prefix}_{ts_str}.xlsx"
    write_excel(excel_path,dict(cat_counts),lane_counts,events,speed_events,
                system,direction,cam_str,src_str,n_lanes,site_name,operator)

    pdf_path=f"{prefix}_{ts_str}.pdf"
    write_pdf(pdf_path,dict(cat_counts),lane_counts,events,speed_events,
              system,direction,cam_str,src_str,n_lanes,
              site_name,road_name,operator,notes)

    html_path=f"{prefix}_{ts_str}.html"
    write_html(html_path,dict(cat_counts),lane_counts,events,speed_events,
               system,direction,cam_str,src_str,n_lanes,site_name,road_name)

    return dict(cat_counts), dict(lane_counts)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 14 — BATCH PROCESSOR
# ══════════════════════════════════════════════════════════════════════════════

def run_batch(video_list_file: str, default_system: str="ANI"):
    """
    Process multiple videos from a text file.
    Format (one line per camera):
        video_path,lanes,direction,camera_id
    Example:
        /videos/cam1_NS.mp4,3,NS,CAM1-NS
        /videos/cam2_NS.mp4,3,NS,CAM2-NS
        /videos/cam3_SN.mp4,2,SN,CAM3-SN
        /videos/cam4_SN.mp4,2,SN,CAM4-SN
    """
    with open(video_list_file) as f:
        lines=[l.strip() for l in f if l.strip() and not l.startswith("#")]

    all_results=[]; combined: Dict[str,int]=defaultdict(int)
    for line in lines:
        parts  =line.split(",")
        video  =parts[0].strip()
        lanes_n=int(parts[1].strip()) if len(parts)>1 else 2
        direc  =parts[2].strip()      if len(parts)>2 else "NS"
        cam_id =parts[3].strip()      if len(parts)>3 else None
        prefix =Path(video).stem

        print(f"\n{'#'*65}\n  PROCESSING: {video}")
        print(f"  Camera: {cam_id}  Dir: {direc}  Lanes: {lanes_n}\n{'#'*65}")
        try:
            cats,lanes_r=run(source=video,show=False,n_lanes=lanes_n,
                             direction=direc,system=default_system,
                             report_prefix=prefix,camera_id=cam_id)
            all_results.append({"video":video,"camera":cam_id,
                                 "direction":direc,"counts":cats})
            for cat,cnt in cats.items(): combined[cat]+=cnt
        except Exception as e:
            print(f"[ERROR] Failed {video}: {e}")

    if all_results:
        print(f"\n{'='*65}\n  BATCH CONSOLIDATED ({len(all_results)} cameras)\n{'='*65}")
        for cat,cnt in sorted(combined.items()):
            print(f"  {cat:<12}: {cnt}")
        print(f"  {'TOTAL':<12}: {sum(combined.values())}")
        ts_str=datetime.now().strftime("%Y%m%d_%H%M%S")
        with open(f"batch_consolidated_{ts_str}.json","w",encoding="utf-8") as f:
            json.dump({"cameras":all_results,"combined":dict(combined),
                       "total":sum(combined.values()),
                       "generated":datetime.now().isoformat()},
                      f,indent=2,ensure_ascii=False)
        print(f"  [JSON]  -> batch_consolidated_{ts_str}.json")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ap=argparse.ArgumentParser(
        description="Niquía Enterprise Vehicle Counter — ANI/INVIAS/COL",
        formatter_class=argparse.RawTextHelpFormatter,epilog=__doc__)

    src=ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--video",  help="Path to video file")
    src.add_argument("--camera", type=int,help="Webcam index (usually 0)")
    src.add_argument("--batch",  help="Text file listing videos for batch")

    ap.add_argument("--output",         default=None,  help="Save annotated video")
    ap.add_argument("--lanes",          type=int,default=2,
                    help="Number of lanes (NS=3, SN=2 for Niquía)")
    ap.add_argument("--direction",      choices=["NS","SN"],default="NS",
                    help="NS=Norte->Sur (Medellín) | SN=Sur->Norte (Copacabana)")
    ap.add_argument("--classification", choices=["ANI","INVIAS","COL"],default="ANI",
                    help="Vehicle classification system")
    ap.add_argument("--report",         default="informe_vehicular",
                    help="Base name for output report files")
    ap.add_argument("--camera-id",      default=None,
                    help="Camera ID string for reports (e.g. CAM1-NS)")
    ap.add_argument("--model",          default="yolov8s.pt",
                    help="YOLO model: yolov8n.pt (fastest) | yolov8s.pt | yolov8m.pt (best)")
    ap.add_argument("--conf",           type=float,default=0.30,
                    help="YOLO detection confidence threshold (default 0.30)")
    ap.add_argument("--skip",           type=int,default=1,
                    help="Run YOLO every N frames (1=every frame)")
    ap.add_argument("--px-per-meter",   type=float,default=12.0,
                    help="Pixels per metre for speed calibration")
    ap.add_argument("--lane-mode",      choices=["auto","uniform"],default="auto",
                    help="auto=white-line detection | uniform=equal-width fallback")
    ap.add_argument("--site",           default="Peaje Niquía",help="Site name for reports")
    ap.add_argument("--road",           default="Autopista Medellín–Bogotá",help="Road name")
    ap.add_argument("--operator",       default="",help="Operator name for reports")
    ap.add_argument("--notes",          default="",help="Additional notes for PDF")
    ap.add_argument("--no-preview",     action="store_true",help="Headless mode (no window)")
    ap.add_argument("--no-speed",       action="store_true",help="Disable speed estimation")
    ap.add_argument("--heatmap",        action="store_true",help="Show trajectory heatmap")
    ap.add_argument("--no-wrong-way",   action="store_true",help="Disable wrong-way alert")
    ap.add_argument("--roi-left",        type=float,default=0.44,
                    help="X-fraction where road starts (0.0–1.0, default 0.44). "
                         "Increase to exclude more left-side clutter.")

    args=ap.parse_args()

    if args.batch:
        run_batch(args.batch,default_system=args.classification)
    else:
        source=args.video if args.video else args.camera
        run(
            source        =source,
            output_path   =args.output,
            show          =not args.no_preview,
            n_lanes       =args.lanes,
            direction     =args.direction,
            system        =args.classification,
            report_prefix =args.report,
            camera_id     =args.camera_id,
            yolo_model    =args.model,
            yolo_conf     =args.conf,
            inference_skip=args.skip,
            px_per_meter  =args.px_per_meter,
            enable_speed  =not args.no_speed,
            lane_mode     =args.lane_mode,
            site_name     =args.site,
            road_name     =args.road,
            operator      =args.operator,
            notes         =args.notes,
            show_heatmap  =args.heatmap,
            show_wrong_way=not args.no_wrong_way,
            roi_left_frac =args.roi_left,
        )
